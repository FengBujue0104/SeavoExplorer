import sys
import os
import re
import json
import time
import base64
import socket
import traceback
import ctypes
import shutil
import stat
import subprocess
import tempfile
import http.client
import urllib.error
import urllib.request
from collections import namedtuple

# 尝试导入OpenCV用于视频缩略图生成
try:
    import cv2
    import numpy as np
    from PIL import Image
    HAS_OPENCV = True
except ImportError:
    HAS_OPENCV = False

from PyQt5.QtWidgets import (QApplication, QMainWindow, QTreeView, QTextEdit,
                                QSplitter, QVBoxLayout, QHBoxLayout, QWidget,
                                QLineEdit, QLabel, QPushButton, QFileDialog, QScrollArea,
                                QMessageBox, QTabWidget, QFileSystemModel,
                                QGroupBox, QMenu, QAbstractItemView, QShortcut,
                                QDialog, QGridLayout, QTableWidget, QTableWidgetItem,
                                QHeaderView, QFormLayout,
                                QRadioButton, QButtonGroup, QInputDialog, QSplashScreen,
                                QToolBar, QToolButton, QSizePolicy, QProgressDialog,
                                QCheckBox, QComboBox, QListWidget, QListWidgetItem,
                                QAction)
from PyQt5.QtCore import QDir, Qt, QModelIndex, QThread, pyqtSignal, QRect, QUrl, QMimeData, QTimer, QEvent
from PyQt5.QtGui import QFont, QPixmap, QImage, QIcon, QPainter, QColor, QPen, QKeySequence, QFontDatabase, QIntValidator

_SPLASH_PIXMAP = None

# 项目文件夹默认命名正则：主板 S 前缀 / 子卡 M 前缀 + 3~4 位编号 + 可选 _注释 后缀。
# 实际使用时请通过 self.folder_regex_mb / self.folder_regex_db 获取。
DEFAULT_MB_RE_TEXT = r'^S(\d{3,4})(?:_(.*))?$'
DEFAULT_DB_RE_TEXT = r'^M(\d{3,4})(?:_(.*))?$'
DEFAULT_MB_RE = re.compile(DEFAULT_MB_RE_TEXT)
DEFAULT_DB_RE = re.compile(DEFAULT_DB_RE_TEXT)

def _resolve_regex(state, custom_text, default_re):
    """根据 state 返回 (re.Pattern, is_fallback)。state 为 custom 时尝试编译 custom_text，失败时兜底到默认正则。"""
    if state == "custom":
        try:
            return re.compile(custom_text), False
        except re.error:
            return default_re, True
    return default_re, False

def _is_regex_safe(pattern):
    """检测正则是否有 ReDoS（灾难性回溯）风险。"""
    import time
    try:
        compiled = re.compile(pattern)
    except re.error as e:
        return False, str(e)
    for test_str in ['A' * 50, 'a' * 50, '1' * 50, '_' * 50]:
        start = time.monotonic()
        try:
            compiled.search(test_str)
            if time.monotonic() - start > 0.5:
                return False, 'match too slow'
        except RuntimeError:
            return False, 'recursion limit'
    return True, ''

APP_VERSION = '0.5.0'
GITHUB_REPO_URL = 'https://github.com/FengBujue0104/SeavoExplorer/'
GITHUB_RELEASES_URL = 'https://github.com/FengBujue0104/SeavoExplorer/releases'
GITHUB_LATEST_RELEASE_API = 'https://api.github.com/repos/FengBujue0104/SeavoExplorer/releases/latest'



def _get_app_dir():
    """返回配置、日志等 sidecar 文件所在目录。"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def _decode_zip_name(raw):
    """解码 zip 条目名：旧式 zip 用 cp437 存中文名，依次尝试 gbk、utf-8 还原，都失败则用原值。"""
    for enc in ('gbk', 'utf-8'):
        try:
            return raw.encode('cp437').decode(enc)
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass
    return raw


# Windows 文件属性
FILE_ATTRIBUTE_HIDDEN = 0x02

# 项目版本文件夹默认子目录模板（集中定义，避免对话框与默认设置各写一份）
DEFAULT_STRUCTURE_FOLDERS = ['BOM', 'SCH', '物料', '评审', '信号测试']

# 各类文件预览的截断阈值（产品行为参数，集中可调）
PREVIEW_PDF_PAGES = 3
PREVIEW_EXCEL_MAX_ROWS = 10
PREVIEW_DOCX_PARAGRAPHS = 20
PREVIEW_DOC_LINES = 50
# 视频预览截取位置(百分比),从左到右排列
VIDEO_PREVIEW_POSITIONS = [0.1, 0.3, 0.5, 0.7, 0.9]

# 按扩展名分类的可预览文件类型（preview_file 与 show_full_image 共用，避免两份手动同步）
IMAGE_EXTS = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif', '.webp', '.svg', '.ico', '.jfif')
VIDEO_EXTS = ('.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.m4v', '.webm', '.mpg', '.mpeg', '.3gp')
ARCHIVE_EXTS = ('.zip', '.rar', '.7z')
TEXT_EXTS = (
    '.txt', '.csv', '.log', '.bom', '.drc', '.rep', '.rpt', '.md', '.json', '.xml',
    '.html', '.htm', '.ini', '.cfg',
    '.py', '.spec', '.toml', '.yaml', '.yml',
)

# 无扩展名的文本类 dotfile（os.path.splitext 无法提取扩展名，需按文件名匹配）
TEXT_DOTFILES = {'.gitignore', '.dockerignore', '.editorconfig', '.env.example', '.gitattributes'}

# 可分类开关的预览类别：(设置 key, 中文名)。顺序即「预览设置」对话框中的展示顺序。
# 二进制兜底与加密类型不在此列（前者只读前1000字节，后者只显示提示，均无卡顿风险）。
PREVIEW_CATEGORIES = (
    ('text', '文本'),
    ('pdf', 'PDF'),
    ('image', '图片'),
    ('video', '视频'),
    ('archive', '压缩包'),
    ('excel', '表格'),
    ('word', '文档'),
)

# 解压安全上限。所有内容先写入压缩包同目录的 staging，完整校验后才提交到最终位置。
ARCHIVE_MAX_ENTRIES = 50000
ARCHIVE_MAX_TOTAL_SIZE = 50 * 1024 * 1024 * 1024
ARCHIVE_MAX_DEPTH = 64
ARCHIVE_MAX_PATH_NODES = 200000
ARCHIVE_MAX_PATH_CHARS = 32 * 1024 * 1024
ARCHIVE_MAX_7Z_OUTPUT = 64 * 1024 * 1024
ARCHIVE_MAX_COMPRESSION_RATIO = 1000
ARCHIVE_RATIO_CHECK_MIN_SIZE = 100 * 1024 * 1024
ARCHIVE_DISK_SAFETY_MIN = 512 * 1024 * 1024
ARCHIVE_COPY_CHUNK_SIZE = 1024 * 1024
FILE_ATTRIBUTE_REPARSE_POINT = 0x0400

_WINDOWS_RESERVED_NAMES = {
    'CON', 'PRN', 'AUX', 'NUL',
    *(f'COM{i}' for i in range(1, 10)),
    *(f'LPT{i}' for i in range(1, 10)),
}


class ArchiveSafetyError(Exception):
    """压缩包未通过安全预检或解压后校验。"""


class ArchiveExtractionCanceled(Exception):
    """用户取消了解压。"""


ArchiveManifestEntry = namedtuple(
    'ArchiveManifestEntry',
    ['name', 'parts', 'size', 'is_dir', 'compressed_size', 'source'],
)


def _is_unc_path(path):
    normalized = str(path or '').replace('/', '\\')
    return normalized.startswith('\\\\')


def _powershell_unc_location_args(folder_path):
    """为 UNC 路径生成不含原始路径字符的安全 PowerShell 参数。"""
    path_data = base64.b64encode(folder_path.encode('utf-8')).decode('ascii')
    script = (
        "$p=[Text.Encoding]::UTF8.GetString([Convert]::FromBase64String('"
        + path_data
        + "'));Set-Location -LiteralPath $p"
    )
    encoded_script = base64.b64encode(script.encode('utf-16le')).decode('ascii')
    return ['-NoLogo', '-NoProfile', '-NoExit', '-EncodedCommand', encoded_script]


def _load_strict_recycle_backend():
    """只加载 Win10/11 的现代回收站后端；不可用时安全失败，绝不降级永久删除。"""
    if sys.platform != 'win32':
        raise RuntimeError('回收站功能仅支持 Windows 10/11')
    winver = sys.getwindowsversion()
    if winver.major < 10:
        raise RuntimeError('当前系统不支持强制回收站模式，已取消删除')
    try:
        from send2trash.win.modern import send2trash as modern_send2trash
    except ImportError as e:
        raise RuntimeError(
            '缺少现代回收站组件，请安装 Send2Trash[nativeLib]==2.1.0'
        ) from e
    return modern_send2trash


def _send_path_to_recycle_strict(path, backend=None):
    """将单一路径强制送入回收站；任何异常都由调用方处理，不做删除回退。"""
    recycle = backend or _load_strict_recycle_backend()
    recycle(path)


def _format_os_error(error):
    """生成包含 Windows/errno 错误码的稳定用户提示。"""
    code = getattr(error, 'winerror', None)
    if code is None:
        code = getattr(error, 'errno', None)
    detail = getattr(error, 'strerror', None) or str(error) or '未知系统错误'
    return f'系统错误 {code}: {detail}' if code is not None else f'系统错误: {detail}'


def _normalize_archive_member_path(name):
    """把归档内路径规范为安全的 Windows 相对路径组件元组。"""
    if not isinstance(name, str):
        raise ArchiveSafetyError('压缩包包含非文本路径')
    if '\x00' in name:
        raise ArchiveSafetyError('压缩包路径包含空字符')

    normalized = name.replace('\\', '/')
    if normalized.startswith('/') or normalized.startswith('//'):
        raise ArchiveSafetyError(f'压缩包包含绝对路径: {name}')
    if re.match(r'^[A-Za-z]:', normalized):
        raise ArchiveSafetyError(f'压缩包包含盘符路径: {name}')
    normalized = normalized.rstrip('/')
    if not normalized:
        raise ArchiveSafetyError('压缩包包含空路径')

    parts = normalized.split('/')
    if len(parts) > ARCHIVE_MAX_DEPTH:
        raise ArchiveSafetyError(f'压缩包路径层级超过 {ARCHIVE_MAX_DEPTH}: {name}')

    for part in parts:
        if part in ('', '.', '..'):
            raise ArchiveSafetyError(f'压缩包包含非法路径段: {name}')
        if len(part) > 255:
            raise ArchiveSafetyError(f'压缩包路径段过长: {name}')
        if part.endswith((' ', '.')):
            raise ArchiveSafetyError(f'压缩包路径含尾随空格或点: {name}')
        if any(ord(ch) < 32 for ch in part) or any(ch in '<>:"|?*' for ch in part):
            raise ArchiveSafetyError(f'压缩包路径包含 Windows 非法字符: {name}')
        device_name = part.split('.', 1)[0].upper()
        if device_name in _WINDOWS_RESERVED_NAMES:
            raise ArchiveSafetyError(f'压缩包路径使用 Windows 保留名称: {name}')
    return tuple(parts)


def _archive_path_key(parts):
    return tuple(part.casefold() for part in parts)


def _build_archive_manifest(raw_entries, archive_size):
    """校验归档条目并生成统一 manifest。raw entry 为字典。"""
    if len(raw_entries) > ARCHIVE_MAX_ENTRIES:
        raise ArchiveSafetyError(f'压缩包条目超过 {ARCHIVE_MAX_ENTRIES} 个')

    manifest = []
    path_types = {}
    path_chars = 0
    total_size = 0

    def register_path(key, path_type):
        nonlocal path_chars
        existing_type = path_types.get(key)
        if existing_type is not None:
            return existing_type
        if len(path_types) >= ARCHIVE_MAX_PATH_NODES:
            raise ArchiveSafetyError(
                f'压缩包唯一路径节点超过 {ARCHIVE_MAX_PATH_NODES} 个'
            )
        path_chars += sum(len(part) for part in key) + max(0, len(key) - 1)
        if path_chars > ARCHIVE_MAX_PATH_CHARS:
            raise ArchiveSafetyError(
                f'压缩包路径文本总量超过 {ARCHIVE_MAX_PATH_CHARS // (1024 ** 2)} MiB'
            )
        path_types[key] = path_type
        return None

    for raw in raw_entries:
        if raw.get('encrypted'):
            raise ArchiveSafetyError(f"暂不支持加密条目: {raw.get('name', '')}")
        if raw.get('is_link') or raw.get('is_special'):
            raise ArchiveSafetyError(f"压缩包包含链接或特殊文件: {raw.get('name', '')}")

        name = raw.get('name', '')
        parts = _normalize_archive_member_path(name)
        is_dir = bool(raw.get('is_dir'))
        try:
            size = int(raw.get('size') or 0)
            compressed_size = int(raw.get('compressed_size') or 0)
        except (TypeError, ValueError) as e:
            raise ArchiveSafetyError(f'压缩包条目大小无效: {name}') from e
        if size < 0 or compressed_size < 0:
            raise ArchiveSafetyError(f'压缩包条目大小无效: {name}')

        key = _archive_path_key(parts)
        for depth in range(1, len(parts)):
            parent_key = _archive_path_key(parts[:depth])
            if path_types.get(parent_key) == 'file':
                raise ArchiveSafetyError(f'压缩包中文件与子路径冲突: {name}')
            register_path(parent_key, 'dir')

        current_type = 'dir' if is_dir else 'file'
        existing_type = register_path(key, current_type)
        if existing_type is not None:
            if existing_type == 'dir' and current_type == 'dir':
                continue
            raise ArchiveSafetyError(f'压缩包包含重复或大小写冲突路径: {name}')

        if not is_dir:
            total_size += size
            if total_size > ARCHIVE_MAX_TOTAL_SIZE:
                raise ArchiveSafetyError(
                    f'压缩包展开总大小超过 {ARCHIVE_MAX_TOTAL_SIZE // (1024 ** 3)} GiB'
                )
            if (
                size >= ARCHIVE_RATIO_CHECK_MIN_SIZE
                and compressed_size > 0
                and size / compressed_size > ARCHIVE_MAX_COMPRESSION_RATIO
            ):
                raise ArchiveSafetyError(f'压缩比异常，疑似压缩炸弹: {name}')

        manifest.append(ArchiveManifestEntry(
            name=name,
            parts=parts,
            size=size,
            is_dir=is_dir,
            compressed_size=compressed_size,
            source=raw.get('source'),
        ))

    if not manifest:
        raise ArchiveSafetyError('压缩包为空')
    if (
        total_size >= ARCHIVE_RATIO_CHECK_MIN_SIZE
        and archive_size > 0
        and total_size / archive_size > ARCHIVE_MAX_COMPRESSION_RATIO
    ):
        raise ArchiveSafetyError('压缩包总体压缩比异常，疑似压缩炸弹')
    return manifest, total_size


def _zip_raw_entries(zf):
    entries = []
    for info in zf.infolist():
        mode = (info.external_attr >> 16) & 0xFFFF
        file_type = stat.S_IFMT(mode)
        is_link = stat.S_ISLNK(mode)
        is_special = bool(file_type and not (stat.S_ISREG(mode) or stat.S_ISDIR(mode) or is_link))
        entries.append({
            'name': _decode_zip_name(info.filename),
            'size': info.file_size,
            'compressed_size': info.compress_size,
            'is_dir': info.is_dir(),
            'encrypted': bool(info.flag_bits & 0x1),
            'is_link': is_link,
            'is_special': is_special,
            'source': info,
        })
    return entries


def _parse_7z_slt_output(output):
    """解析 7z -slt 输出，并确保 EOF 前最后一个条目不会丢失。"""
    entries = []
    record = {}
    in_entries = False

    def flush_record():
        nonlocal record
        if not in_entries or 'Path' not in record:
            record = {}
            return
        if len(entries) >= ARCHIVE_MAX_ENTRIES:
            raise ArchiveSafetyError(f'压缩包条目超过 {ARCHIVE_MAX_ENTRIES} 个')
        attrs = record.get('Attributes', '')
        entries.append({
            'name': record['Path'],
            'size': record.get('Size', 0),
            'compressed_size': record.get('Packed Size', 0),
            'is_dir': record.get('Folder') == '+' or 'D' in attrs,
            'encrypted': record.get('Encrypted') == '+',
            'is_link': bool(record.get('Symbolic Link') or record.get('Hard Link')),
            'is_special': record.get('Alternate Stream') == '+' or record.get('Anti') == '+',
            'source': dict(record),
        })
        record = {}

    for raw_line in output.splitlines():
        line = raw_line.strip()
        if line == '----------':
            flush_record()
            in_entries = True
            continue
        if not line:
            flush_record()
            continue
        if ' = ' not in line:
            continue
        key, value = line.split(' = ', 1)
        if key in ('Size', 'Packed Size'):
            try:
                value = int(value)
            except ValueError:
                value = 0
        record[key] = value
    flush_record()
    return entries


def _archive_identity(path):
    st = os.stat(path)
    return st.st_size, st.st_mtime_ns


def _ensure_archive_disk_space(parent_dir, total_size):
    reserve = max(ARCHIVE_DISK_SAFETY_MIN, total_size // 10)
    free = shutil.disk_usage(parent_dir).free
    if free < total_size + reserve:
        required = total_size + reserve
        raise ArchiveSafetyError(
            f'磁盘空间不足，至少需要 {required / (1024 ** 3):.2f} GiB 可用空间'
        )


def _safe_stage_target(stage_dir, parts):
    target = os.path.join(stage_dir, *parts)
    stage_real = os.path.realpath(stage_dir)
    target_real = os.path.realpath(target)
    try:
        if os.path.commonpath([stage_real, target_real]) != stage_real:
            raise ArchiveSafetyError('压缩包目标路径越过临时解压目录')
    except ValueError as e:
        raise ArchiveSafetyError('压缩包目标路径位于其他磁盘') from e
    return target


def _extract_zip_to_stage(zf, manifest, stage_dir, is_canceled, progress_callback):
    total_written = 0
    for index, entry in enumerate(manifest, 1):
        if is_canceled():
            raise ArchiveExtractionCanceled()
        target_path = _safe_stage_target(stage_dir, entry.parts)
        if entry.is_dir:
            os.makedirs(target_path, exist_ok=True)
            continue
        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        file_written = 0
        with zf.open(entry.source) as source, open(target_path, 'xb') as target:
            while True:
                if is_canceled():
                    raise ArchiveExtractionCanceled()
                chunk = source.read(ARCHIVE_COPY_CHUNK_SIZE)
                if not chunk:
                    break
                file_written += len(chunk)
                total_written += len(chunk)
                if file_written > entry.size or total_written > ARCHIVE_MAX_TOTAL_SIZE:
                    raise ArchiveSafetyError(f'条目展开大小与声明不符: {entry.name}')
                target.write(chunk)
        if file_written != entry.size:
            raise ArchiveSafetyError(f'条目展开大小与声明不符: {entry.name}')
        if progress_callback and (index == len(manifest) or index % 100 == 0):
            progress_callback(f'正在解压：{index}/{len(manifest)}')


def _validate_staged_tree(stage_dir, manifest, is_canceled=lambda: False):
    expected_files = {}
    expected_dirs = set()
    for entry in manifest:
        if is_canceled():
            raise ArchiveExtractionCanceled()
        key = _archive_path_key(entry.parts)
        for depth in range(1, len(entry.parts)):
            expected_dirs.add(_archive_path_key(entry.parts[:depth]))
        if entry.is_dir:
            expected_dirs.add(key)
        else:
            expected_files[key] = entry.size

    actual_files = {}
    actual_dirs = set()
    stage_real = os.path.realpath(stage_dir)
    for root, dirs, files in os.walk(stage_dir, followlinks=False):
        if is_canceled():
            raise ArchiveExtractionCanceled()
        for names, is_dir in ((dirs, True), (files, False)):
            for name in names:
                if is_canceled():
                    raise ArchiveExtractionCanceled()
                path = os.path.join(root, name)
                st = os.lstat(path)
                if os.path.islink(path) or getattr(st, 'st_file_attributes', 0) & FILE_ATTRIBUTE_REPARSE_POINT:
                    raise ArchiveSafetyError(f'解压结果包含链接或 reparse point: {name}')
                real_path = os.path.realpath(path)
                if os.path.commonpath([stage_real, real_path]) != stage_real:
                    raise ArchiveSafetyError(f'解压结果越过临时目录: {name}')
                rel = os.path.relpath(path, stage_dir).replace('\\', '/')
                parts = _normalize_archive_member_path(rel)
                key = _archive_path_key(parts)
                if is_dir:
                    actual_dirs.add(key)
                else:
                    actual_files[key] = st.st_size

    if actual_files != expected_files:
        raise ArchiveSafetyError('解压后的文件集合或大小与压缩包清单不一致')
    if actual_dirs != expected_dirs:
        raise ArchiveSafetyError('解压后的目录集合与压缩包清单不一致')


def _unique_destination_candidates(parent_dir, name, is_dir):
    yield os.path.join(parent_dir, name)
    if is_dir:
        stem, ext = name, ''
    else:
        stem, ext = os.path.splitext(name)
    counter = 1
    while counter <= 100000:
        yield os.path.join(parent_dir, f'{stem} ({counter}){ext}')
        counter += 1
    raise ArchiveSafetyError('无法生成唯一的解压目标名称')


def _rename_without_overwrite(source, parent_dir, name, is_dir):
    for candidate in _unique_destination_candidates(parent_dir, name, is_dir):
        if os.path.lexists(candidate):
            continue
        try:
            os.rename(source, candidate)
            return candidate
        except FileExistsError:
            continue
        except OSError as e:
            if getattr(e, 'winerror', None) in (80, 183):
                continue
            raise
    raise ArchiveSafetyError('无法提交解压结果')


def _commit_staged_extraction(stage_dir, parent_dir, archive_name):
    top_items = os.listdir(stage_dir)
    if not top_items:
        raise ArchiveSafetyError('压缩包解压后没有内容')
    if len(top_items) == 1:
        name = top_items[0]
        source = os.path.join(stage_dir, name)
        destination = _rename_without_overwrite(source, parent_dir, name, os.path.isdir(source))
        try:
            os.rmdir(stage_dir)
        except OSError:
            shutil.rmtree(stage_dir, ignore_errors=True)
        return destination, not os.path.exists(stage_dir)
    destination = _rename_without_overwrite(stage_dir, parent_dir, archive_name or '解压内容', True)
    return destination, True


def _subprocess_window_options():
    options = {}
    if sys.platform == 'win32':
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0
        options['startupinfo'] = startupinfo
        options['creationflags'] = subprocess.CREATE_NO_WINDOW
    return options


def _run_7z_process(
    cmd,
    timeout,
    is_canceled=lambda: False,
    process_callback=None,
    max_output_bytes=ARCHIVE_MAX_7Z_OUTPUT,
):
    """运行可取消的 7-Zip 命令，控制台输出固定为 UTF-8。"""
    with tempfile.TemporaryFile() as stdout_file, tempfile.TemporaryFile() as stderr_file:
        process = subprocess.Popen(
            cmd,
            stdin=subprocess.DEVNULL,
            stdout=stdout_file,
            stderr=stderr_file,
            **_subprocess_window_options(),
        )

        def stop_process():
            if process.poll() is not None:
                return
            try:
                process.terminate()
            except OSError:
                pass
            try:
                process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                try:
                    process.kill()
                except OSError:
                    pass
                process.wait()

        deadline = time.monotonic() + timeout
        try:
            if process_callback:
                process_callback(process)
            while process.poll() is None:
                if is_canceled():
                    stop_process()
                    raise ArchiveExtractionCanceled()
                output_size = (
                    os.fstat(stdout_file.fileno()).st_size
                    + os.fstat(stderr_file.fileno()).st_size
                )
                if output_size > max_output_bytes:
                    stop_process()
                    raise ArchiveSafetyError(
                        f'7-Zip 输出超过 {max_output_bytes // (1024 ** 2)} MiB 安全上限'
                    )
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    stop_process()
                    raise ArchiveSafetyError('7-Zip 操作超时')
                try:
                    process.wait(timeout=min(0.05, remaining))
                except subprocess.TimeoutExpired:
                    continue

            output_size = (
                os.fstat(stdout_file.fileno()).st_size
                + os.fstat(stderr_file.fileno()).st_size
            )
            if output_size > max_output_bytes:
                raise ArchiveSafetyError(
                    f'7-Zip 输出超过 {max_output_bytes // (1024 ** 2)} MiB 安全上限'
                )
            if is_canceled():
                raise ArchiveExtractionCanceled()
            stdout_file.seek(0)
            stderr_file.seek(0)
            stdout = stdout_file.read()
            stderr = stderr_file.read()
        except BaseException:
            stop_process()
            raise
        finally:
            if process_callback:
                process_callback(None)

    try:
        stdout_text = stdout.decode('utf-8')
        stderr_text = stderr.decode('utf-8')
    except UnicodeDecodeError as e:
        raise ArchiveSafetyError('7-Zip 输出不是有效 UTF-8，已停止处理') from e
    if is_canceled():
        raise ArchiveExtractionCanceled()
    if process.returncode != 0:
        detail = (stderr_text or stdout_text).strip()
        if len(detail) > 2000:
            detail = detail[-2000:]
        raise ArchiveSafetyError(f'7-Zip 操作失败: {detail or process.returncode}')
    return stdout_text


def _inspect_7z_archive(
    sevenzip,
    archive_path,
    is_canceled=lambda: False,
    process_callback=None,
):
    cmd = [sevenzip, 'l', '-slt', '-sccUTF-8', '-p', '--', archive_path]
    output = _run_7z_process(cmd, 120, is_canceled, process_callback)
    return _parse_7z_slt_output(output)


def _extract_7z_to_stage(
    sevenzip,
    archive_path,
    stage_dir,
    is_canceled=lambda: False,
    process_callback=None,
):
    cmd = [
        sevenzip, 'x', '-aos', '-p', '-bd', '-bb0', '-sccUTF-8',
        f'-o{stage_dir}', '--', archive_path,
    ]
    _run_7z_process(cmd, 300, is_canceled, process_callback)


def _transactional_extract_archive(
    archive_path,
    sevenzip=None,
    is_canceled=lambda: False,
    progress_callback=None,
    process_callback=None,
):
    """安全解压到 staging，完整校验后以不覆盖方式提交，返回最终目标路径。"""
    archive_path = os.path.abspath(archive_path)
    if not os.path.isfile(archive_path):
        raise ArchiveSafetyError('压缩包不存在')
    ext = os.path.splitext(archive_path)[1].lower()
    if ext not in ARCHIVE_EXTS:
        raise ArchiveSafetyError(f'不支持的压缩包类型: {ext}')
    if ext != '.zip' and not sevenzip:
        raise ArchiveSafetyError('未找到7-Zip，请在设置中指定7z.exe路径或安装7-Zip')

    parent_dir = os.path.dirname(archive_path)
    archive_name = os.path.splitext(os.path.basename(archive_path))[0]
    identity_before = _archive_identity(archive_path)
    manifest = None
    total_size = 0
    zip_file = None

    if progress_callback:
        progress_callback('正在检查压缩包安全性...')
    try:
        if ext == '.zip':
            import zipfile
            zip_file = zipfile.ZipFile(archive_path, 'r')
            manifest, total_size = _build_archive_manifest(
                _zip_raw_entries(zip_file),
                identity_before[0],
            )
        else:
            raw_entries = _inspect_7z_archive(
                sevenzip,
                archive_path,
                is_canceled,
                process_callback,
            )
            manifest, total_size = _build_archive_manifest(raw_entries, identity_before[0])

        if _archive_identity(archive_path) != identity_before:
            raise ArchiveSafetyError('压缩包在检查期间发生变化，已停止解压')
        _ensure_archive_disk_space(parent_dir, total_size)
        if is_canceled():
            raise ArchiveExtractionCanceled()

        stage_dir = tempfile.mkdtemp(prefix='.seavo-extract-', dir=parent_dir)
        stage_consumed = False
        try:
            if progress_callback:
                progress_callback('正在解压到安全临时目录...')
            if ext == '.zip':
                _extract_zip_to_stage(
                    zip_file,
                    manifest,
                    stage_dir,
                    is_canceled,
                    progress_callback,
                )
            else:
                _extract_7z_to_stage(
                    sevenzip,
                    archive_path,
                    stage_dir,
                    is_canceled,
                    process_callback,
                )
            if _archive_identity(archive_path) != identity_before:
                raise ArchiveSafetyError('压缩包在解压期间发生变化，已丢弃临时结果')
            if is_canceled():
                raise ArchiveExtractionCanceled()
            if progress_callback:
                progress_callback('正在校验并提交解压结果...')
            _validate_staged_tree(stage_dir, manifest, is_canceled)
            if is_canceled():
                raise ArchiveExtractionCanceled()
            destination, stage_consumed = _commit_staged_extraction(
                stage_dir,
                parent_dir,
                archive_name,
            )
            return destination
        finally:
            if not stage_consumed and os.path.isdir(stage_dir):
                shutil.rmtree(stage_dir, ignore_errors=True)
    finally:
        if zip_file is not None:
            zip_file.close()


def _first_available_font(candidates):
    """返回候选列表中系统已安装的第一个字体名，都没有则返回空串。"""
    try:
        families = set(QFontDatabase().families())
    except Exception:
        return ''
    for name in candidates:
        if name in families:
            return name
    return ''


# UI 字体候选：微软雅黑系列优先，其次思源/苹方，最后通用无衬线
_UI_FONT_CANDIDATES = [
    'Microsoft YaHei UI', 'Microsoft YaHei', '微软雅黑',
    'Source Han Sans SC', 'Noto Sans CJK SC', 'PingFang SC',
    'Segoe UI', 'Arial',
]
# 等宽字体候选：用于代码/文本/压缩包等预览
_MONO_FONT_CANDIDATES = ['Cascadia Mono', 'Consolas', 'Source Code Pro', 'Courier New']


def get_mono_font(size=10):
    """获取用于预览区的等宽字体。"""
    name = _first_available_font(_MONO_FONT_CANDIDATES) or 'Courier New'
    return QFont(name, size)


def apply_app_font(app):
    """为整个应用设置清晰的全局界面字体。"""
    name = _first_available_font(_UI_FONT_CANDIDATES)
    if not name:
        return
    font = QFont(name, 10)
    try:
        font.setStyleStrategy(QFont.PreferAntialias)
    except Exception:
        pass
    app.setFont(font)


def get_splash_pixmap():
    """预渲染启动画面，首次调用后缓存"""
    global _SPLASH_PIXMAP
    if _SPLASH_PIXMAP is not None:
        return _SPLASH_PIXMAP
    splash_pix = QPixmap(400, 300)
    splash_pix.fill(Qt.black)
    brand_color = QColor(233, 74, 22)
    painter = QPainter(splash_pix)
    painter.setFont(QFont("Arial", 36, QFont.Bold))
    painter.setPen(brand_color)
    painter.drawText(QRect(0, 100, 400, 50), Qt.AlignCenter, "SEAVO")
    painter.setPen(QPen(brand_color, 2, Qt.DashLine))
    painter.drawRoundedRect(50, 80, 300, 80, 10, 10)
    painter.end()
    _SPLASH_PIXMAP = splash_pix
    return _SPLASH_PIXMAP

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import olefile
except ImportError:
    olefile = None

FolderInfo = namedtuple('FolderInfo', ['sort_key', 'path', 'number', 'comment', 'source'])

class ZoomableImageLabel(QLabel):
    """支持滚轮缩放 + 按钮缩放 + 鼠标拖拽平移的图片显示控件。"""
    ZOOM_MIN = 0.1
    ZOOM_MAX = 10.0
    ZOOM_STEP = 1.25  # 每次滚轮/按钮的缩放倍率

    def __init__(self, parent=None):
        super().__init__(parent)
        self._original_pixmap = None
        self._zoom = 1.0
        self._dragging = False
        self._drag_start_pos = None
        self._drag_start_scroll = None
        self.setAlignment(Qt.AlignCenter)
        self.setCursor(Qt.OpenHandCursor)  # 默认手型提示可拖拽

    def setPixmap(self, pixmap):  # 覆盖:记录原始图并应用当前缩放
        self._original_pixmap = pixmap
        self._zoom = 1.0
        super().setPixmap(pixmap)

    def set_zoom_callback(self, cb):
        """设置缩放变化回调,签名: cb()"""
        self._zoom_cb = cb

    def zoom_in(self):
        self._apply_zoom(self._zoom * self.ZOOM_STEP)

    def zoom_out(self):
        self._apply_zoom(self._zoom / self.ZOOM_STEP)

    def zoom_reset(self):
        self._apply_zoom(1.0)

    def _apply_zoom(self, factor):
        if self._original_pixmap is None:
            return
        factor = max(self.ZOOM_MIN, min(self.ZOOM_MAX, factor))
        self._zoom = factor
        if factor == 1.0:
            super().setPixmap(self._original_pixmap)
        else:
            new_size = self._original_pixmap.size() * factor
            scaled = self._original_pixmap.scaled(new_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            super().setPixmap(scaled)
        # 通知外部更新缩放百分比显示
        if hasattr(self, '_zoom_cb') and self._zoom_cb:
            self._zoom_cb()

    def wheelEvent(self, event):  # 滚轮缩放
        delta = event.angleDelta().y()
        if delta > 0:
            self.zoom_in()
        elif delta < 0:
            self.zoom_out()
        event.accept()

    # ---- 鼠标拖拽平移 ----
    def _scroll_area(self):
        """获取父级 QScrollView,用于滚动。"""
        p = self.parent()
        while p is not None and not isinstance(p, QScrollArea):
            p = p.parent()
        return p

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._original_pixmap is not None:
            self._dragging = True
            self._drag_start_pos = event.pos()
            sa = self._scroll_area()
            if sa:
                self._drag_start_scroll = (sa.horizontalScrollBar().value(), sa.verticalScrollBar().value())
            self.setCursor(Qt.ClosedHandCursor)
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging and self._drag_start_pos is not None:
            delta = event.pos() - self._drag_start_pos
            sa = self._scroll_area()
            if sa and self._drag_start_scroll is not None:
                sa.horizontalScrollBar().setValue(self._drag_start_scroll[0] - delta.x())
                sa.verticalScrollBar().setValue(self._drag_start_scroll[1] - delta.y())
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._dragging = False
            self.setCursor(Qt.OpenHandCursor)
        super().mouseReleaseEvent(event)

    def get_zoom_text(self):
        return f'{int(self._zoom * 100)}%'


class OpenWithDialog(QDialog):
    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.result = None
        self.setWindowTitle('打开方式')
        self.setGeometry(300, 300, 400, 150)
        
        layout = QVBoxLayout()
        label = QLabel(f'选择打开文件的方式：\n\n{os.path.basename(self.file_path)}')
        label.setWordWrap(True)
        layout.addWidget(label)
        
        button_layout = QHBoxLayout()
        direct_open_btn = QPushButton('直接打开')
        direct_open_btn.clicked.connect(self.direct_open)
        explorer_btn = QPushButton('在资源管理器中打开')
        explorer_btn.clicked.connect(self.explorer_open)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(direct_open_btn)
        button_layout.addWidget(explorer_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
    def direct_open(self):
        self.result = 'direct'
        self.accept()
        
    def explorer_open(self):
        self.result = 'explorer'
        self.accept()

class CommentEditDialog(QDialog):
    """注释编辑对话框"""
    def __init__(self, title, current_comment, parent=None):
        super().__init__(parent)
        self.current_comment = current_comment
        self.new_comment = current_comment
        self.setWindowTitle(title)
        self.initUI()
        
    def initUI(self):
        self.setGeometry(300, 300, 400, 150)
        
        layout = QVBoxLayout()
        
        self.comment_edit = QLineEdit(self.current_comment)
        self.comment_edit.setPlaceholderText('请输入注释内容')
        self.comment_edit.setMaxLength(200)  # 限制最大长度
        # 按Enter键保存
        self.comment_edit.returnPressed.connect(self.save)
        
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton('保存')
        save_btn.clicked.connect(self.save)
        
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addWidget(QLabel('注释内容：'))
        layout.addWidget(self.comment_edit)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # 设置焦点到输入框
        self.comment_edit.setFocus()
        
    def save(self):
        self.new_comment = self.comment_edit.text().strip()
        self.accept()
        
    def get_comment(self):
        return self.new_comment

class RenameDialog(QDialog):
    """重命名对话框，将文件名和扩展名分开编辑"""
    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.old_name = os.path.basename(file_path)
        self.is_file = os.path.isfile(file_path)
        
        if self.is_file:
            self.old_base_name, self.old_ext = os.path.splitext(self.old_name)
        else:
            self.old_base_name = self.old_name
            self.old_ext = ''
        
        self.setWindowTitle('重命名')
        self.initUI()
        
    def initUI(self):
        self.setGeometry(300, 300, 400, 150)
        
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel(f'原名称：{self.old_name}'))
        
        self.name_edit = QLineEdit(self.old_base_name)
        self.name_edit.setPlaceholderText('文件名')
        self.name_edit.returnPressed.connect(self.save)
        
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel('名称：'))
        name_layout.addWidget(self.name_edit)
        
        if self.is_file:
            self.ext_edit = QLineEdit(self.old_ext)
            self.ext_edit.setPlaceholderText('扩展名')
            self.ext_edit.setFixedWidth(80)
            name_layout.addWidget(QLabel('扩展名：'))
            name_layout.addWidget(self.ext_edit)
        else:
            self.ext_edit = None
        
        layout.addLayout(name_layout)
        
        button_layout = QHBoxLayout()
        save_btn = QPushButton('确定')
        save_btn.clicked.connect(self.save)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        self.name_edit.setFocus()
        self.name_edit.selectAll()
        
    def save(self):
        new_base = self.name_edit.text().strip()
        if not new_base:
            QMessageBox.warning(self, '警告', '名称不能为空')
            return
        
        if self.ext_edit:
            new_ext = self.ext_edit.text().strip()
            if not new_ext.startswith('.'):
                new_ext = '.' + new_ext if new_ext else ''
            self.new_name = new_base + new_ext
        else:
            self.new_name = new_base
        
        if self.new_name == self.old_name:
            self.reject()
            return
            
        self.accept()
        
    def get_new_name(self):
        return self.new_name

class UpdateDownloadThread(QThread):
    """更新文件下载线程，支持临时文件、重试与断点续传。"""
    progress_changed = pyqtSignal(int, int, float, int, bool, int)
    status_changed = pyqtSignal(str)
    download_completed = pyqtSignal(str)
    download_failed = pyqtSignal(str)
    download_canceled = pyqtSignal(str, int)

    CHUNK_SIZE = 1024 * 1024
    DOWNLOAD_TIMEOUT = 90
    MAX_RETRIES = 3

    def __init__(self, url, save_path, expected_size=0, parent=None):
        super().__init__(parent)
        self.url = url
        self.save_path = save_path
        self.expected_size = int(expected_size or 0)
        self.part_path = save_path + '.part'

    def _make_request(self, start=0):
        headers = {'User-Agent': f'SeavoExplorer/{APP_VERSION}'}
        if start > 0:
            headers['Range'] = f'bytes={start}-'
        return urllib.request.Request(self.url, headers=headers)

    def _sleep_with_cancel(self, seconds):
        end_time = time.monotonic() + seconds
        while time.monotonic() < end_time:
            if self.isInterruptionRequested():
                return False
            self.msleep(100)
        return True

    def _partial_size(self):
        try:
            if os.path.exists(self.part_path):
                return os.path.getsize(self.part_path)
        except OSError:
            pass
        return 0

    def _reset_partial(self):
        """删除下载临时文件；成功或文件不存在返回 None，失败返回具体异常。"""
        try:
            os.remove(self.part_path)
        except FileNotFoundError:
            return None
        except OSError as e:
            return e
        return None

    def _download_once(self, attempt):
        existing = self._partial_size()
        if self.expected_size and existing >= self.expected_size:
            existing = 0
            self._reset_partial()
        request = self._make_request(existing)
        with urllib.request.urlopen(request, timeout=self.DOWNLOAD_TIMEOUT) as response:
            code = getattr(response, 'status', response.getcode())
            resumed = existing > 0 and code == 206
            if existing > 0 and code != 206:
                self.status_changed.emit('服务器未接受断点续传，正在从头重新下载...')
                existing = 0
                resumed = False
                self._reset_partial()
            total = self.expected_size
            content_length = int(response.headers.get('Content-Length') or 0)
            if not total:
                total = existing + content_length if content_length else 0
            mode = 'ab' if resumed else 'wb'
            downloaded = existing if resumed else 0
            start_time = time.monotonic()
            last_emit = 0
            os.makedirs(os.path.dirname(self.save_path) or '.', exist_ok=True)
            self.progress_changed.emit(downloaded, total, 0.0, -1, resumed, attempt)
            with open(self.part_path, mode) as f:
                while True:
                    if self.isInterruptionRequested():
                        f.flush()
                        try:
                            os.fsync(f.fileno())
                        except OSError:
                            pass
                        self.download_canceled.emit(self.part_path, downloaded)
                        return False
                    chunk = response.read(self.CHUNK_SIZE)
                    if not chunk:
                        break
                    f.write(chunk)
                    downloaded += len(chunk)
                    now = time.monotonic()
                    if now - last_emit >= 0.25:
                        elapsed = max(now - start_time, 0.001)
                        speed = max(downloaded - existing, 0) / elapsed
                        eta = int((total - downloaded) / speed) if total > downloaded and speed > 0 else -1
                        self.progress_changed.emit(downloaded, total, speed, eta, resumed, attempt)
                        last_emit = now
                f.flush()
                try:
                    os.fsync(f.fileno())
                except OSError:
                    pass
            if total and downloaded < total:
                raise http.client.IncompleteRead(b'', total - downloaded)
            if os.path.exists(self.save_path):
                try:
                    os.chmod(self.save_path, 0o666)
                except OSError:
                    pass
            os.replace(self.part_path, self.save_path)
            self.progress_changed.emit(downloaded, total, 0.0, -1, resumed, attempt)
            self.download_completed.emit(self.save_path)
            return True

    def run(self):
        last_error = None
        for attempt in range(1, self.MAX_RETRIES + 2):
            if self.isInterruptionRequested():
                self.download_canceled.emit(self.part_path, self._partial_size())
                return
            try:
                if attempt > 1:
                    self.status_changed.emit(f'网络中断，正在重试 ({attempt - 1}/{self.MAX_RETRIES})...')
                if self._download_once(attempt):
                    return
                return
            except urllib.error.HTTPError as e:
                last_error = e
                if e.code not in (500, 502, 503, 504) or attempt > self.MAX_RETRIES:
                    break
            except urllib.error.URLError as e:
                last_error = e
                if attempt > self.MAX_RETRIES:
                    break
            except (socket.timeout, TimeoutError, http.client.IncompleteRead, http.client.RemoteDisconnected, ConnectionError) as e:
                last_error = e
                if attempt > self.MAX_RETRIES:
                    break
            except OSError as e:
                last_error = e
                break
            except Exception as e:
                last_error = e
                break
            if not self._sleep_with_cancel(min(2 ** (attempt - 1), 8)):
                self.download_canceled.emit(self.part_path, self._partial_size())
                return
        message = str(last_error) if last_error else '未知错误'
        if isinstance(last_error, urllib.error.HTTPError):
            message = f'GitHub 返回错误：HTTP {last_error.code}'
        elif isinstance(last_error, urllib.error.URLError):
            message = f'网络连接失败：{last_error.reason}'
        cleanup_error = self._reset_partial()
        if cleanup_error is not None:
            message += (
                f'\n下载临时文件未能清理：{self.part_path}'
                f'\n{cleanup_error}'
            )
        self.download_failed.emit(message)


class ArchiveExtractThread(QThread):
    """在后台执行事务式解压；取消时会中止 ZIP 复制或终止 7-Zip 子进程。"""
    status_changed = pyqtSignal(str)
    extract_completed = pyqtSignal(str)
    extract_failed = pyqtSignal(str)
    extract_canceled = pyqtSignal()

    def __init__(self, archive_path, sevenzip=None, parent=None):
        super().__init__(parent)
        self.archive_path = archive_path
        self.sevenzip = sevenzip
        self._process = None

    def _set_process(self, process):
        self._process = process

    def cancel(self):
        self.requestInterruption()
        process = self._process
        if process is not None and process.poll() is None:
            try:
                process.terminate()
            except OSError:
                pass

    def run(self):
        try:
            destination = _transactional_extract_archive(
                self.archive_path,
                sevenzip=self.sevenzip,
                is_canceled=self.isInterruptionRequested,
                progress_callback=self.status_changed.emit,
                process_callback=self._set_process,
            )
            # 提交是最后一个原子步骤；一旦返回，磁盘结果已成功，之后到达的取消不能改报失败。
            self.extract_completed.emit(destination)
        except ArchiveExtractionCanceled:
            self.extract_canceled.emit()
        except Exception as e:
            self.extract_failed.emit(str(e))
        finally:
            self._process = None


class FolderScanThread(QThread):
    """文件夹扫描线程，用于异步加载项目文件夹"""
    scan_completed = pyqtSignal(list, list)  # 发射(主板文件夹列表, 子卡文件夹列表)
    scan_started = pyqtSignal()
    scan_progress = pyqtSignal(str)  # 发射当前扫描的目录
    
    MAX_FILES = 50000
    MAX_MATCHES_WARNING = 500

    def __init__(self, settings, include_subfolders, comments, sort_by_number=False, mb_regex=None, db_regex=None):
        super().__init__()
        self.settings = settings
        self.include_subfolders = include_subfolders
        self.comments = comments
        self.sort_by_number = sort_by_number
        # 线程安全：在创建时快照正则，不在运行中读共享状态
        self._mb_regex = mb_regex or DEFAULT_MB_RE
        self._db_regex = db_regex or DEFAULT_DB_RE

    def _scan_directory(self, directory, dir_name, motherboard_folders, daughterboard_folders):
        """扫描单个目录，收集匹配的项目文件夹"""
        if self.isInterruptionRequested() or not os.path.exists(directory):
            return
        try:
            items = os.listdir(directory)
            for item in items:
                if self.isInterruptionRequested():
                    return
                item_path = os.path.join(directory, item)
                if os.path.isdir(item_path) and not os.path.islink(item_path):
                    # 先匹配主板，再匹配子卡（双正则独立）
                    mb_match = self._mb_regex.match(item)
                    db_match = self._db_regex.match(item)
                    if mb_match:
                        number = mb_match.group(1)
                        folder_comment = mb_match.group(2) if mb_match.group(2) else ''
                        internal_comment = self.comments.get(item_path, folder_comment)
                        motherboard_folders.append((int(number), item_path, number, internal_comment, dir_name))
                    if db_match:
                        number = db_match.group(1)
                        folder_comment = db_match.group(2) if db_match.group(2) else ''
                        internal_comment = self.comments.get(item_path, folder_comment)
                        daughterboard_folders.append((int(number), item_path, number, internal_comment, dir_name))
                    if self.include_subfolders:
                        self._scan_directory(item_path, dir_name, motherboard_folders, daughterboard_folders)
        except Exception as e:
            if not self.isInterruptionRequested():
                self.scan_progress.emit(f"扫描目录 {directory} 时出错: {str(e)}")

    def run(self):
        """线程运行方法"""
        root_dirs = self.settings or []
        dir_order = {name: idx for idx, (name, _) in enumerate(root_dirs)}
        motherboard_folders = []
        daughterboard_folders = []
        for dir_name, root_dir in root_dirs:
            if self.isInterruptionRequested():
                return
            self.scan_progress.emit(f"正在扫描: {root_dir}")
            self._scan_directory(root_dir, dir_name, motherboard_folders, daughterboard_folders)
        if self.isInterruptionRequested():
            return
        if self.sort_by_number:
            motherboard_folders.sort(key=lambda x: x[0])
            daughterboard_folders.sort(key=lambda x: x[0])
        else:
            motherboard_folders.sort(key=lambda x: (dir_order.get(x[4], 999), x[0]))
            daughterboard_folders.sort(key=lambda x: (dir_order.get(x[4], 999), x[0]))
        if not self.isInterruptionRequested():
            self.scan_completed.emit(motherboard_folders, daughterboard_folders)


class FolderStatsThread(QThread):
    """递归统计某文件夹的文件数与总大小（off UI 线程，避免大目录/网络盘冻结界面）。"""
    stats_ready = pyqtSignal(int, int, int, bool)  # (token, file_count, total_size, truncated)

    MAX_FILES = 50000  # 软上限：超过即停，UI 显示 50000+

    def __init__(self, root, token):
        super().__init__()
        self.root = root
        self.token = token

    def run(self):
        count = 0
        total = 0
        truncated = False
        try:
            for dirpath, dirnames, filenames in os.walk(self.root):
                if self.isInterruptionRequested():
                    return
                for name in filenames:
                    if self.isInterruptionRequested():
                        return
                    count += 1
                    try:
                        total += os.path.getsize(os.path.join(dirpath, name))
                    except OSError:
                        # 权限/失联/已删除等：跳过单个文件，不中断整体
                        pass
                    if count >= self.MAX_FILES:
                        truncated = True
                        break
                if truncated:
                    break
        except Exception:
            # os.walk 顶层异常（root 失联等）：发已累计的部分结果
            pass
        if not self.isInterruptionRequested():
            self.stats_ready.emit(self.token, count, total, truncated)


class FileSearchThread(QThread):
    """在某个文件夹下递归搜索文件：按文件名 + 扩展名 + 修改时间过滤（off UI 线程）。"""
    search_ready = pyqtSignal(int, list, bool)  # (token, results, truncated)

    MAX_RESULTS = 2000  # 软上限：超过即停，UI 提示

    def __init__(self, root, token, name_filter, exts, mtime_after):
        super().__init__()
        self.root = root
        self.token = token
        self.name_filter = (name_filter or '').lower()
        self.exts = exts            # None 或一组扩展名小写（含点）
        self.mtime_after = mtime_after  # None 或时间戳下限（含）

    def run(self):
        results = []
        truncated = False
        try:
            for dirpath, dirnames, filenames in os.walk(self.root):
                if self.isInterruptionRequested():
                    return
                for name in filenames:
                    if self.isInterruptionRequested():
                        return
                    ext = os.path.splitext(name)[1].lower()
                    if self.exts is not None and ext not in self.exts:
                        continue
                    if self.name_filter and self.name_filter not in name.lower():
                        continue
                    full = os.path.join(dirpath, name)
                    try:
                        st_mtime = os.path.getmtime(full)
                        size = os.path.getsize(full)
                    except OSError:
                        continue  # 权限/失联/已删除：跳过
                    if self.mtime_after is not None and st_mtime < self.mtime_after:
                        continue
                    rel = os.path.relpath(full, self.root)
                    results.append((name, rel, full, size, st_mtime))
                    if len(results) >= self.MAX_RESULTS:
                        truncated = True
                        break
                if truncated:
                    break
        except Exception:
            pass
        if not self.isInterruptionRequested():
            self.search_ready.emit(self.token, results, truncated)


class NewProjectDialog(QDialog):
    def __init__(self, parent=None, default_folder=None):
        super().__init__(parent)
        self.project_type = '主板'
        self.pcb_number = ''
        self.comment = ''
        self.target_folder = default_folder or os.path.expanduser('~')
        self.parent_window = parent
        self.setWindowTitle('新建项目文件夹')
        self.setGeometry(300, 300, 500, 300)
        
        layout = QVBoxLayout()
        
        # 目标文件夹选择
        folder_group = QGroupBox('目标文件夹')
        folder_layout = QHBoxLayout()
        self.folder_label = QLabel(self.target_folder)
        self.folder_label.setStyleSheet('font-family: "Cascadia Mono", "Consolas", "Courier New", monospace;')
        self.folder_label.setToolTip('点击更改目标文件夹')
        self.folder_label.setCursor(Qt.PointingHandCursor)
        self.folder_label.mousePressEvent = self.select_folder
        
        folder_btn = QPushButton('选择文件夹')
        folder_btn.clicked.connect(self.select_folder)
        
        folder_layout.addWidget(self.folder_label)
        folder_layout.addWidget(folder_btn)
        folder_group.setLayout(folder_layout)
        
        type_group = QGroupBox('项目类型')
        type_layout = QVBoxLayout()
        self.type_group = QButtonGroup()
        self.motherboard_radio = QRadioButton('主板 (S开头)')
        self.motherboard_radio.setChecked(True)
        self.motherboard_radio.toggled.connect(lambda: self.set_project_type('主板'))
        self.daughterboard_radio = QRadioButton('子卡 (M开头)')
        self.daughterboard_radio.toggled.connect(lambda: self.set_project_type('子卡'))
        self.type_group.addButton(self.motherboard_radio)
        self.type_group.addButton(self.daughterboard_radio)
        type_layout.addWidget(self.motherboard_radio)
        type_layout.addWidget(self.daughterboard_radio)
        type_group.setLayout(type_layout)
        
        info_group = QGroupBox('项目信息')
        info_layout = QFormLayout()
        self.number_edit = QLineEdit()
        self.number_edit.setPlaceholderText('请输入3-4位数字')
        self.number_edit.textChanged.connect(self.on_number_changed)
        self.comment_edit = QLineEdit()
        self.comment_edit.setPlaceholderText('请输入项目注释')
        self.comment_edit.textChanged.connect(self.on_comment_changed)
        self.preview_label = QLabel('')
        self.preview_label.setStyleSheet('font-weight: bold; color: blue;')
        info_layout.addRow('PCB编号：', self.number_edit)
        info_layout.addRow('项目注释：', self.comment_edit)
        info_layout.addRow('预览：', self.preview_label)
        info_group.setLayout(info_layout)
        
        button_layout = QHBoxLayout()
        self.create_btn = QPushButton('创建')
        self.create_btn.clicked.connect(self.create_project)
        self.create_btn.setEnabled(False)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(self.create_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addWidget(folder_group)
        layout.addWidget(type_group)
        layout.addWidget(info_group)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
    def set_project_type(self, project_type):
        self.project_type = project_type
        self.update_preview()
        
    def on_number_changed(self, text):
        self.pcb_number = text
        self.update_preview()
        
    def on_comment_changed(self, text):
        self.comment = text
        self.update_preview()
        
    def update_preview(self):
        if not self.pcb_number:
            self.preview_label.setText('')
            self.create_btn.setEnabled(False)
            return
        if not re.match(r'^\d{3,4}$', self.pcb_number):
            self.preview_label.setText('编号必须是3-4位数字')
            self.preview_label.setStyleSheet('font-weight: bold; color: red;')
            self.create_btn.setEnabled(False)
            return
        prefix = 'S' if self.project_type == '主板' else 'M'
        if self.comment:
            folder_name = f'{prefix}{self.pcb_number}_{self.comment}'
        else:
            folder_name = f'{prefix}{self.pcb_number}'
        self.preview_label.setText(folder_name)
        self.preview_label.setStyleSheet('font-weight: bold; color: blue;')
        self.create_btn.setEnabled(True)
        
    def select_folder(self, event=None):
        """选择目标文件夹"""
        folder_path = QFileDialog.getExistingDirectory(self, '选择目标文件夹', self.target_folder)
        if not folder_path:
            return
        self.target_folder = folder_path
        self.folder_label.setText(folder_path)
        if not self.parent_window:
            return
        self.parent_window.default_new_project_folder = folder_path
        settings = getattr(self.parent_window, 'settings', None) or []
        for name, path in settings:
            if path == folder_path:
                return
        folder_name = os.path.basename(folder_path) or "自定义路径"
        settings.append((folder_name, folder_path))
        self.parent_window.settings = settings
        self.parent_window.save_settings_to_file(
            self.parent_window.settings,
            self.parent_window.include_subfolders,
            folder_path
        )
    
    def create_project(self):
        if not re.match(r'^\d{3,4}$', self.pcb_number):
            QMessageBox.warning(self, '警告', 'PCB编号必须是3-4位数字')
            return
        prefix = 'S' if self.project_type == '主板' else 'M'
        if self.comment:
            folder_name = f'{prefix}{self.pcb_number}_{self.comment}'
        else:
            folder_name = f'{prefix}{self.pcb_number}'
        target_path = os.path.join(self.target_folder, folder_name)
        if os.path.exists(target_path):
            QMessageBox.warning(self, '警告', f'文件夹 {folder_name} 已存在')
            return
        try:
            os.makedirs(target_path)
            QMessageBox.information(self, '成功', f'项目文件夹 {folder_name} 已创建')
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, '错误', f'创建文件夹失败: {str(e)}')
            
    def get_project_info(self):
        prefix = 'S' if self.project_type == '主板' else 'M'
        if self.comment:
            folder_name = f'{prefix}{self.pcb_number}_{self.comment}'
        else:
            folder_name = f'{prefix}{self.pcb_number}'
        return {
            'type': self.project_type,
            'number': self.pcb_number,
            'comment': self.comment,
            'folder_name': folder_name,
            'full_path': os.path.join(self.target_folder, folder_name)
        }

class NewStructureDialog(QDialog):
    def __init__(self, project_folder=None, parent=None):
        super().__init__(parent)
        self.version = '00'
        self.selected_folders = {name: True for name in DEFAULT_STRUCTURE_FOLDERS}
        self.custom_folders = []
        self.project_folder = project_folder
        self.parent_window = parent
        
        # 加载保存的文件夹结构设置
        if parent and hasattr(parent, 'folder_structure'):
            saved_structure = parent.folder_structure
            if 'version' in saved_structure:
                self.version = saved_structure['version']
            if 'selected_folders' in saved_structure:
                self.selected_folders = saved_structure['selected_folders']
            if 'custom_folders' in saved_structure:
                # 过滤掉空字符串
                self.saved_custom_folders = [f for f in saved_structure['custom_folders'] if f.strip()]
        
        # 设置窗口标题
        title = '新建文件夹内部结构'
        if self.project_folder:
            project_name = os.path.basename(self.project_folder)
            title += f' - 项目：{project_name}'
        self.setWindowTitle(title)
        self.setGeometry(300, 300, 450, 420)
        
        layout = QVBoxLayout()
        
        # 添加项目文件夹信息标签
        if self.project_folder:
            project_info_group = QGroupBox('项目信息')
            project_info_layout = QHBoxLayout()
            project_info_label = QLabel(f'当前项目文件夹：')
            project_path_label = QLabel(self.project_folder)
            project_path_label.setWordWrap(True)
            project_path_label.setToolTip(self.project_folder)
            project_info_layout.addWidget(project_info_label)
            project_info_layout.addWidget(project_path_label, 1)
            project_info_group.setLayout(project_info_layout)
            layout.addWidget(project_info_group)
        
        version_group = QGroupBox('版本选择')
        version_layout = QHBoxLayout()
        version_label = QLabel('版本号：V')
        self.version_edit = QLineEdit('00')
        self.version_edit.setMaxLength(2)
        self.version_edit.setFixedWidth(50)
        self.version_edit.setValidator(QIntValidator(0, 99))
        self.version_edit.textChanged.connect(self.on_version_changed)
        version_layout.addWidget(version_label)
        version_layout.addWidget(self.version_edit)
        version_layout.addStretch()
        version_group.setLayout(version_layout)
        
        folders_group = QGroupBox('常用文件夹（默认勾选）')
        folders_layout = QVBoxLayout()
        
        self.folder_names = list(DEFAULT_STRUCTURE_FOLDERS)
        self.folder_checkboxes = {}
        checkbox_layout = QGridLayout()
        for i, name in enumerate(self.folder_names):
            checkbox = QPushButton(name)
            checkbox.setCheckable(True)
            checkbox.setChecked(self.selected_folders.get(name, True))
            checkbox.toggled.connect(lambda c, n=name: self.on_folder_toggled(n, c))
            self.folder_checkboxes[name] = checkbox
            checkbox_layout.addWidget(checkbox, i // 2, i % 2)
        folders_layout.addLayout(checkbox_layout)
        
        # 添加全部恢复默认按钮
        reset_btn = QPushButton('全部恢复默认')
        reset_btn.clicked.connect(self.reset_to_defaults)
        folders_layout.addWidget(reset_btn, alignment=Qt.AlignRight)
        
        folders_group.setLayout(folders_layout)
        
        custom_group = QGroupBox('自定义文件夹')
        custom_layout = QVBoxLayout()
        self.custom_container = QWidget()
        self.custom_container_layout = QVBoxLayout(self.custom_container)
        self.custom_container_layout.setContentsMargins(0, 0, 0, 0)
        add_btn = QPushButton('+ 添加自定义文件夹')
        add_btn.clicked.connect(self.add_custom_folder)
        custom_layout.addWidget(self.custom_container)
        custom_layout.addWidget(add_btn)
        custom_group.setLayout(custom_layout)
        
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setFixedHeight(100)
        
        button_layout = QHBoxLayout()
        create_btn = QPushButton('创建')
        create_btn.clicked.connect(self.create_structure)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(create_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addWidget(version_group)
        layout.addWidget(folders_group)
        layout.addWidget(custom_group)
        layout.addWidget(QLabel('文件夹结构预览：'))
        layout.addWidget(self.preview_text)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # 初始化UI控件状态
        self.version_edit.setText(self.version)
        
        # 添加保存的自定义文件夹
        if hasattr(self, 'saved_custom_folders'):
            for folder_name in self.saved_custom_folders:
                self.add_custom_folder_with_text(folder_name)
        
        self.update_preview()
        
    def on_version_changed(self, text):
        if text and len(text) <= 2:
            self.version = text.zfill(2)
        self.update_preview()
        
    def on_folder_toggled(self, folder_name, checked):
        self.selected_folders[folder_name] = checked
        self.update_preview()
    
    def reset_to_defaults(self):
        """将常用文件夹恢复为默认勾选状态，并删除所有自定义文件夹"""
        self.selected_folders = {name: True for name in self.folder_names}
        
        for name, checkbox in self.folder_checkboxes.items():
            checkbox.setChecked(True)
        
        if hasattr(self, 'custom_folders'):
            self.custom_folders.clear()
            while self.custom_container_layout.count() > 0:
                item = self.custom_container_layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
        
        self.update_preview()
        
    def add_custom_folder(self):
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 5, 0, 5)
        edit = QLineEdit()
        edit.setPlaceholderText('输入文件夹名称')
        edit.textChanged.connect(self.update_preview)
        remove_btn = QPushButton('-')
        remove_btn.setFixedWidth(30)
        remove_btn.clicked.connect(lambda: self.remove_custom_folder(row_widget, edit))
        row_layout.addWidget(edit)
        row_layout.addWidget(remove_btn)
        self.custom_container_layout.addWidget(row_widget)
        self.custom_folders.append(edit)
        self.update_preview()
        
    def add_custom_folder_with_text(self, folder_name):
        """添加带有预设文本的自定义文件夹"""
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 5, 0, 5)
        edit = QLineEdit(folder_name)
        edit.textChanged.connect(self.update_preview)
        remove_btn = QPushButton('-')
        remove_btn.setFixedWidth(30)
        remove_btn.clicked.connect(lambda: self.remove_custom_folder(row_widget, edit))
        row_layout.addWidget(edit)
        row_layout.addWidget(remove_btn)
        self.custom_container_layout.addWidget(row_widget)
        self.custom_folders.append(edit)
        self.update_preview()
        
    def remove_custom_folder(self, widget, edit):
        widget.deleteLater()
        if edit in self.custom_folders:
            self.custom_folders.remove(edit)
        self.update_preview()
        
    def update_preview(self):
        preview = f'V{self.version}/\n'
        for folder_name, checked in self.selected_folders.items():
            if checked:
                preview += f'  - {folder_name}\n'
        for edit in self.custom_folders:
            folder_name = edit.text().strip()
            if folder_name:
                preview += f'  - {folder_name}\n'
        self.preview_text.setPlainText(preview)
        
    def create_structure(self):
        if not re.match(r'^\d{2}$', self.version):
            QMessageBox.warning(self, '警告', '版本号必须是两位数字')
            return
        self.save_folder_structure()
        self.accept()
        
    def save_folder_structure(self):
        """保存文件夹结构设置到主窗口"""
        if self.parent_window:
            # 收集当前设置
            custom_folders = [edit.text().strip() for edit in self.custom_folders if edit.text().strip()]
            
            # 更新主窗口的文件夹结构设置
            self.parent_window.folder_structure = {
                'version': self.version,
                'selected_folders': self.selected_folders,
                'custom_folders': custom_folders
            }
            
            # 保存到文件
            self.parent_window.save_settings_to_file(
                self.parent_window.settings,
                self.parent_window.include_subfolders,
                self.parent_window.default_new_project_folder
            )
    
    def reject(self):
        """重写取消方法(closeEvent 会统一保存设置,此处无需重复)"""
        super().reject()
        
    def closeEvent(self, event):
        """重写关闭事件，保存设置"""
        self.save_folder_structure()
        super().closeEvent(event)
    
    def get_structure_info(self):
        selected_folders = [name for name, checked in self.selected_folders.items() if checked]
        custom_folders = [edit.text().strip() for edit in self.custom_folders if edit.text().strip()]
        return {
            'version': self.version,
            'selected_folders': selected_folders,
            'custom_folders': custom_folders
        }

class _ReorderableTableDialog(QDialog):
    """带可重排 QTableWidget（self.path_list）的对话框基类，封装上移/下移/删除行。
    子类负责建表与 add_path/save_settings。_swap_rows 通用处理任意列数，保留 text/flags/勾选态。"""

    @staticmethod
    def _clone_cell(item):
        new = QTableWidgetItem(item.text() if item is not None else '')
        if item is not None:
            new.setFlags(item.flags())
            if item.flags() & Qt.ItemIsUserCheckable:
                new.setCheckState(item.checkState())
        return new

    def _swap_rows(self, row1, row2):
        for col in range(self.path_list.columnCount()):
            # 先克隆两个单元格再写回：setItem 会销毁原 C++ 对象，必须在覆盖前完成克隆
            new_for_row1 = self._clone_cell(self.path_list.item(row2, col))
            new_for_row2 = self._clone_cell(self.path_list.item(row1, col))
            self.path_list.setItem(row1, col, new_for_row1)
            self.path_list.setItem(row2, col, new_for_row2)

    def remove_path(self):
        current_row = self.path_list.currentRow()
        if current_row >= 0:
            self.path_list.removeRow(current_row)

    def move_up(self):
        current_row = self.path_list.currentRow()
        if current_row > 0:
            self._swap_rows(current_row, current_row - 1)
            self.path_list.setCurrentCell(current_row - 1, 0)

    def move_down(self):
        current_row = self.path_list.currentRow()
        if current_row >= 0 and current_row < self.path_list.rowCount() - 1:
            self._swap_rows(current_row, current_row + 1)
            self.path_list.setCurrentCell(current_row + 1, 0)


class SettingsDialog(_ReorderableTableDialog):
    def __init__(self, current_paths, include_subfolders=False, sort_by_number=False, parent=None, show_hidden=False, regex_state='default', custom_mb_regex='', custom_db_regex=''):
        super().__init__(parent)
        self.current_paths = current_paths if current_paths is not None else []
        self.paths = list(self.current_paths)
        self.include_subfolders = include_subfolders
        self.sort_by_number = sort_by_number
        self.show_hidden = show_hidden
        self.regex_state = regex_state
        self.custom_mb_regex = custom_mb_regex
        self.custom_db_regex = custom_db_regex
        self.setWindowTitle('项目文件夹设置')
        self.setGeometry(300, 300, 550, 450)
        
        layout = QVBoxLayout()
        
        self.path_list = QTableWidget(0, 2)
        self.path_list.setHorizontalHeaderLabels(['名称', '路径'])
        self.path_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.path_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for name, path in self.current_paths:
            row = self.path_list.rowCount()
            self.path_list.insertRow(row)
            self.path_list.setItem(row, 0, QTableWidgetItem(name))
            self.path_list.setItem(row, 1, QTableWidgetItem(path))
        
        button_layout = QHBoxLayout()
        add_btn = QPushButton('+ 添加文件夹')
        add_btn.clicked.connect(self.add_path)
        remove_btn = QPushButton('- 删除选中')
        remove_btn.clicked.connect(self.remove_path)
        button_layout.addWidget(add_btn)
        button_layout.addWidget(remove_btn)
        
        sort_layout = QHBoxLayout()
        up_btn = QPushButton('↑ 上移')
        up_btn.clicked.connect(self.move_up)
        down_btn = QPushButton('↓ 下移')
        down_btn.clicked.connect(self.move_down)
        sort_layout.addWidget(up_btn)
        sort_layout.addWidget(down_btn)
        sort_layout.addStretch()
        
        self.include_subfolders_checkbox = QPushButton('包含所有子文件夹(可能会导致一段时间卡顿无响应，请勿退出)')
        self.include_subfolders_checkbox.setCheckable(True)
        self.include_subfolders_checkbox.setChecked(self.include_subfolders)
        
        self.sort_by_number_checkbox = QPushButton('跨路径序号从小到大排序')
        self.sort_by_number_checkbox.setCheckable(True)
        self.sort_by_number_checkbox.setChecked(self.sort_by_number)
        
        self.show_hidden_checkbox = QPushButton('显示隐藏文件（以.开头或系统隐藏属性）')
        self.show_hidden_checkbox.setCheckable(True)
        self.show_hidden_checkbox.setChecked(self.show_hidden)
        
        # --- 正则设置区域 ---
        regex_group = QGroupBox('项目文件夹命名规则')
        regex_layout = QVBoxLayout()
        
        self.regex_default_rb = QRadioButton('默认 (主板 S + 编号, 子卡 M + 编号)')
        self.regex_custom_rb = QRadioButton('自定义正则')
        regex_layout.addWidget(self.regex_default_rb)
        regex_layout.addWidget(self.regex_custom_rb)
        self.regex_custom_rb.clicked.connect(lambda checked: self._on_regex_mode_changed('custom'))
        self.regex_default_rb.clicked.connect(lambda checked: self._on_regex_mode_changed('default'))
        
        # 主板正则输入
        mb_row = QHBoxLayout()
        mb_row.addWidget(QLabel('主板:'))
        self.regex_mb_edit = QLineEdit()
        self.regex_mb_edit.setPlaceholderText(r'^S(\d{3,4})(?:_(.*))?$')
        mb_row.addWidget(self.regex_mb_edit)
        regex_layout.addLayout(mb_row)
        
        # 子卡正则输入
        db_row = QHBoxLayout()
        db_row.addWidget(QLabel('子卡:'))
        self.regex_db_edit = QLineEdit()
        self.regex_db_edit.setPlaceholderText(r'^M(\d{3,4})(?:_(.*))?$')
        db_row.addWidget(self.regex_db_edit)
        regex_layout.addLayout(db_row)
        
        # 测试按钮和状态
        test_row = QHBoxLayout()
        self.regex_test_btn = QPushButton('测试')
        self.regex_test_btn.clicked.connect(self._test_regex)
        test_row.addWidget(self.regex_test_btn)
        test_row.addStretch()
        regex_layout.addLayout(test_row)
        
        self.regex_status_label = QLabel()
        self.regex_status_label.setWordWrap(True)
        regex_layout.addWidget(self.regex_status_label)
        
        regex_group.setLayout(regex_layout)
        layout.addWidget(regex_group)
        
        save_btn = QPushButton('保存设置')
        save_btn.clicked.connect(self.save_settings)
        
        layout.addWidget(QLabel('项目文件路径（可调整顺序）：'))
        layout.addWidget(self.path_list)
        layout.addLayout(button_layout)
        layout.addLayout(sort_layout)
        layout.addWidget(self.include_subfolders_checkbox)
        layout.addWidget(self.sort_by_number_checkbox)
        layout.addWidget(self.show_hidden_checkbox)
        layout.addWidget(save_btn)
        self.setLayout(layout)
        
        # 初始化 UI 状态
        self._refresh_regex_ui()
        
    def _refresh_regex_ui(self):
        """根据 self.regex_state 同步 UI 状态和编辑框内容。"""
        is_custom = self.regex_state == 'custom'
        self.regex_custom_rb.setChecked(is_custom)
        self.regex_default_rb.setChecked(not is_custom)
        self.regex_mb_edit.setText(self.custom_mb_regex if is_custom else DEFAULT_MB_RE_TEXT)
        self.regex_db_edit.setText(self.custom_db_regex if is_custom else DEFAULT_DB_RE_TEXT)
        self.regex_mb_edit.setEnabled(is_custom)
        self.regex_db_edit.setEnabled(is_custom)
        self.regex_test_btn.setEnabled(is_custom)
        self._test_regex()

    def _on_regex_mode_changed(self, state):
        print(f'DEBUG: _on_regex_mode_changed called with state={state}')
        self.regex_state = state
        self._refresh_regex_ui()
        print(f'DEBUG: regex_state={self.regex_state}, mb_edit enabled={self.regex_mb_edit.isEnabled()}')


    def _test_regex(self):
        """实时验证两个自定义正则并更新状态标签。"""
        if self.regex_state == 'default':
            self.regex_status_label.setText(f'✅ 使用默认：主板 {DEFAULT_MB_RE_TEXT}，子卡 {DEFAULT_DB_RE_TEXT}')
            self.regex_status_label.setStyleSheet('color: #27ae60;')
            return
        mb_text = self.regex_mb_edit.text().strip()
        db_text = self.regex_db_edit.text().strip()
        errors = []
        if not mb_text:
            errors.append('主板正则不能为空')
        try:
            re.compile(mb_text)
        except re.error as e:
            errors.append(f'主板正则无效: {e}')
        if not db_text:
            errors.append('子卡正则不能为空')
        try:
            re.compile(db_text)
        except re.error as e:
            errors.append(f'子卡正则无效: {e}')
        if errors:
            self.regex_status_label.setText('❌ ' + '; '.join(errors) + '（保存后将回退到默认）')
            self.regex_status_label.setStyleSheet('color: #c0392b;')
        else:
            self.regex_status_label.setText(f'✅ 主板: {mb_text}  |  子卡: {db_text}')
            self.regex_status_label.setStyleSheet('color: #27ae60;')

    def add_path(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择项目文件夹')
        if folder_path:
            default_name = os.path.basename(folder_path)
            name, ok = QInputDialog.getText(self, '输入名称', '请输入显示名称：', text=default_name)
            if ok:
                if not name.strip():
                    name = default_name
                row = self.path_list.rowCount()
                self.path_list.insertRow(row)
                self.path_list.setItem(row, 0, QTableWidgetItem(name))
                self.path_list.setItem(row, 1, QTableWidgetItem(folder_path))

    def save_settings(self):
        paths = []
        for row in range(self.path_list.rowCount()):
            name_item = self.path_list.item(row, 0)
            path_item = self.path_list.item(row, 1)
            if name_item and path_item:
                paths.append((name_item.text(), path_item.text()))
        if not paths:
            QMessageBox.warning(self, '警告', '至少需要保留一个项目文件夹')
            return
        self.paths = paths
        self.include_subfolders = self.include_subfolders_checkbox.isChecked()
        self.sort_by_number = self.sort_by_number_checkbox.isChecked()
        self.show_hidden = self.show_hidden_checkbox.isChecked()
        self.regex_state = 'custom' if self.regex_custom_rb.isChecked() else 'default'
        self.custom_mb_regex = self.regex_mb_edit.text().strip()
        self.custom_db_regex = self.regex_db_edit.text().strip()
        self.accept()
        
    def get_settings(self):
        return self.paths, self.include_subfolders, self.sort_by_number, self.show_hidden, self.regex_state, self.custom_mb_regex, self.custom_db_regex


class SevenZipSettingsDialog(QDialog):
    """配置 RAR/7Z 预览授权，以及预览和解压共用的 7z.exe 路径。"""

    def __init__(self, current_path='', enable_7zip=False, parent=None):
        super().__init__(parent)
        self.archive_tool_path = str(current_path or '')
        self.enable_7zip = bool(enable_7zip)
        self.setWindowTitle('7-Zip 设置')
        self.setMinimumWidth(560)

        layout = QVBoxLayout()

        self.enable_checkbox = QCheckBox('启用 .rar / .7z 的 7-Zip 预览')
        self.enable_checkbox.setChecked(self.enable_7zip)
        layout.addWidget(self.enable_checkbox)

        warning_label = QLabel(
            '安全提示：启用后，预览 .rar / .7z 时会启动外部 7z.exe。'
            '请仅使用可信来源的官方 7-Zip，及时更新，并谨慎处理来源不明的压缩包。'
        )
        warning_label.setWordWrap(True)
        warning_label.setStyleSheet('color: #8a4b08;')
        layout.addWidget(warning_label)

        path_layout = QHBoxLayout()
        path_layout.addWidget(QLabel('7z.exe 路径：'))
        self.path_edit = QLineEdit(self.archive_tool_path)
        self.path_edit.setPlaceholderText('留空时自动检测 7-Zip 安装位置')
        browse_btn = QPushButton('浏览...')
        browse_btn.clicked.connect(self.browse_path)
        path_layout.addWidget(self.path_edit, 1)
        path_layout.addWidget(browse_btn)
        layout.addLayout(path_layout)

        button_layout = QHBoxLayout()
        save_btn = QPushButton('保存设置')
        save_btn.clicked.connect(self.save_settings)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

        self.setLayout(layout)

    def browse_path(self):
        current = self.path_edit.text().strip().strip('"')
        start_dir = os.path.dirname(current) if current else ''
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            '选择 7z.exe',
            start_dir,
            '7-Zip 可执行文件 (7z.exe);;可执行文件 (*.exe)',
        )
        if file_path:
            self.path_edit.setText(file_path)

    def save_settings(self):
        raw_path = self.path_edit.text().strip().strip('"')
        enable_7zip = self.enable_checkbox.isChecked()
        archive_tool_path = ''
        if raw_path:
            archive_tool_path = os.path.abspath(
                os.path.normpath(os.path.expandvars(os.path.expanduser(raw_path)))
            )
            if enable_7zip and not os.path.isfile(archive_tool_path):
                QMessageBox.warning(self, '路径无效', '指定的 7z.exe 文件不存在。')
                return
            if enable_7zip and os.path.basename(archive_tool_path).casefold() != '7z.exe':
                QMessageBox.warning(self, '路径无效', '请选择名称为 7z.exe 的可执行文件。')
                return

        self.archive_tool_path = archive_tool_path
        self.enable_7zip = enable_7zip
        self.accept()

    def get_settings(self):
        return self.archive_tool_path, self.enable_7zip


class QuickAccessSettingsDialog(_ReorderableTableDialog):
    def __init__(self, current_paths, parent=None):
        super().__init__(parent)
        self.current_paths = current_paths if current_paths else []
        self.paths = list(self.current_paths)
        self.setWindowTitle('快捷访问设置')
        self.setGeometry(300, 300, 550, 400)
        
        layout = QVBoxLayout()
        
        self.path_list = QTableWidget(0, 3)
        self.path_list.setHorizontalHeaderLabels(['名称', '路径', '不显示预览'])
        self.path_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.path_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.path_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        
        for item in self.current_paths:
            if len(item) == 3:
                name, path, no_preview = item
            else:
                name, path = item
                no_preview = False
            row = self.path_list.rowCount()
            self.path_list.insertRow(row)
            self.path_list.setItem(row, 0, QTableWidgetItem(name))
            self.path_list.setItem(row, 1, QTableWidgetItem(path))
            check_item = QTableWidgetItem()
            check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            check_item.setCheckState(Qt.Checked if no_preview else Qt.Unchecked)
            self.path_list.setItem(row, 2, check_item)
        
        button_layout = QHBoxLayout()
        add_btn = QPushButton('+ 添加文件夹')
        add_btn.clicked.connect(self.add_path)
        remove_btn = QPushButton('- 删除选中')
        remove_btn.clicked.connect(self.remove_path)
        button_layout.addWidget(add_btn)
        button_layout.addWidget(remove_btn)
        
        sort_layout = QHBoxLayout()
        up_btn = QPushButton('↑ 上移')
        up_btn.clicked.connect(self.move_up)
        down_btn = QPushButton('↓ 下移')
        down_btn.clicked.connect(self.move_down)
        sort_layout.addWidget(up_btn)
        sort_layout.addWidget(down_btn)
        sort_layout.addStretch()
        
        reset_btn = QPushButton('恢复默认')
        reset_btn.clicked.connect(self.reset_to_default)
        
        save_btn = QPushButton('保存设置')
        save_btn.clicked.connect(self.save_settings)
        
        hint_label = QLabel('提示：大文件夹与网络文件夹建议勾选"不显示预览"')
        hint_label.setStyleSheet('color: gray; font-size: 11px;')
        
        layout.addWidget(QLabel('快捷访问路径（可调整顺序）：'))
        layout.addWidget(self.path_list)
        layout.addLayout(button_layout)
        layout.addLayout(sort_layout)
        layout.addWidget(reset_btn)
        layout.addWidget(save_btn)
        layout.addWidget(hint_label)
        self.setLayout(layout)
    
    def _default_name_from_path(self, path):
        """从路径提取显示名,对 UNC 路径特殊处理。"""
        p = path.replace('\\', '/').rstrip('/')
        base = os.path.basename(p)
        if base:
            return base
        # UNC 路径如 //server/share 取 share
        parts = [x for x in p.split('/') if x]
        if len(parts) >= 2 and parts[0] == '':
            # //server/share -> 取 share
            return parts[2] if len(parts) >= 3 else parts[-1]
        return parts[-1] if parts else '网络文件夹'

    def add_path(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择文件夹')
        if folder_path:
            default_name = self._default_name_from_path(folder_path)
            name, ok = QInputDialog.getText(self, '输入名称', '请输入显示名称：', text=default_name)
            if ok:
                if not name.strip():
                    name = default_name
                row = self.path_list.rowCount()
                self.path_list.insertRow(row)
                self.path_list.setItem(row, 0, QTableWidgetItem(name))
                self.path_list.setItem(row, 1, QTableWidgetItem(folder_path))
                check_item = QTableWidgetItem()
                check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                check_item.setCheckState(Qt.Unchecked)
                self.path_list.setItem(row, 2, check_item)

    def reset_to_default(self):
        if hasattr(self.parent(), '_get_default_quick_access_paths'):
            default_paths = self.parent()._get_default_quick_access_paths()
            self.path_list.setRowCount(0)
            for item in default_paths:
                if len(item) == 3:
                    name, path, no_preview = item
                else:
                    name, path = item
                    no_preview = False
                row = self.path_list.rowCount()
                self.path_list.insertRow(row)
                self.path_list.setItem(row, 0, QTableWidgetItem(name))
                self.path_list.setItem(row, 1, QTableWidgetItem(path))
                check_item = QTableWidgetItem()
                check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                check_item.setCheckState(Qt.Checked if no_preview else Qt.Unchecked)
                self.path_list.setItem(row, 2, check_item)
    
    def save_settings(self):
        paths = []
        for row in range(self.path_list.rowCount()):
            name_item = self.path_list.item(row, 0)
            path_item = self.path_list.item(row, 1)
            check_item = self.path_list.item(row, 2)
            if name_item and path_item:
                no_preview = check_item.checkState() == Qt.Checked if check_item else False
                paths.append((name_item.text(), path_item.text(), no_preview))
        self.paths = paths
        self.accept()
    
    def get_settings(self):
        return self.paths


class WizardDialog(QDialog):
    """新手向导：分页介绍几项核心功能，仅首次自动弹出，可随时跳过。"""

    # (标题, HTML 正文)
    PAGES = [
        (
            '欢迎使用 SeavoExplorer',
            '''
            <p>SeavoExplorer 是面向 S/M 主板项目的文件浏览器，帮你快速定位项目、
            预览工程文档、整理版本目录。</p>
            <p>下面用几步介绍几项核心功能。你可以随时点击<b>“跳过”</b>关闭向导，
            之后也能在菜单 <b>帮助 → 新手向导</b> 中重新打开。</p>
            '''
        ),
        (
            '一、配置并浏览项目',
            '''
            <p>首次使用请先在菜单 <b>设置 → 项目文件夹设置</b> 中添加包含项目的根目录。</p>
            <p>程序会自动扫描其中符合命名规则的文件夹（以 <code>S</code> 或 <code>M</code>
            开头 + 3~4 位数字，可选 <code>_注释</code>，如 <code>S1234_样机</code>）。</p>
            <ul>
            <li><b>单击</b>左侧项目行 → 在右侧文件树中查看该项目文件</li>
            <li><b>双击编号列</b> → 在资源管理器中打开</li>
            <li>顶部<b>搜索框</b> → 实时过滤项目列表</li>
            </ul>
            '''
        ),
        (
            '二、预览工程文档',
            '''
            <p>在右侧文件树中<b>单击</b>文件即可在下方预览区查看内容，无需打开外部程序：</p>
            <ul>
            <li>文本 / PDF / Excel / Word / 图片 / 视频缩略图</li>
            <li>压缩包（.zip/.rar/.7z）以树状结构显示内容</li>
            </ul>
            <p><b>双击</b>文件用系统默认程序打开；切换到<b>元数据</b>标签可查看文件详情。</p>
            <p>对文件夹右键可选择<b>“在终端中打开”</b>，程序会优先尝试 Windows Terminal，
            再回退到 PowerShell / cmd。</p>
            '''
        ),
        (
            '三、文件操作（支持多选）',
            '''
            <p>文件树支持按住 <b>Ctrl</b> / <b>Shift</b> 多选，再进行批量操作：</p>
            <ul>
            <li><b>Ctrl+C / 右键复制</b>：复制到剪贴板，可在资源管理器粘贴，也保留程序内“粘贴副本”</li>
            <li><b>Ctrl+V / 右键粘贴副本</b>：粘贴到选中文件夹或当前项目（自动处理重名）</li>
            <li><b>F2 / 右键重命名</b>：重命名单个文件</li>
            <li><b>右键保存版本</b>：为文件生成日期版本副本（如 S1200-10_20260708.dsn），自动递增字母后缀</li>
            <li><b>右键归档到old文件夹</b>：将选中文件移入同目录下的 old/ 文件夹（自动创建），支持多选</li>
            <li><b>Delete</b>：移入回收站</li>
            <li>右键还可“添加到 zip 压缩包”“智能解压”</li>
            </ul>
            '''
        ),
        (
            '四、快捷访问与项目结构',
            '''
            <p>快捷访问栏位于窗口顶部，适合放常用项目根目录、资料目录或网络目录。</p>
            <ul>
            <li><b>+</b> 按钮 → 打开快捷访问设置</li>
            <li><b>右键快捷访问按钮</b> → 可在终端中打开、删除该快捷访问、或进入快捷访问设置</li>
            </ul>
            <p>左侧 <b>“新建项目文件夹”</b>：按规则创建新的 S/M 项目根目录。</p>
            <p>选中项目后的 <b>“新建文件夹内部结构”</b>：在项目内创建版本目录（如 <code>V01</code>）
            及 BOM / SCH / 物料 / 评审 / 信号测试 等标准子文件夹。</p>
            <p>更详细的说明请见菜单 <b>帮助 → 使用帮助</b>。祝使用愉快！</p>
            '''
        ),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('新手向导')
        self.setMinimumSize(560, 460)
        icon_path = parent._resource_path('favicon.ico') if parent and hasattr(parent, '_resource_path') else ''
        if icon_path and os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.current_page = 0

        layout = QVBoxLayout(self)

        self.title_label = QLabel()
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        self.title_label.setFont(title_font)
        self.title_label.setStyleSheet('color: #2c3e50;')
        layout.addWidget(self.title_label)

        self.step_label = QLabel()
        self.step_label.setStyleSheet('color: #7f8c8d;')
        layout.addWidget(self.step_label)

        self.body = QTextEdit()
        self.body.setReadOnly(True)
        layout.addWidget(self.body, 1)

        btn_layout = QHBoxLayout()
        self.skip_btn = QPushButton('跳过')
        self.skip_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.skip_btn)
        btn_layout.addStretch()
        self.prev_btn = QPushButton('上一步')
        self.prev_btn.clicked.connect(self.go_prev)
        btn_layout.addWidget(self.prev_btn)
        self.next_btn = QPushButton('下一步')
        self.next_btn.clicked.connect(self.go_next)
        btn_layout.addWidget(self.next_btn)
        layout.addLayout(btn_layout)

        self._render_page()

    def _render_page(self):
        title, html = self.PAGES[self.current_page]
        self.title_label.setText(title)
        self.step_label.setText(f'第 {self.current_page + 1} / {len(self.PAGES)} 步')
        self.body.setHtml(html)
        self.prev_btn.setEnabled(self.current_page > 0)
        is_last = self.current_page == len(self.PAGES) - 1
        self.next_btn.setText('完成' if is_last else '下一步')
        self.skip_btn.setVisible(not is_last)

    def go_prev(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._render_page()

    def go_next(self):
        if self.current_page < len(self.PAGES) - 1:
            self.current_page += 1
            self._render_page()
        else:
            self.accept()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # 确定配置文件的保存位置
        # 对于打包后的程序，使用EXE文件所在目录
        self.app_dir = _get_app_dir()
        
        # 初始化配置和注释文件路径
        self.CONFIG_FILE = os.path.join(self.app_dir, 'seavoexplorer.json')
        self.COMMENTS_FILE = os.path.join(self.app_dir, 'seavo_comments.json')
        
        # 确保app_dir目录存在
        os.makedirs(self.app_dir, exist_ok=True)

        # 加载配置/注释期间累积的警告，待 UI 就绪后统一弹出（此时主窗口尚未构建，不能直接弹框）
        self._pending_load_warnings = []

        self.current_folder = None
        self.filtered_folders = {'主板': [], '子卡': []}
        self.include_subfolders = False
        self.sort_by_number = False
        self.archive_tool_path = ''
        self.pinned_folders = []
        self.hidden_folders = []
        self.comments = self.load_comments() or {}
        self.clipboard_path = None
        self.clipboard_paths = []
        self._extract_jobs = {}
        self._close_after_extract_cancel = False

        self.settings = self.load_settings()
        
        self.initUI()
        # 异步加载文件夹，提高启动速度
        self.load_filtered_folders_async()
        # 记录待恢复的上次项目，待扫描完成（表就绪）后由 on_scan_completed 触发恢复
        self._pending_restore_project = getattr(self, 'last_project_path', None)
        # UI 就绪后弹出加载期累积的警告（配置/注释损坏等）
        if self._pending_load_warnings:
            QTimer.singleShot(0, self._show_pending_load_warnings)
        # 首次运行自动弹出新手向导（窗口显示后再弹，避免阻塞启动）
        if not getattr(self, 'wizard_shown', False):
            QTimer.singleShot(0, self.show_wizard)

    def _show_pending_load_warnings(self):
        if self._pending_load_warnings:
            QMessageBox.warning(self, '提示', '\n'.join(self._pending_load_warnings))
            self._pending_load_warnings = []

    def initUI(self):
        self.setWindowTitle('主板项目文件浏览器')
        self.setGeometry(100, 100, 1400, 900)
        self._restore_window_geometry()
        
        icon_path = self._resource_path('favicon.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 0, 5, 5)
        
        # 快捷访问工具栏
        self.quick_access_toolbar = QWidget()
        quick_access_layout = QHBoxLayout(self.quick_access_toolbar)
        quick_access_layout.setContentsMargins(5, 0, 5, 0)
        quick_access_layout.setSpacing(2)
        self.quick_access_toolbar.setFixedHeight(28)
        quick_access_label = QLabel('快捷访问:')
        quick_access_layout.addWidget(quick_access_label)
        self.quick_access_add_btn = QPushButton('+')
        self.quick_access_add_btn.setFixedSize(22, 22)
        self.quick_access_add_btn.setToolTip('打开快捷访问设置')
        self.quick_access_add_btn.clicked.connect(self.show_quick_access_settings_dialog)
        quick_access_layout.addWidget(self.quick_access_add_btn)

        self.quick_access_buttons = []
        self._create_quick_access_buttons(quick_access_layout)
        quick_access_layout.addStretch()
        
        main_layout.addWidget(self.quick_access_toolbar)
        
        content_layout = QHBoxLayout()
        main_layout.addLayout(content_layout)
        
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(5, 5, 5, 5)
        
        folder_search_layout = QHBoxLayout()
        folder_search_label = QLabel('文件夹搜索:')
        self.folder_search_edit = QLineEdit()
        self.folder_search_edit.textChanged.connect(self.filter_folders)
        folder_search_layout.addWidget(folder_search_label)
        folder_search_layout.addWidget(self.folder_search_edit)
        
        self.motherboard_group = QGroupBox('主板')
        motherboard_layout = QVBoxLayout(self.motherboard_group)
        self.motherboard_table = QTableWidget(0, 2)
        self.motherboard_table.setHorizontalHeaderLabels(['编号', '注释'])
        self.motherboard_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.motherboard_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.motherboard_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.motherboard_table.setSelectionMode(QTableWidget.SingleSelection)
        self.motherboard_table.cellClicked.connect(self.on_folder_cell_clicked)
        self.motherboard_table.cellDoubleClicked.connect(self.on_folder_cell_double_clicked)
        self.motherboard_table.verticalHeader().setVisible(False)
        self.motherboard_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.motherboard_table.customContextMenuRequested.connect(lambda pos: self._show_folder_context_menu(self.motherboard_table, pos))
        motherboard_layout.addWidget(self.motherboard_table)
        
        self.daughterboard_group = QGroupBox('子卡')
        daughterboard_layout = QVBoxLayout(self.daughterboard_group)
        self.daughterboard_table = QTableWidget(0, 2)
        self.daughterboard_table.setHorizontalHeaderLabels(['编号', '注释'])
        self.daughterboard_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.daughterboard_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.daughterboard_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.daughterboard_table.setSelectionMode(QTableWidget.SingleSelection)
        self.daughterboard_table.cellClicked.connect(self.on_folder_cell_clicked)
        self.daughterboard_table.cellDoubleClicked.connect(self.on_folder_cell_double_clicked)
        self.daughterboard_table.verticalHeader().setVisible(False)
        self.daughterboard_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.daughterboard_table.customContextMenuRequested.connect(lambda pos: self._show_folder_context_menu(self.daughterboard_table, pos))
        daughterboard_layout.addWidget(self.daughterboard_table)
        
        button_layout = QHBoxLayout()
        self.new_project_btn = QPushButton('新建项目文件夹')
        self.new_project_btn.clicked.connect(self.new_project)
        button_layout.addWidget(self.new_project_btn)
        self.new_structure_btn = QPushButton('新建文件夹内部结构')
        self.new_structure_btn.clicked.connect(self.new_folder_structure)
        self.new_structure_btn.setEnabled(False)
        button_layout.addWidget(self.new_structure_btn)
        
        left_layout.addLayout(folder_search_layout)
        left_layout.addWidget(self.motherboard_group)
        left_layout.addWidget(self.daughterboard_group)
        left_layout.addLayout(button_layout)
        
        right_layout = QVBoxLayout()

        # 面包屑路径栏（置于文件树上方，跟随选中项，从项目根开始）
        self._breadcrumb_path = None
        self._breadcrumb_buttons = []
        self.breadcrumb_bar = self._build_breadcrumb_bar()
        right_layout.addWidget(self.breadcrumb_bar)

        # 文件搜索条（独立于左侧文件夹搜索，仅搜当前项目内文件）
        self._build_search_bar(right_layout)

        self.file_tree = QTreeView()
        self.file_model = QFileSystemModel()
        self.file_model.setFilter(QDir.NoDotAndDotDot | QDir.AllEntries)
        self.file_tree.setModel(self.file_model)
        self.file_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.file_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.file_tree.setColumnWidth(0, 300)
        self.file_tree.setSortingEnabled(True)
        self.file_tree.sortByColumn(0, Qt.AscendingOrder)
        self.file_tree.clicked.connect(self.on_file_clicked)
        self.file_tree.doubleClicked.connect(self.on_file_double_clicked)
        self.file_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.file_tree.customContextMenuRequested.connect(self.on_file_tree_context_menu)
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self)
        self.copy_shortcut.activated.connect(self.copy_selected_items)
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self)
        self.paste_shortcut.activated.connect(self.paste_to_selected_target)

        self.tabs = QTabWidget()
        
        # 文件预览容器
        self.preview_container = QWidget()
        self.preview_layout = QVBoxLayout(self.preview_container)
        
        # 文本预览
        self.preview_tab = QTextEdit()
        self.preview_tab.setReadOnly(True)
        self.preview_tab.setFont(get_mono_font(10))
        self.preview_layout.addWidget(self.preview_tab)

        # 预览被关闭时显示的按钮：点击后才真正读取并预览该文件
        self.preview_button = QPushButton('显示预览')
        self.preview_button.setMinimumHeight(40)
        self.preview_button.clicked.connect(self._on_preview_button_clicked)
        self.preview_layout.addWidget(self.preview_button)
        self.preview_button.hide()
        self._manual_preview_path = None
        self._scheduled_preview_path = None
        self._preview_timer = QTimer(self)
        self._preview_timer.setSingleShot(True)
        self._preview_timer.setInterval(250)
        self._preview_timer.timeout.connect(self._execute_pending_preview)
        
        # 图片预览
        self.image_scroll_area = QScrollArea()
        self.image_scroll_area.setAlignment(Qt.AlignCenter)
        self.image_scroll_area.setWidgetResizable(True)
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setCursor(Qt.PointingHandCursor)  # 点击手势
        self.image_label.mousePressEvent = self.show_full_image  # 点击事件
        self.image_scroll_area.setWidget(self.image_label)
        self.preview_layout.addWidget(self.image_scroll_area)
        self.image_scroll_area.hide()  # 初始隐藏图片预览
        
        self.tabs.addTab(self.preview_container, '文件预览')
        
        # 元数据预览
        self.metadata_tab = QTextEdit()
        self.metadata_tab.setReadOnly(True)
        self.tabs.addTab(self.metadata_tab, '元数据')
        
        # 右侧面板（文件树+预览）
        right_widget = QWidget()
        right_widget.setLayout(right_layout)

        # 文件树与搜索结果面板共享同一区段（搜索时互斥显隐）
        self.file_area = QWidget()
        file_area_layout = QVBoxLayout(self.file_area)
        file_area_layout.setContentsMargins(0, 0, 0, 0)
        file_area_layout.setSpacing(0)
        file_area_layout.addWidget(self.file_tree)
        # 结果面板（默认隐藏，命中时替换 file_tree 显隐）
        self.search_results_panel = self._build_search_results_panel()
        file_area_layout.addWidget(self.search_results_panel)
        self.search_results_panel.hide()
        self._search_thread = None
        self._search_token = 0

        right_layout.addWidget(self.file_area)
        right_layout.addWidget(self.tabs)
        right_layout.setStretch(0, 0)  # 面包屑（固定高度 26px，不参与拉伸）
        right_layout.setStretch(1, 0)  # 搜索条（固定高度 30px，不参与拉伸）
        right_layout.setStretch(2, 2)  # file_area（文件树/结果面板，主内容区）
        right_layout.setStretch(3, 1)  # tabs（预览/元数据）
        
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.addWidget(left_panel)
        self.splitter.addWidget(right_widget)
        self.splitter.setStretchFactor(1, 3)
        content_layout.addWidget(self.splitter)

        # 状态栏右侧：项目文件数/大小统计（常驻，不与 showMessage 临时消息冲突）
        self._stats_token = 0
        self._stats_thread = None
        self.folder_stats_label = QLabel('')
        self.folder_stats_label.setStyleSheet('color: #555; padding: 0 8px;')
        self.statusBar().addPermanentWidget(self.folder_stats_label)
        # 在状态栏右侧添加回收站按钮
        self.statusBar().addPermanentWidget(self._create_recycle_btn())

        self.create_menu()
        # 控件全部建好后再恢复分栏位置，否则 setSizes 会被后续布局覆盖
        self._restore_splitter_sizes()
    
    def _restore_window_geometry(self):
        """用持久化的窗口几何与最大化标志恢复主窗口。

        - 校验几何（非四元、含非正宽高、完全在屏幕外）非法则保持默认；
        - 若 window_maximized 为真：先确保有合法的"普通几何"作为取消最大化后的回弹尺寸，再 showMaximized()；
        - 任何异常都安全退化，不抛错。
        """
        geo = getattr(self, 'window_geometry', None)
        maximized = bool(getattr(self, 'window_maximized', False))
        geo_applied = False
        try:
            if isinstance(geo, (list, tuple)) and len(geo) == 4:
                x, y, w, h = (int(v) for v in geo)
                if w > 0 and h > 0:
                    # 屏幕边界检查：窗口矩形与所在屏幕的可用桌面区域必须有交集
                    rect = QRect(x, y, w, h)
                    desktop = QApplication.desktop()
                    screen_rect = desktop.availableGeometry(rect.center())
                    if screen_rect.intersects(rect):
                        self.setGeometry(rect)
                        geo_applied = True
        except (TypeError, ValueError):
            pass
        # 几何非法但要求最大化：保持当前默认几何（来自 initUI 的 setGeometry(100,100,1400,900)），
        # 这样用户取消最大化后会回到合理尺寸，而不是 Qt 内部某个微小默认。
        if maximized:
            self.showMaximized()

    def _restore_splitter_sizes(self):
        """用持久化的分栏尺寸恢复主分栏；值缺失或非法则保持默认，不抛错。"""
        sizes = getattr(self, 'splitter_sizes', None)
        try:
            if not (isinstance(sizes, (list, tuple)) and len(sizes) == 2):
                return
            sizes = [int(v) for v in sizes]
            if any(v < 0 for v in sizes) or sum(sizes) <= 0:
                return
            self.splitter.setSizes(sizes)
        except (TypeError, ValueError):
            return

    def create_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu('文件')
        file_menu.addAction('新建项目', self.new_project)
        file_menu.addAction('新建文件夹内部结构', self.new_folder_structure)
        file_menu.addAction('刷新(快捷键F5)', self.refresh_all)
        file_menu.addAction('退出', self.close)
        settings_menu = menubar.addMenu('设置')
        settings_menu.addAction('项目文件夹设置', self.show_settings_dialog)
        settings_menu.addAction('快捷访问设置', self.show_quick_access_settings_dialog)
        settings_menu.addAction('7-Zip路径设置', self.show_7zip_settings_dialog)
        settings_menu.addAction('预览设置', self.show_preview_settings_dialog)
        self.show_hidden_action = QAction('显示隐藏文件', self, checkable=True)
        self.show_hidden_action.setChecked(getattr(self, 'show_hidden', False))
        self.show_hidden_action.triggered.connect(self._on_toggle_show_hidden)
        settings_menu.addAction(self.show_hidden_action)
        settings_menu.addAction('恢复已隐藏项目', self.show_restore_hidden_projects_dialog)
        help_menu = menubar.addMenu('帮助')
        help_menu.addAction('新手向导', self.show_wizard)
        help_menu.addAction('使用帮助', self.show_help)
        help_menu.addAction('检查更新', self.check_for_updates)
        help_menu.addAction('关于', self.show_about)
    
    def _on_toggle_show_hidden(self, checked):
        self.show_hidden = checked
        self.save_settings_to_file(self.settings, self.include_subfolders)
        if self.current_folder:
            self._apply_hidden_files_filter()
            self.refresh_file_tree()

    def refresh_file_tree(self):
        """刷新文件树显示（重新应用过滤器）。"""
        if not self.current_folder:
            return
        self.file_model.setRootPath(QDir().rootPath())
        QApplication.processEvents()
        self.file_model.setRootPath(self.current_folder)
        self.file_tree.setRootIndex(self.file_model.index(self.current_folder))

    def _init_default_settings(self):
        """初始化默认设置"""
        self.project_paths = []
        self.include_subfolders = False
        self.sort_by_number = False
        self.archive_tool_path = ''
        self.enable_7zip = False
        self.default_new_project_folder = os.path.expanduser("~")
        self.folder_structure = {
            'version': '00',
            'selected_folders': {name: True for name in DEFAULT_STRUCTURE_FOLDERS},
            'custom_folders': []
        }
        self.quick_access_paths = self._get_default_quick_access_paths()
        self.pinned_folders = []
        self.hidden_folders = []
        self.show_hidden = False
        self.regex_state = 'default'
        self.custom_mb_regex = ''
        self.custom_db_regex = ''
        self.wizard_shown = False
        for key, _name in PREVIEW_CATEGORIES:
            # 视频预览默认关闭(需手动开启),避免无 OpenCV 时卡顿
            default = False if key == 'video' else True
            setattr(self, f'preview_{key}_enabled', default)
        # 窗口几何与主分栏位置（None 表示用内置默认，由 _restore_* 校验后恢复）
        self.window_geometry = None
        self.splitter_sizes = None
        self.window_maximized = False
        # 上次打开的项目绝对路径（None 表示从未选过项目）；快捷访问不会更新它
        self.last_project_path = None

    def _get_default_quick_access_paths(self):
        """获取默认快捷访问路径"""
        import ctypes
        from ctypes import wintypes
        
        default_paths = []
        
        # 获取系统特殊文件夹路径
        CSIDL_DESKTOP = 0x00
        CSIDL_MYPICTURES = 0x27
        CSIDL_DOWNLOADS = 0x28
        
        def get_special_folder(csidl):
            try:
                buf = ctypes.create_unicode_buffer(wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(0, csidl, 0, 0, buf)
                return buf.value if buf.value else None
            except Exception:
                return None
        
        # 桌面
        desktop = get_special_folder(CSIDL_DESKTOP)
        if desktop and os.path.exists(desktop):
            default_paths.append(('桌面', desktop, False))
        
        pictures = get_special_folder(CSIDL_MYPICTURES)
        if pictures and os.path.exists(pictures):
            default_paths.append(('图片', pictures, False))
        
        for letter in ['C', 'D', 'E']:
            drive_path = f'{letter}:\\'
            if os.path.exists(drive_path):
                default_paths.append((f'{letter}:', drive_path, True))
        
        return default_paths

    def load_settings(self):
        self._init_default_settings()
        if not os.path.exists(self.CONFIG_FILE):
            return self.project_paths
        try:
            with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except (json.JSONDecodeError, ValueError):
            # 配置损坏：备份后用默认值，并提示用户，避免静默丢失全部配置
            bak = self._backup_corrupt_file(self.CONFIG_FILE)
            self._pending_load_warnings.append(
                '配置文件已损坏，已恢复默认设置' + (f'（原文件备份为 {os.path.basename(bak)}）' if bak else ''))
            self._init_default_settings()
            return self.project_paths
        except Exception:
            return self.project_paths
        try:
            if 'project_paths' in config_data and config_data['project_paths']:
                if not isinstance(config_data['project_paths'], list):
                    raise TypeError('project_paths 类型错误')
                self.project_paths = config_data['project_paths']
            if 'include_subfolders' in config_data:
                self.include_subfolders = config_data['include_subfolders']
            if 'sort_by_number' in config_data:
                self.sort_by_number = config_data['sort_by_number']
            if 'default_new_project_folder' in config_data:
                self.default_new_project_folder = config_data['default_new_project_folder']
            if 'folder_structure' in config_data:
                self.folder_structure = config_data['folder_structure']
            if 'archive_tool_path' in config_data:
                self.archive_tool_path = config_data['archive_tool_path']
            enable_7zip = config_data.get('enable_7zip', False)
            self.enable_7zip = enable_7zip if isinstance(enable_7zip, bool) else False
            if 'quick_access_paths' in config_data:
                self.quick_access_paths = config_data['quick_access_paths']
            if 'pinned_folders' in config_data:
                self.pinned_folders = config_data['pinned_folders']
            if 'hidden_folders' in config_data:
                self.hidden_folders = config_data['hidden_folders']
            if 'wizard_shown' in config_data:
                self.wizard_shown = config_data['wizard_shown']
            if 'show_hidden' in config_data:
                self.show_hidden = config_data['show_hidden']
            self.regex_state = config_data.get('regex_state', 'default')
            self.custom_mb_regex = config_data.get('custom_mb_regex', '')
            self.custom_db_regex = config_data.get('custom_db_regex', '')
            for key, _name in PREVIEW_CATEGORIES:
                cfg_key = f'preview_{key}_enabled'
                if cfg_key in config_data:
                    setattr(self, cfg_key, config_data[cfg_key])
            if 'window_geometry' in config_data:
                self.window_geometry = config_data['window_geometry']
            if 'splitter_sizes' in config_data:
                self.splitter_sizes = config_data['splitter_sizes']
            if 'window_maximized' in config_data:
                self.window_maximized = bool(config_data['window_maximized'])
            if 'last_project_path' in config_data:
                self.last_project_path = config_data['last_project_path']
        except Exception:
            self._init_default_settings()
        return self.project_paths
    
    def make_file_hidden(self, file_path):
        """将文件设置为隐藏属性"""
        try:
            if sys.platform == 'win32':
                ctypes.windll.kernel32.SetFileAttributesW(file_path, FILE_ATTRIBUTE_HIDDEN)
        except Exception:
            pass

    def safe_write_json(self, file_path, data, make_hidden=True):
        """原子地写入JSON文件：先写临时文件再 os.replace 替换，避免写入中途崩溃丢失原文件"""
        tmp_path = file_path + '.tmp'
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            # 目标已存在时先解除隐藏/只读，否则 os.replace 在 Windows 上可能失败
            if os.path.exists(file_path):
                try:
                    os.chmod(file_path, 0o666)
                    if sys.platform == 'win32':
                        ctypes.windll.kernel32.SetFileAttributesW(file_path, 0x80)  # FILE_ATTRIBUTE_NORMAL
                except Exception:
                    pass
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp_path, file_path)  # 原子替换
            os.chmod(file_path, 0o644)
            if make_hidden:
                self.make_file_hidden(file_path)
            return True
        except Exception:
            # 清理残留临时文件，避免污染目录
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            return False

    def save_settings_to_file(self, paths, include_subfolders=False, default_new_project_folder=None):
        try:
            config_data = {
                'project_paths': paths,
                'include_subfolders': include_subfolders,
                'sort_by_number': getattr(self, 'sort_by_number', False),
                'archive_tool_path': getattr(self, 'archive_tool_path', ''),
                'enable_7zip': bool(getattr(self, 'enable_7zip', False)),
                'quick_access_paths': getattr(self, 'quick_access_paths', []),
                'pinned_folders': getattr(self, 'pinned_folders', []),
                'hidden_folders': getattr(self, 'hidden_folders', []),
                'wizard_shown': getattr(self, 'wizard_shown', False),
                'show_hidden': getattr(self, 'show_hidden', False),
                'regex_state': getattr(self, 'regex_state', 'default'),
                'custom_mb_regex': getattr(self, 'custom_mb_regex', ''),
                'custom_db_regex': getattr(self, 'custom_db_regex', '')
            }
            for key, _name in PREVIEW_CATEGORIES:
                cfg_key = f'preview_{key}_enabled'
                config_data[cfg_key] = getattr(self, cfg_key, True)
            config_data['window_geometry'] = getattr(self, 'window_geometry', None)
            config_data['splitter_sizes'] = getattr(self, 'splitter_sizes', None)
            config_data['window_maximized'] = bool(getattr(self, 'window_maximized', False))
            config_data['last_project_path'] = getattr(self, 'last_project_path', None)
            if hasattr(self, 'folder_structure'):
                config_data['folder_structure'] = self.folder_structure
            if default_new_project_folder:
                config_data['default_new_project_folder'] = default_new_project_folder
            elif hasattr(self, 'default_new_project_folder'):
                config_data['default_new_project_folder'] = self.default_new_project_folder
            else:
                config_data['default_new_project_folder'] = os.path.expanduser('~')
            if self.safe_write_json(self.CONFIG_FILE, config_data):
                return True
            else:
                QMessageBox.critical(self, '错误', '保存设置失败')
                return False
        except Exception as e:
            QMessageBox.critical(self, '错误', f'保存设置失败: {str(e)}')
            return False

    def _backup_corrupt_file(self, file_path):
        """将损坏的配置/注释文件改名备份为 .bak，避免被下次保存静默覆盖。返回备份路径或 None"""
        try:
            if not os.path.exists(file_path):
                return None
            bak = file_path + '.bak'
            try:
                if sys.platform == 'win32':
                    ctypes.windll.kernel32.SetFileAttributesW(file_path, 0x80)
            except Exception:
                pass
            if os.path.exists(bak):
                os.remove(bak)
            os.replace(file_path, bak)
            return bak
        except Exception:
            return None

    def load_comments(self):
        """加载项目注释"""
        if not os.path.exists(self.COMMENTS_FILE):
            return {}
        try:
            with open(self.COMMENTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, ValueError):
            # 文件损坏：备份而非静默返回 {}（否则下次保存会用空字典永久覆盖所有注释）
            bak = self._backup_corrupt_file(self.COMMENTS_FILE)
            self._pending_load_warnings.append(
                '注释文件已损坏，已忽略' + (f'并备份为 {os.path.basename(bak)}' if bak else ''))
            return {}
        except Exception:
            return {}

    def save_comments(self):
        """保存项目注释"""
        if self.safe_write_json(self.COMMENTS_FILE, self.comments):
            return True
        else:
            QMessageBox.warning(self, '警告', '保存注释失败，但不影响程序使用')
            return False
    
    def show_settings_dialog(self):
        dialog = SettingsDialog(self.settings, self.include_subfolders, self.sort_by_number, self, getattr(self, 'show_hidden', False), getattr(self, 'regex_state', 'default'), getattr(self, 'custom_mb_regex', ''), getattr(self, 'custom_db_regex', ''))
        if dialog.exec_():
            new_paths, new_include_subfolders, new_sort_by_number, new_show_hidden, new_regex_state, new_custom_mb_regex, new_custom_db_regex = dialog.get_settings()
            self.sort_by_number = new_sort_by_number
            self.show_hidden = new_show_hidden
            self.regex_state = new_regex_state
            self.custom_mb_regex = new_custom_mb_regex
            self.custom_db_regex = new_custom_db_regex
            if self.save_settings_to_file(new_paths, new_include_subfolders):
                self.settings = new_paths
                self.include_subfolders = new_include_subfolders
                QMessageBox.information(self, '成功', '设置已保存')
                self.load_filtered_folders()
                # 刷新文件树以应用隐藏文件设置
                if self.current_folder:
                    self._apply_hidden_files_filter()
                    self.refresh_file_tree()
    
    def _apply_hidden_files_filter(self):
        """根据 show_hidden 设置更新文件模型过滤器。"""
        if getattr(self, 'show_hidden', False):
            # 显示隐藏文件：只过滤 . 和 ..
            self.file_model.setFilter(QDir.NoDotAndDotDot | QDir.AllEntries | QDir.Hidden)
        else:
            # 默认：过滤 . .. 和隐藏文件
            self.file_model.setFilter(QDir.NoDotAndDotDot | QDir.AllEntries)
        
    
    def show_7zip_settings_dialog(self):
        """显示7-Zip设置对话框"""
        dialog = SevenZipSettingsDialog(
            current_path=self.archive_tool_path,
            enable_7zip=getattr(self, 'enable_7zip', False),
            parent=self,
        )
        if dialog.exec_():
            new_path, new_enable = dialog.get_settings()
            old_path = self.archive_tool_path
            old_enable = getattr(self, 'enable_7zip', False)
            self.archive_tool_path = new_path
            self.enable_7zip = new_enable
            if self.save_settings_to_file(self.settings, self.include_subfolders):
                QMessageBox.information(self, '成功', '7-Zip 设置已保存')
            else:
                self.archive_tool_path = old_path
                self.enable_7zip = old_enable

    def show_preview_settings_dialog(self):
        """显示预览开关对话框：分别控制各类文件的自动预览。"""
        dialog = QDialog(self)
        dialog.setWindowTitle('预览设置')
        layout = QVBoxLayout()
        layout.addWidget(QLabel('关闭后，点击对应类型的文件不会自动读取预览，\n而是在预览区显示一个「显示预览」按钮，需手动点击才加载。'))

        # 每个类别一个复选框，标签附带常见扩展名示例
        EXAMPLES = {
            'text': 'txt/csv/log/md 等',
            'pdf': 'pdf',
            'image': 'jpg/png/gif 等',
            'video': 'mp4/avi/mov 等',
            'archive': 'zip/rar/7z',
            'excel': 'xlsx/xlsm/xls',
            'word': 'docx/doc',
        }
        checkboxes = {}
        for key, name in PREVIEW_CATEGORIES:
            example = EXAMPLES.get(key, '')
            cb = QCheckBox(f'{name}预览（{example}）' if example else f'{name}预览')
            cb.setChecked(getattr(self, f'preview_{key}_enabled', True))
            layout.addWidget(cb)
            checkboxes[key] = cb

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton('确定')
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)
        if dialog.exec_():
            for key, cb in checkboxes.items():
                setattr(self, f'preview_{key}_enabled', cb.isChecked())
            self.save_settings_to_file(self.settings, self.include_subfolders)

    def _create_quick_access_buttons(self, layout):
        for btn in list(self.quick_access_buttons):
            layout.removeWidget(btn)
            btn.deleteLater()
        self.quick_access_buttons.clear()

        insert_index = layout.indexOf(self.quick_access_add_btn) + 1
        if insert_index <= 0:
            insert_index = layout.count()

        for item in self.quick_access_paths:
            if len(item) == 3:
                name, path, no_preview = item
            else:
                name, path = item
                no_preview = False
            item_data = (name, path, no_preview)
            btn = QPushButton(name)
            btn.setToolTip(path)
            btn.setFixedHeight(22)
            btn.setMinimumWidth(btn.fontMetrics().width('000000') + 16)
            btn.setContextMenuPolicy(Qt.CustomContextMenu)
            btn.customContextMenuRequested.connect(
                lambda pos, b=btn, data=item_data: self._show_quick_access_context_menu(b, data, pos)
            )
            # 单击/双击分开处理：单击用延迟定时器(双击可取消)，双击直接打开资源管理器
            # 避免 QPushButton.clicked 在第一次点击 release 即触发导致双击也会跑一次单击
            _single_shot = QTimer(btn)  # parent=btn,随按钮一起释放
            _single_shot.setSingleShot(True)
            _single_shot.setInterval(250)

            if no_preview:
                btn.setStyleSheet(
                    "QPushButton { background-color: #e8e8e8; border: 1px solid #bbb; border-radius: 3px; "
                    "color: #555; font-style: italic; }"
                    "QPushButton:hover { background-color: #d8d8d8; }"
                )
                # "不预览"按钮:无论单双击,只打开一次资源管理器
                _opened = {'done': False}
                def _open_external_once(p=path, _flag=_opened):
                    if _flag['done']:
                        return  # 已打开,忽略重复触发
                    _flag['done'] = True
                    self._open_quick_access_external(p)
                    # 250ms 后重置标志,允许下次点击
                    QTimer.singleShot(300, lambda: _flag.update(done=False))
                btn.clicked.connect(lambda checked, fn=_open_external_once: fn())
            else:
                _single_shot.timeout.connect(lambda p=path: self._open_quick_access_path(p))
                # clicked 只启动定时器；若 250ms 内发生双击,_on_double_click 会 stop 掉它
                btn.clicked.connect(lambda checked, t=_single_shot: t.start())
                def _on_double_click(event, p=path, t=_single_shot):
                    # 取消即将触发的单击动作
                    if t.isActive():
                        t.stop()
                    self._open_with_shell(p)
                    QPushButton.mouseDoubleClickEvent(btn, event)
                btn.mouseDoubleClickEvent = _on_double_click

            layout.insertWidget(insert_index, btn)
            insert_index += 1
            self.quick_access_buttons.append(btn)

    def _show_quick_access_context_menu(self, button, item_data, pos):
        """快捷访问按钮右键菜单。"""
        name, path, no_preview = item_data
        menu = QMenu(self)
        terminal_action = menu.addAction('在终端中打开')
        delete_action = menu.addAction('删除快捷访问')
        menu.addSeparator()
        settings_action = menu.addAction('打开快捷访问设置')
        chosen = menu.exec_(button.mapToGlobal(pos))
        if chosen == terminal_action:
            self.open_folder_in_terminal(path)
        elif chosen == delete_action:
            reply = QMessageBox.question(
                self,
                '确认删除',
                f'确定要删除快捷访问“{name}”吗？\n{path}',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                self._remove_quick_access_item(item_data)
        elif chosen == settings_action:
            self.show_quick_access_settings_dialog()

    def _open_quick_access_external(self, path):
        try:
            path = os.path.normpath(path)
            if os.path.exists(path):
                os.startfile(path)
            else:
                QMessageBox.warning(self, '警告', f'路径不存在: {path}')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'无法打开文件夹: {str(e)}')
    
    def _open_quick_access_path(self, path):
        # os.path.exists() 对 UNC/网络路径不可靠,直接尝试加载;失败时 QFileSystemModel 不会显示内容
        self.current_folder = path
        self.file_model.setRootPath(path)
        self.file_tree.setRootIndex(self.file_model.index(path))
        self.new_structure_btn.setEnabled(False)
        self._update_folder_status_bar()
        # 面包屑焦点回到快捷访问根
        self._breadcrumb_path = path
        self._rebuild_breadcrumb()
        self._close_search_results()
        self._set_search_bar_enabled(True)

    def _get_selected_file_paths(self):
        selection_model = self.file_tree.selectionModel()
        if not selection_model:
            return []

        paths = []
        seen = set()
        for index in selection_model.selectedRows(0):
            file_path = self.file_model.filePath(index)
            normalized_path = os.path.normpath(file_path)
            if os.path.exists(file_path) and normalized_path not in seen:
                seen.add(normalized_path)
                paths.append(file_path)
        return paths

    def _get_primary_selected_file_path(self):
        selected_paths = self._get_selected_file_paths()
        if selected_paths:
            return selected_paths[0]

        current_index = self.file_tree.currentIndex()
        if current_index.isValid():
            file_path = self.file_model.filePath(current_index)
            if os.path.exists(file_path):
                return file_path
        return None

    def _select_single_file_index(self, index):
        if not index.isValid():
            return

        selection_model = self.file_tree.selectionModel()
        if not selection_model:
            return

        selection_model.clearSelection()
        self.file_tree.setCurrentIndex(index)
        selection_model.select(index, selection_model.Select | selection_model.Rows)

    def _copy_paths_to_clipboard(self, file_paths):
        valid_paths = []
        seen = set()
        for file_path in file_paths:
            normalized_path = os.path.normpath(file_path)
            if os.path.exists(file_path) and normalized_path not in seen:
                seen.add(normalized_path)
                valid_paths.append(file_path)

        if not valid_paths:
            QMessageBox.warning(self, '警告', '没有可复制的文件或文件夹')
            return False

        mime_data = QMimeData()
        mime_data.setUrls([QUrl.fromLocalFile(path) for path in valid_paths])
        mime_data.setText('\n'.join(valid_paths))
        QApplication.clipboard().setMimeData(mime_data)

        self.clipboard_paths = list(valid_paths)
        if len(valid_paths) == 1:
            self.clipboard_path = valid_paths[0]
            self.statusBar().showMessage(f"已复制，可在资源管理器中粘贴: {os.path.basename(valid_paths[0])}")
        else:
            self.clipboard_path = None
            self.statusBar().showMessage(f"已复制 {len(valid_paths)} 个项目，可在资源管理器中粘贴")
        return True

    def _copy_path_to_clipboard(self, file_path):
        """复制文件到系统剪贴板，并保留程序内粘贴副本功能"""
        return self._copy_paths_to_clipboard([file_path])

    def copy_selected_items(self):
        selected_paths = self._get_selected_file_paths()
        if selected_paths:
            return self._copy_paths_to_clipboard(selected_paths)

        selected_path = self._get_primary_selected_file_path()
        if selected_path:
            return self._copy_path_to_clipboard(selected_path)
        return False

    def paste_to_selected_target(self):
        target = self.current_folder
        selected_paths = self._get_selected_file_paths()
        if len(selected_paths) == 1:
            selected_path = selected_paths[0]
            if os.path.isdir(selected_path):
                target = selected_path
            else:
                target = os.path.dirname(selected_path)
        elif len(selected_paths) > 1:
            self.statusBar().showMessage("已选择多个项目，将粘贴到当前项目文件夹")
        else:
            selected_path = self._get_primary_selected_file_path()
            if selected_path:
                if os.path.isdir(selected_path):
                    target = selected_path
                else:
                    target = os.path.dirname(selected_path)

        if target and os.path.exists(target):
            self.paste_copy(target)
            return True
        return False

    def _create_unique_zip_path(self, parent_dir, base_name):
        zip_name = base_name + '.zip'
        zip_path = os.path.join(parent_dir, zip_name)
        counter = 1
        while os.path.exists(zip_path):
            zip_name = f"{base_name} ({counter}).zip"
            zip_path = os.path.join(parent_dir, zip_name)
            counter += 1
        return zip_path, zip_name

    def _write_path_to_zip(self, zf, source_path, base_dir):
        if os.path.isfile(source_path):
            zf.write(source_path, os.path.relpath(source_path, base_dir))
            return

        folder_arcname = os.path.relpath(source_path, base_dir).replace('\\', '/') + '/'
        zf.writestr(folder_arcname, '')
        for root, dirs, files in os.walk(source_path):
            rel_root = os.path.relpath(root, base_dir)
            if rel_root != '.':
                dir_arcname = rel_root.replace('\\', '/') + '/'
                zf.writestr(dir_arcname, '')
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                dir_arcname = os.path.relpath(dir_path, base_dir).replace('\\', '/') + '/'
                zf.writestr(dir_arcname, '')
            for file_name in files:
                file_path = os.path.join(root, file_name)
                arcname = os.path.relpath(file_path, base_dir)
                zf.write(file_path, arcname)

    def _move_paths_to_recycle(self, file_paths):
        valid_paths = []
        seen = set()
        for file_path in file_paths:
            absolute_path = os.path.abspath(file_path)
            normalized_path = os.path.normcase(os.path.normpath(absolute_path))
            if os.path.exists(absolute_path) and normalized_path not in seen:
                seen.add(normalized_path)
                valid_paths.append(absolute_path)

        # 同时选中父目录和其子项时只处理父目录，避免父目录已回收后子项被误报失败。
        filtered_paths = []
        for path in valid_paths:
            nested = False
            for other in valid_paths:
                if path == other:
                    continue
                try:
                    if os.path.commonpath([path, other]) == other:
                        nested = True
                        break
                except ValueError:
                    continue
            if not nested:
                filtered_paths.append(path)
        valid_paths = filtered_paths

        if not valid_paths:
            QMessageBox.warning(self, '警告', '没有可移入回收站的文件或文件夹')
            return False

        if len(valid_paths) == 1:
            prompt = (
                f'确定将以下项目移入回收站？\n\n{os.path.basename(valid_paths[0])}'
                '\n\n若该位置不支持回收站，操作会失败并保留原文件。'
            )
        else:
            names = '\n'.join('· ' + os.path.basename(p) for p in valid_paths[:10])
            if len(valid_paths) > 10:
                names += f'\n… 等共 {len(valid_paths)} 个项目'
            prompt = (
                f'确定将以下 {len(valid_paths)} 个项目移入回收站？\n\n{names}'
                '\n\n若某个位置不支持回收站，该项目会保留并单独报告失败。'
            )
        if QMessageBox.question(self, '确认删除', prompt,
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return False

        try:
            backend = _load_strict_recycle_backend()
        except OSError as e:
            QMessageBox.warning(
                self,
                '错误',
                f'无法启用安全回收站功能：{_format_os_error(e)}',
            )
            return False
        except Exception as e:
            QMessageBox.warning(self, '错误', f'无法启用安全回收站功能：{str(e)}')
            return False

        successes = []
        failures = []
        for path in valid_paths:
            try:
                _send_path_to_recycle_strict(path, backend=backend)
                successes.append(path)
            except OSError as e:
                failures.append((path, _format_os_error(e)))
            except Exception as e:
                failures.append((path, str(e)))

        if failures:
            details = '\n'.join(
                f'· {os.path.basename(path)}：{message or "无法进入回收站"}'
                for path, message in failures[:10]
            )
            if len(failures) > 10:
                details += f'\n… 另有 {len(failures) - 10} 项失败'
            QMessageBox.warning(
                self,
                '部分项目未删除' if successes else '移入回收站失败',
                '以下项目未能进入回收站，程序没有执行永久删除：\n\n' + details,
            )

        if successes:
            if len(successes) == 1:
                self.statusBar().showMessage(
                    f'已移入回收站: {os.path.basename(successes[0])}'
                )
            else:
                self.statusBar().showMessage(f'已移入回收站: {len(successes)} 个项目')
        return bool(successes)

    def _create_recycle_btn(self):
        """创建回收站按钮"""
        btn = QPushButton('🗑')
        btn.setFixedSize(30, 22)
        btn.setToolTip('打开回收站')
        btn.clicked.connect(self.open_recycle_bin)
        return btn
    
    def _remove_quick_access_item(self, item_data):
        """删除一个快捷访问项，按名称、路径和预览模式精确匹配。"""
        target_name, target_path, target_no_preview = item_data
        target_path = os.path.normcase(os.path.normpath(target_path))
        kept_paths = []
        removed = False
        for item in self.quick_access_paths:
            if len(item) == 3:
                name, path, no_preview = item
            else:
                name, path = item
                no_preview = False
            item_path = os.path.normcase(os.path.normpath(path))
            if (
                not removed
                and name == target_name
                and item_path == target_path
                and bool(no_preview) == bool(target_no_preview)
            ):
                removed = True
                continue
            kept_paths.append(item)

        if not removed:
            QMessageBox.warning(self, '警告', '未找到该快捷访问项')
            return False

        self.quick_access_paths = kept_paths
        self.save_settings_to_file(self.settings, self.include_subfolders)
        self._refresh_quick_access_toolbar()
        self.statusBar().showMessage('已删除快捷访问')
        return True

    def _refresh_quick_access_toolbar(self):
        """刷新快捷访问工具栏，保留标题、加号和分隔空间。"""
        quick_access_layout = self.quick_access_toolbar.layout()
        for btn in list(self.quick_access_buttons):
            quick_access_layout.removeWidget(btn)
            btn.deleteLater()
        self.quick_access_buttons.clear()
        self._create_quick_access_buttons(quick_access_layout)

    def show_quick_access_settings_dialog(self):
        """显示快捷访问设置对话框"""
        dialog = QuickAccessSettingsDialog(self.quick_access_paths, self)
        if dialog.exec_():
            new_paths = dialog.get_settings()
            self.quick_access_paths = new_paths
            # 保存设置
            self.save_settings_to_file(self.settings, self.include_subfolders)
            # 更新工具栏
            self._refresh_quick_access_toolbar()
            QMessageBox.information(self, '成功', '快捷访问设置已保存')
    
    def open_recycle_bin(self):
        """打开回收站"""
        try:
            os.startfile('shell:RecycleBinFolder')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'无法打开回收站: {str(e)}')
    
    def load_filtered_folders_async(self):
        """异步加载过滤后的文件夹"""
        # 若已有扫描线程在运行，先停止旧线程并断开其信号，避免旧结果回填到新一轮扫描，
        # 以及 QThread 仍在运行时被覆盖销毁触发警告
        old = getattr(self, 'scan_thread', None)
        if old is not None and old.isRunning():
            try:
                old.scan_completed.disconnect(self.on_scan_completed)
                old.scan_progress.disconnect(self.on_scan_progress)
            except (TypeError, RuntimeError):
                pass
            old.requestInterruption()
            old.quit()
            old.wait(3000)
            old.deleteLater()

        # 清空表格
        self.motherboard_table.setRowCount(0)
        self.daughterboard_table.setRowCount(0)
        self.filtered_folders = {'主板': [], '子卡': []}

        # 在状态栏显示加载信息
        self.statusBar().showMessage("正在扫描文件夹...")

        # 解析正则（自定义正则无效时自动回退到默认）
        self._regex_fallback = False
        mb_regex, mb_fallback = _resolve_regex(self.regex_state, self.custom_mb_regex, DEFAULT_MB_RE)
        db_regex, db_fallback = _resolve_regex(self.regex_state, self.custom_db_regex, DEFAULT_DB_RE)
        self._regex_fallback = mb_fallback or db_fallback
        self.folder_regex_mb = mb_regex
        self.folder_regex_db = db_regex
        # 创建并启动扫描线程
        self.scan_thread = FolderScanThread(self.settings, self.include_subfolders, self.comments, self.sort_by_number, mb_regex=self.folder_regex_mb, db_regex=self.folder_regex_db)
        self.scan_thread.scan_completed.connect(self.on_scan_completed)
        self.scan_thread.scan_progress.connect(self.on_scan_progress)
        self.scan_thread.start()

    def closeEvent(self, event):
        """退出时确保后台线程已结束，避免 QThread 被销毁时仍在运行。
        统一处理:disconnect 信号 → requestInterruption → quit → wait → deleteLater。"""
        running_extracts = [
            thread for thread in getattr(self, '_extract_jobs', {}) if thread.isRunning()
        ]
        if running_extracts:
            self._close_after_extract_cancel = True
            for thread in running_extracts:
                thread.cancel()
            self.statusBar().showMessage('正在取消解压并清理临时文件，完成后将自动退出...')
            event.ignore()
            return

        # 下载线程单独处理(结构不同,无 _signal_map 中的信号)
        update_thread = getattr(self, 'update_download_thread', None)
        if update_thread is not None:
            try:
                for sig in ('progress_changed', 'status_changed', 'download_completed', 'download_failed', 'download_canceled'):
                    try: getattr(update_thread, sig).disconnect()
                    except (TypeError, RuntimeError): pass
            except Exception: pass
            try:
                if update_thread.isRunning():
                    update_thread.requestInterruption()
                    update_thread.quit()
                    update_thread.wait(5000)
            except Exception: pass
            update_thread.deleteLater()
            self.update_download_thread = None
        # 其余线程统一处理
        _signal_map = {
            '_stats_thread': ('stats_ready',),
            '_search_thread': ('search_ready',),
            'scan_thread': ('scan_completed', 'scan_progress'),
        }
        for attr in ('_stats_thread', '_search_thread', 'scan_thread'):
            t = getattr(self, attr, None)
            if t is not None:
                try:
                    for sig_name in _signal_map.get(attr, ()):
                        sig = getattr(t, sig_name, None)
                        if sig is not None:
                            try: sig.disconnect()
                            except (TypeError, RuntimeError): pass
                    if t.isRunning():
                        t.requestInterruption()
                        t.quit()
                        t.wait(2000)
                    t.deleteLater()
                except Exception:
                    pass
                if attr == 'scan_thread':
                    self.scan_thread = None
                else:
                    setattr(self, attr, None)
        # 记住窗口几何/最大化标志/主分栏位置，保存失败绝不阻塞关闭
        try:
            # 最小化时先还原，避免 normalGeometry 未覆盖的极端退化抓到极小化坐标
            if self.isMinimized():
                self.showNormal()
            # normalGeometry() 返回非最大化时的几何；若窗口从未被最大化过，某些平台返回 0 尺寸 → 回退到 geometry()
            ngeo = self.normalGeometry()
            if ngeo.width() <= 0 or ngeo.height() <= 0:
                ngeo = self.geometry()
            self.window_geometry = [ngeo.x(), ngeo.y(), ngeo.width(), ngeo.height()]
            self.window_maximized = bool(self.isMaximized())
            if hasattr(self, 'splitter'):
                self.splitter_sizes = list(self.splitter.sizes())
            self.save_settings_to_file(self.settings, self.include_subfolders)
        except Exception:
            pass
        super().closeEvent(event)

    def changeEvent(self, event):
        """禁止全屏：任何走向全屏的状态变更（F11、外部 API 调用等）都立即拉回普通/最大化态。"""
        try:
            if event.type() == QEvent.WindowStateChange and self.isFullScreen():
                # 清掉全屏位，保留其它状态位（如最大化）
                self.setWindowState(self.windowState() & ~Qt.WindowFullScreen)
                try:
                    self.statusBar().showMessage('已禁用全屏模式', 3000)
                except Exception:
                    pass
        except Exception:
            pass
        super().changeEvent(event)

    def on_scan_progress(self, message):
        """处理扫描进度更新"""
        self.statusBar().showMessage(message)
    
    def on_scan_completed(self, motherboard_folders, daughterboard_folders):
        """处理扫描完成信号"""
        # 清空现有表格数据
        self.motherboard_table.setRowCount(0)
        self.daughterboard_table.setRowCount(0)
        self.filtered_folders = {'主板': [], '子卡': []}
        
        visible_motherboard_folders = [folder for folder in motherboard_folders if folder[1] not in self.hidden_folders]
        visible_daughterboard_folders = [folder for folder in daughterboard_folders if folder[1] not in self.hidden_folders]

        # 填充主板表格
        for folder in visible_motherboard_folders:
            row_position = self.motherboard_table.rowCount()
            self.motherboard_table.insertRow(row_position)
            number_item = QTableWidgetItem(folder[2])
            number_item.setData(Qt.UserRole, folder[1])
            number_item.setData(Qt.UserRole + 1, folder[4])
            self.motherboard_table.setItem(row_position, 0, number_item)
            comment_item = QTableWidgetItem(folder[3])
            self.motherboard_table.setItem(row_position, 1, comment_item)
            self.filtered_folders['主板'].append(FolderInfo(*folder))
        
        # 填充子卡表格
        for folder in visible_daughterboard_folders:
            row_position = self.daughterboard_table.rowCount()
            self.daughterboard_table.insertRow(row_position)
            number_item = QTableWidgetItem(folder[2])
            number_item.setData(Qt.UserRole, folder[1])
            number_item.setData(Qt.UserRole + 1, folder[4])
            self.daughterboard_table.setItem(row_position, 0, number_item)
            comment_item = QTableWidgetItem(folder[3])
            self.daughterboard_table.setItem(row_position, 1, comment_item)
            self.filtered_folders['子卡'].append(FolderInfo(*folder))
        
        hidden_count = len(motherboard_folders) + len(daughterboard_folders) - len(visible_motherboard_folders) - len(visible_daughterboard_folders)
        message = f"共找到 {len(visible_motherboard_folders)} 个主板文件夹，{len(visible_daughterboard_folders)} 个子卡文件夹"
        if hidden_count:
            message += f"，已隐藏 {hidden_count} 个项目"
        self.statusBar().showMessage(message)

        if self.pinned_folders:
            self._apply_pin_order(self.motherboard_table)
            self._apply_pin_order(self.daughterboard_table)

        # 首次扫描完成后尝试恢复上次打开的项目（仅一次，刷新不再触发）
        self._restore_last_project()

    def _restore_last_project(self):
        """扫描完成后恢复上次打开的项目。仅在首次扫描、且用户未抢先手动选行时执行。
        路径不存在或命名不再符合 S/M 格式则清掉脏数据；表里找不到（被隐藏/根路径配置变了）则静默保留。"""
        target = getattr(self, '_pending_restore_project', None)
        self._pending_restore_project = None
        if not target:
            return
        # 用户在扫描期间已手动选了别的行 → 不抢用户选择
        if getattr(self, 'current_folder', None) is not None:
            return
        try:
            if not os.path.isdir(target):
                self.last_project_path = None
                return
            name = os.path.basename(target)
            match = self.folder_regex_mb.match(name) or self.folder_regex_db.match(name)
            if not match:
                self.last_project_path = None
                return
            table = self.motherboard_table if match.group(1) == 'S' else self.daughterboard_table
            for row in range(table.rowCount()):
                if table.item(row, 0).data(Qt.UserRole) == target:
                    table.selectRow(row)
                    table.scrollToItem(table.item(row, 0))
                    self._select_project_path(target)
                    # 恢复也视作一次"选中"，刷新 last_project_path 保持一致
                    self.last_project_path = target
                    return
            # 表里没找到（被隐藏 / 根路径配置变了 / 未被扫到）：静默，不清字段
        except Exception:
            pass

    def load_filtered_folders(self):
        """同步加载过滤后的文件夹（保留接口兼容）"""
        self.load_filtered_folders_async()
    

    def refresh_all(self):
        """刷新项目列表和文件树（菜单/F5 共用）。"""
        self.load_filtered_folders()
        if not self.current_folder:
            return
        # 强制 QFileSystemModel 释放并重新获取目录句柄：先切到根目录，泵事件，再切回
        self.file_model.setRootPath(QDir().rootPath())
        QApplication.processEvents()
        self.file_model.setRootPath(self.current_folder)
        self.file_tree.setRootIndex(self.file_model.index(self.current_folder))
        QApplication.processEvents()
        self.file_tree.collapseAll()
    def filter_folders(self, text):
        for row in range(self.motherboard_table.rowCount()):
            number = self.motherboard_table.item(row, 0).text()
            comment = self.motherboard_table.item(row, 1).text()
            show = text.lower() in number.lower() or text.lower() in comment.lower()
            self.motherboard_table.setRowHidden(row, not show)
        for row in range(self.daughterboard_table.rowCount()):
            number = self.daughterboard_table.item(row, 0).text()
            comment = self.daughterboard_table.item(row, 1).text()
            show = text.lower() in number.lower() or text.lower() in comment.lower()
            self.daughterboard_table.setRowHidden(row, not show)
    
    def _get_effective_folder_comment(self, folder_path):
        if folder_path in self.comments:
            return self.comments[folder_path]
        folder_name = os.path.basename(folder_path)
        match = self.folder_regex_mb.match(folder_name) or self.folder_regex_db.match(folder_name)
        if match:
            return match.group(2) if match.group(2) else ''
        return ''

    def _show_folder_context_menu(self, table, pos):
        row = table.rowAt(pos.y())
        if row < 0:
            return
        folder_path = table.item(row, 0).data(Qt.UserRole)
        if not folder_path:
            return
        menu = QMenu(self)
        if folder_path in self.pinned_folders:
            pin_action = menu.addAction('取消置顶')
        else:
            pin_action = menu.addAction('置顶')
        terminal_action = menu.addAction('在终端中打开')
        menu.addSeparator()
        hide_action = menu.addAction('隐藏项目')
        action_pos = table.viewport().mapToGlobal(pos)
        chosen = menu.exec_(action_pos)
        if chosen == pin_action:
            if folder_path in self.pinned_folders:
                self.pinned_folders.remove(folder_path)
            else:
                self.pinned_folders.append(folder_path)
            self._apply_pin_order(table)
            self.save_settings_to_file(self.settings, self.include_subfolders)
        elif chosen == terminal_action:
            self.open_folder_in_terminal(folder_path)
        elif chosen == hide_action:
            if folder_path not in self.hidden_folders:
                self.hidden_folders.append(folder_path)
            if folder_path in self.pinned_folders:
                self.pinned_folders.remove(folder_path)
            if folder_path == self.current_folder:
                self.current_folder = None
                self.file_model.setRootPath('')
                self.file_tree.setRootIndex(QModelIndex())
                self._reset_preview()
                self.metadata_tab.clear()
                self.new_structure_btn.setEnabled(False)
                # 清空面包屑
                self._breadcrumb_path = None
                self._rebuild_breadcrumb()
                # 关闭文件搜索结果面板 + 禁用搜索条
                self._set_search_bar_enabled(False)
                # 清空文件统计 + 取消在跑的统计线程
                self._stats_token += 1
                self.folder_stats_label.setText('')
                stats_t = getattr(self, '_stats_thread', None)
                if stats_t is not None and stats_t.isRunning():
                    stats_t.requestInterruption()
                    stats_t.quit()
            # 同步清空「上次项目」记录，避免下次恢复指向已隐藏/删除的项目
            if getattr(self, 'last_project_path', None) == folder_path:
                self.last_project_path = None
            self.save_settings_to_file(self.settings, self.include_subfolders)
            self.load_filtered_folders()

    def show_restore_hidden_projects_dialog(self):
        """恢复已隐藏的项目"""
        hidden_folders = [path for path in self.hidden_folders if os.path.exists(path)]
        missing_folders = [path for path in self.hidden_folders if not os.path.exists(path)]
        if missing_folders:
            self.hidden_folders = hidden_folders
            self.save_settings_to_file(self.settings, self.include_subfolders)

        if not hidden_folders:
            QMessageBox.information(self, '提示', '当前没有已隐藏项目')
            return

        labels = [f'{os.path.basename(path)}    {path}' for path in hidden_folders]
        label, ok = QInputDialog.getItem(self, '恢复已隐藏项目', '请选择要恢复的项目：', labels, 0, False)
        if not ok or not label:
            return

        index = labels.index(label)
        restored_path = hidden_folders[index]
        self.hidden_folders.remove(restored_path)
        self.save_settings_to_file(self.settings, self.include_subfolders)
        self.load_filtered_folders()
        self.locate_new_folder(restored_path)
        self.statusBar().showMessage(f"已恢复项目: {os.path.basename(restored_path)}")

    def _apply_pin_order(self, table):
        pinned_rows = []
        normal_rows = []
        for row in range(table.rowCount()):
            folder_path = table.item(row, 0).data(Qt.UserRole)
            if folder_path in self.pinned_folders:
                pinned_rows.append(row)
            else:
                normal_rows.append(row)
        pinned_rows.sort(key=lambda r: self.pinned_folders.index(table.item(r, 0).data(Qt.UserRole)), reverse=True)
        new_order = pinned_rows + normal_rows
        items_data = []
        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.takeItem(row, col)
                row_data.append(item)
            items_data.append(row_data)
        table.setRowCount(0)
        for row_idx in new_order:
            row_position = table.rowCount()
            table.insertRow(row_position)
            for col, item in enumerate(items_data[row_idx]):
                if item:
                    table.setItem(row_position, col, item)
        for row in range(table.rowCount()):
            folder_path = table.item(row, 0).data(Qt.UserRole)
            font = table.item(row, 0).font()
            if folder_path in self.pinned_folders:
                font.setBold(True)
                table.item(row, 0).setFont(font)
                table.item(row, 1).setFont(font)
            else:
                font.setBold(False)
                table.item(row, 0).setFont(font)
                table.item(row, 1).setFont(font)

    def _build_breadcrumb_bar(self):
        """构建面包屑容器（横向布局，单行，末尾留弹性空间）。段在 _rebuild_breadcrumb 里动态填充。"""
        bar = QWidget()
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(4, 2, 4, 2)
        layout.setSpacing(2)
        layout.addStretch()  # 占位，段插在它前面
        bar.setFixedHeight(26)
        self._breadcrumb_layout = layout
        return bar

    def _clear_breadcrumb(self):
        """移除所有已建的段控件（按钮 + 分隔符），保留末尾的 stretch。"""
        for w in self._breadcrumb_buttons:
            self._breadcrumb_layout.removeWidget(w)
            w.deleteLater()
        self._breadcrumb_buttons = []

    def _rebuild_breadcrumb(self):
        """按 current_folder（根）+ _breadcrumb_path（焦点）重绘面包屑。
        段：[根 basename, rel 各级]；中间段可点，末段不可点。超长则中间省略。"""
        self._clear_breadcrumb()
        root = getattr(self, 'current_folder', None)
        if not root:
            return
        focus = self._breadcrumb_path or root
        try:
            rel = os.path.relpath(focus, root)
        except (ValueError, TypeError):
            rel = '.'
        # focus 不在 root 之下：退化为只显示根
        if rel == '.' or rel.startswith('..'):
            parts = []
        else:
            parts = [p for p in rel.split(os.sep) if p and p != '.']
        # 构建 (路径, 显示名) 元组列表：根 + 逐级
        segments = [(root, os.path.basename(root) or root)]
        acc = root
        for p in parts:
            acc = os.path.join(acc, p)
            segments.append((acc, p))

        # 超长省略：保留首段 + 末两段，中间塞不可点的 …
        MAX_SEGS = 5
        if len(segments) > MAX_SEGS:
            segments = segments[:1] + [(None, '…')] + segments[-2:]

        insert_at = self._breadcrumb_layout.count() - 1  # stretch 之前
        last_idx = len(segments) - 1
        for i, (path, name) in enumerate(segments):
            if i > 0:
                sep = QLabel('›')
                sep.setStyleSheet('color: #999;')
                self._breadcrumb_layout.insertWidget(insert_at, sep)
                insert_at += 1
                self._breadcrumb_buttons.append(sep)
            is_last = (i == last_idx)
            is_ellipsis = (path is None)
            if is_last or is_ellipsis:
                lbl = QLabel(name)
                if is_last:
                    lbl.setStyleSheet('font-weight: bold; color: #2c3e50; padding: 0 4px;')
                else:
                    lbl.setStyleSheet('color: #999; padding: 0 2px;')
                self._breadcrumb_layout.insertWidget(insert_at, lbl)
                insert_at += 1
                self._breadcrumb_buttons.append(lbl)
            else:
                btn = QToolButton()
                btn.setText(name)
                btn.setAutoRaise(True)
                btn.setCursor(Qt.PointingHandCursor)
                btn.setStyleSheet(
                    'QToolButton { border: none; color: #2575c0; padding: 0 4px; }'
                    'QToolButton:hover { color: #1a4d80; text-decoration: underline; }'
                )
                btn.clicked.connect(lambda checked, p=path: self._on_breadcrumb_clicked(p))
                self._breadcrumb_layout.insertWidget(insert_at, btn)
                insert_at += 1
                self._breadcrumb_buttons.append(btn)

    def _on_breadcrumb_clicked(self, path):
        """点击中间段：在文件树里选中并滚动到该目录，不改树根。"""
        root = getattr(self, 'current_folder', None)
        if not (root and path and os.path.isdir(path)):
            self._rebuild_breadcrumb()  # 路径已失效，按当前状态重绘
            return
        # 必须在 root 之下
        try:
            rel = os.path.relpath(path, root)
        except (ValueError, TypeError):
            return
        if rel.startswith('..'):
            return
        idx = self.file_model.index(path)
        if idx.isValid():
            self.file_tree.setCurrentIndex(idx)
            self.file_tree.scrollTo(idx)
            self.file_tree.expand(idx)
        self._reset_preview()
        self._breadcrumb_path = path
        self._rebuild_breadcrumb()

    # ---- 文件搜索（文件名 + 类型 + 日期）----

    # 类型下拉：显示名 -> 扩展名集合（None 表示不限）
    _SEARCH_TYPE_EXTS = None  # 占位，运行时如下表
    _TYPE_OPTIONS = None

    def _init_search_type_options(self):
        if self.__class__._TYPE_OPTIONS is not None:
            return
        opts = [  # (显示名, 扩展名集合或None)
            ('全部', None),
            ('文本', TEXT_EXTS),
            ('PDF', ('.pdf',)),
            ('图片', IMAGE_EXTS),
            ('视频', VIDEO_EXTS),
            ('压缩包', ARCHIVE_EXTS),
            ('表格', ('.xlsx', '.xlsm', '.xls')),
            ('文档', ('.docx', '.doc')),
        ]
        self.__class__._TYPE_OPTIONS = opts

    def _build_search_bar(self, parent_layout):
        """构建文件搜索条：关键词 + 类型下拉 + 日期下拉 + 搜索/关闭按钮。"""
        self._init_search_type_options()
        bar = QWidget()
        h = QHBoxLayout(bar)
        h.setContentsMargins(4, 2, 4, 2)
        h.setSpacing(4)
        h.addWidget(QLabel('文件搜索:'))
        self.search_name_edit = QLineEdit()
        self.search_name_edit.setPlaceholderText('文件名关键词（留空=不限）')
        self.search_name_edit.returnPressed.connect(self._start_file_search)
        h.addWidget(self.search_name_edit, 1)
        self.search_type_combo = QComboBox()
        for name, _exts in self._TYPE_OPTIONS:
            self.search_type_combo.addItem(name)
        h.addWidget(self.search_type_combo)
        self.search_date_combo = QComboBox()
        for label in ('不限', '近7天', '近30天', '近半年', '近一年'):
            self.search_date_combo.addItem(label)
        h.addWidget(self.search_date_combo)
        self.search_btn = QPushButton('搜索')
        self.search_btn.clicked.connect(self._start_file_search)
        h.addWidget(self.search_btn)
        bar.setFixedHeight(30)
        self.search_bar = bar
        # 未选项目前禁用
        for w in (self.search_name_edit, self.search_type_combo, self.search_date_combo, self.search_btn):
            w.setEnabled(False)
        parent_layout.addWidget(bar)

    def _build_search_results_panel(self):
        """搜索结果面板：顶部状态行 + 结果列表。默认隐藏。"""
        panel = QWidget()
        v = QVBoxLayout(panel)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(2)
        head = QHBoxLayout()
        head.setContentsMargins(4, 2, 4, 2)
        self.search_status_label = QLabel('')
        self.search_status_label.setStyleSheet('color: #555;')
        head.addWidget(self.search_status_label, 1)
        close_btn = QPushButton('关闭')
        close_btn.clicked.connect(self._close_search_results)
        head.addWidget(close_btn)
        v.addLayout(head)
        self.search_results_list = QListWidget()
        self.search_results_list.itemDoubleClicked.connect(self._on_search_result_double_clicked)
        v.addWidget(self.search_results_list, 1)
        return panel

    def _search_exts_for_combo(self):
        idx = self.search_type_combo.currentIndex()
        if idx < 0 or idx >= len(self._TYPE_OPTIONS):
            return None
        return self._TYPE_OPTIONS[idx][1]

    def _search_mtime_after(self):
        """返回时间戳下限或 None。now 用 self._search_now_ts（可被测试注入）。"""
        import time as _time
        text = self.search_date_combo.currentText()
        if text == '不限' or not text:
            return None
        now = getattr(self, '_search_now_ts', None) or _time.time()
        days = {'近7天': 7, '近30天': 30, '近半年': 182, '近一年': 365}.get(text)
        if days is None:
            return None
        return now - days * 86400

    def _start_file_search(self):
        root = getattr(self, 'current_folder', None)
        if not (root and os.path.isdir(root)):
            return
        self._search_token += 1
        token = self._search_token
        old = getattr(self, '_search_thread', None)
        if old is not None:
            # 先断开信号，避免 wait 超时后旧线程的信号发往已 deleteLater 的对象
            try:
                old.search_ready.disconnect(self._on_search_ready)
            except (TypeError, RuntimeError):
                pass
            if old.isRunning():
                old.requestInterruption()
                old.quit()
                old.wait(2000)
            old.deleteLater()
            self._search_thread = None
        name = self.search_name_edit.text()
        exts = self._search_exts_for_combo()
        mtime_after = self._search_mtime_after()
        # 切到结果面板
        self.file_tree.hide()
        self.search_results_panel.show()
        self.search_results_list.clear()
        self.search_status_label.setText('搜索中…')
        self._search_thread = FileSearchThread(root, token, name, exts, mtime_after)
        self._search_thread.search_ready.connect(self._on_search_ready)
        self._search_thread.start()

    def _on_search_ready(self, token, results, truncated):
        if token != self._search_token:
            return  # 过期，丢弃
        self.search_results_list.clear()
        # 按所在目录分组、排序（目录升序 -> 文件名升序），避免杂乱
        root = getattr(self, 'current_folder', '') or ''
        def dir_key(item):
            d = os.path.dirname(item[1])
            return (d.lower(), d)
        results = sorted(results, key=lambda it: (dir_key(it)[0], it[0].lower()))
        cur_group = None
        for name, rel, full, size, mtime in results:
            dpart = os.path.dirname(rel)
            if dpart != cur_group:
                cur_group = dpart
                title = dpart if dpart else os.path.basename(root) or '（项目根）'
                grp = QListWidgetItem(f'📁  {title}')
                grp.setFlags(Qt.NoItemFlags)  # 组标题不可选/不可双击定位
                f = grp.font(); f.setBold(True)
                grp.setFont(f)
                grp.setBackground(QColor('#eef3f8'))
                self.search_results_list.addItem(grp)
            it = QListWidgetItem(f'    {name}    ({self.format_file_size(size)})')
            it.setData(Qt.UserRole, full)
            it.setToolTip(full)
            self.search_results_list.addItem(it)
        n = len(results)
        if truncated:
            self.search_status_label.setText(f'找到 {n}+ 项（结果过多已截断，请细化条件）')
        elif n == 0:
            self.search_status_label.setText('未找到匹配文件')
        else:
            self.search_status_label.setText(f'找到 {n} 项（双击在文件树中定位）')

    def _on_search_result_double_clicked(self, item):
        """双击结果：回显文件树，逐级展开父链并定位到该文件。"""
        full = item.data(Qt.UserRole)
        if not full or not os.path.exists(full):
            return
        # 回到文件树（关闭结果面板会 show file_tree）
        self._close_search_results()
        focus = full if os.path.isdir(full) else os.path.dirname(full)
        # 逐级展开父链，确保深层目录可见（QFileSystemModel 懒加载需边展开边泵事件）
        self._expand_ancestors(focus)
        # 选中文件/目录本身并滚动到位
        target = focus if os.path.isdir(full) else full
        try:
            idx = self.file_model.index(target)
            if idx.isValid():
                self.file_tree.setCurrentIndex(idx)
                # 加载可能异步：泵几轮事件后再滚动
                for _ in range(5):
                    QApplication.processEvents()
                self.file_tree.scrollTo(idx, QAbstractItemView.PositionAtCenter)
        except Exception:
            pass
        self._breadcrumb_path = focus
        self._rebuild_breadcrumb()

    def _expand_ancestors(self, path):
        """逐级展开从 current_folder 到 path 的父链，边展开边泵事件以触发懒加载。"""
        root = os.path.normpath(getattr(self, 'current_folder', '') or '')
        cur = os.path.normpath(path)
        chain = []
        while cur and cur != root and os.path.commonpath([cur, root]) == root and cur != os.path.dirname(cur):
            chain.append(cur)
            cur = os.path.dirname(cur)
        for p in reversed(chain):
            try:
                idx = self.file_model.index(p)
                if idx.isValid():
                    self.file_tree.setExpanded(idx, True)
                    QApplication.processEvents()
            except Exception:
                pass

    def _close_search_results(self):
        """关闭结果面板，回显文件树。断开信号并等待线程结束，避免无效搜索浪费 CPU。"""
        t = getattr(self, '_search_thread', None)
        if t is not None:
            # 断开信号：让在跑线程的结果静默丢弃，无需改 token 影响后续搜索
            try:
                t.search_ready.disconnect(self._on_search_ready)
            except (TypeError, RuntimeError):
                pass
            if t.isRunning():
                t.requestInterruption()
                t.quit()
                t.wait(2000)
            t.deleteLater()
            self._search_thread = None
        self.search_results_list.clear()
        self.search_status_label.setText('')
        self.search_results_panel.hide()
        self.file_tree.show()

    def _set_search_bar_enabled(self, enabled):
        for w in (self.search_name_edit, self.search_type_combo, self.search_date_combo, self.search_btn):
            w.setEnabled(enabled)
        if not enabled:
            self._close_search_results()

    def _select_project_path(self, folder_path):
        """把某个项目路径设为当前：刷新右侧 file_tree、清预览、启用「新建结构」按钮、更新状态栏。
        供点击项目行与启动恢复共用，避免两份同步漂移。路径不存在则静默跳过。"""
        if not (folder_path and os.path.exists(folder_path)):
            return
        self.current_folder = folder_path
        self.file_model.setRootPath(folder_path)
        self.file_tree.setRootIndex(self.file_model.index(folder_path))
        self._reset_preview()
        self.metadata_tab.clear()
        self.new_structure_btn.setEnabled(True)
        self._update_folder_status_bar()
        # 面包屑焦点回到项目根
        self._breadcrumb_path = folder_path
        self._rebuild_breadcrumb()
        # 切了项目：清空旧搜索结果，启用文件搜索条
        self._close_search_results()
        self._set_search_bar_enabled(True)

    def on_folder_cell_clicked(self, row, column):
        table = self.sender()
        folder_path = table.item(row, 0).data(Qt.UserRole)
        if folder_path and os.path.exists(folder_path):
            self._select_project_path(folder_path)
            # 仅点击项目行才更新「上次项目」；快捷访问不走这里
            self.last_project_path = folder_path
    
    def on_folder_cell_double_clicked(self, row, column):
        """双击表格项：根据列执行不同操作"""
        table = self.sender()
        folder_path = table.item(row, 0).data(Qt.UserRole)
        
        if not folder_path or not os.path.exists(folder_path):
            return
        
        # 根据列执行不同操作
        if column == 0:  # 双击编号列：打开文件夹
            self._open_with_shell(folder_path)
        elif column == 1:  # 双击注释列：修改注释
            # 优先编辑当前界面显示的有效注释：JSON 覆盖值优先，否则使用文件夹名后缀
            current_comment = self._get_effective_folder_comment(folder_path)
            stored_comment = self.comments.get(folder_path)

            # 弹出对话框编辑注释
            folder_name = os.path.basename(folder_path)
            dialog = CommentEditDialog(f'编辑项目注释 - {folder_name}', current_comment, self)
            if dialog.exec_():
                new_comment = dialog.get_comment()
                if new_comment != stored_comment:
                    if new_comment:
                        self.comments[folder_path] = new_comment
                    else:
                        # 如果注释为空，从存储中删除，界面会回退显示文件夹名后缀注释
                        self.comments.pop(folder_path, None)
                    # 保存注释
                    self.save_comments()
                    # 刷新文件夹列表
                    self.load_filtered_folders()
                    # 定位回原来的文件夹
                    self.locate_new_folder(folder_path)
    

    
    def on_file_clicked(self, index):
        self._reset_preview()
        file_path = self.file_model.filePath(index)
        file_info = self.file_model.fileInfo(index)
        self.extract_metadata(file_info)
        self._update_breadcrumb_for_item(file_path, file_info)
        if file_info.isFile():
            self._scheduled_preview_path = file_path
            self._preview_timer.start()

    def _cancel_pending_preview(self):
        """取消已安排的待处理预览，避免与双击冲突。"""
        timer = getattr(self, '_preview_timer', None)
        if timer is not None:
            try:
                if timer.isActive():
                    timer.stop()
            except RuntimeError:
                pass
        self._scheduled_preview_path = None

    def _execute_pending_preview(self):
        """执行待处理的预览，如果路径仍然有效且窗口未关闭。"""
        path = getattr(self, '_scheduled_preview_path', None)
        self._scheduled_preview_path = None
        # 窗口已关闭时不访问 GUI 对象
        if self.isVisible() and path and os.path.exists(path):
            self.preview_file(path)


    def _update_breadcrumb_for_item(self, file_path, file_info):
        """面包屑跟随：文件取其所在目录，目录取自身。"""
        focus = os.path.dirname(file_path) if file_info.isFile() else file_path
        self._breadcrumb_path = focus
        self._rebuild_breadcrumb()

    
    def on_file_tree_context_menu(self, position):
        """文件树右键菜单"""
        index = self.file_tree.indexAt(position)
        menu = QMenu(self)

        if index.isValid():
            clicked_path = self.file_model.filePath(index)
            selected_paths = self._get_selected_file_paths()
            if clicked_path not in selected_paths:
                self._select_single_file_index(index)

            selected_paths = self._get_selected_file_paths()
            if not selected_paths:
                return

            multi_selected = len(selected_paths) > 1
            file_path = selected_paths[0]
            file_info = self.file_model.fileInfo(index)
            ext = os.path.splitext(file_path)[1].lower()
            is_archive = ext in ['.zip', '.rar', '.7z']

            copy_action = menu.addAction('复制')
            add_to_zip_action = menu.addAction('添加到zip压缩包')

            paste_copy_action = None
            rename_action = None
            terminal_action = None
            extract_action = None
            recycle_action = None

            if not multi_selected:
                paste_copy_action = menu.addAction('粘贴副本')
                rename_action = menu.addAction('重命名')
                save_version_action = menu.addAction('保存版本')
                if os.path.isdir(file_path):
                    terminal_action = menu.addAction('在终端中打开')
                menu.addSeparator()
                if is_archive:
                    extract_action = menu.addAction('智能解压')
                menu.addSeparator()
                archive_action = menu.addAction('归档到old文件夹')
                recycle_action = menu.addAction('移入回收站')
                paste_copy_action.setEnabled(self._has_pasteable_clipboard())
            else:
                menu.addSeparator()
                archive_action = menu.addAction('归档到old文件夹')
                recycle_action = menu.addAction('移入回收站')

            action = menu.exec_(self.file_tree.viewport().mapToGlobal(position))

            if action == copy_action:
                self._copy_paths_to_clipboard(selected_paths)
            elif action == add_to_zip_action:
                if multi_selected:
                    self.add_paths_to_zip(selected_paths)
                else:
                    self.add_to_zip(file_path)
            elif not multi_selected and action == paste_copy_action:
                self.paste_copy(file_path)
            elif not multi_selected and action == rename_action:
                self.rename_item(file_path)
            elif not multi_selected and action == save_version_action:
                self.save_file_version(file_path)
            elif not multi_selected and terminal_action and action == terminal_action:
                self.open_folder_in_terminal(file_path)
            elif not multi_selected and extract_action and action == extract_action:
                self.smart_extract(file_path)
            elif action == recycle_action:
                if multi_selected:
                    self._move_paths_to_recycle(selected_paths)
                else:
                    self.move_to_recycle(file_path)
            elif action == archive_action:
                self.archive_to_old_folder(selected_paths)
        else:
            # 右键点击了空白区域
            paste_copy_action = menu.addAction('粘贴副本')
            paste_copy_action.setEnabled(self._has_pasteable_clipboard())

            action = menu.exec_(self.file_tree.viewport().mapToGlobal(position))

            if action == paste_copy_action:
                # 粘贴到当前文件夹
                if self.current_folder and os.path.exists(self.current_folder):
                    self.paste_copy(self.current_folder)
                else:
                    QMessageBox.warning(self, "警告", "请先选择一个项目文件夹")
    
    def _has_pasteable_clipboard(self):
        """剪贴板中是否有可粘贴的有效路径"""
        for p in (self.clipboard_paths or []):
            if os.path.exists(p):
                return True
        return bool(self.clipboard_path) and os.path.exists(self.clipboard_path)

    def paste_copy(self, target_path):
        """粘贴副本到目标路径，支持多选"""
        sources = [p for p in (self.clipboard_paths or []) if os.path.exists(p)]
        if not sources and self.clipboard_path and os.path.exists(self.clipboard_path):
            sources = [self.clipboard_path]
        if not sources:
            QMessageBox.warning(self, "警告", "剪贴板中没有有效的文件")
            return

        # 目标目录：若 target_path 是目录则用它，否则用其所在目录
        target_dir = target_path if os.path.isdir(target_path) else os.path.dirname(target_path)

        pasted = 0
        errors = []
        for source in sources:
            try:
                self._paste_single(source, target_dir)
                pasted += 1
            except Exception as e:
                errors.append(f"{os.path.basename(source)}: {str(e)}")

        if errors:
            QMessageBox.warning(self, "错误", "粘贴副本失败:\n" + "\n".join(errors))
        if pasted == 1:
            self.statusBar().showMessage(f"已粘贴副本: {os.path.basename(sources[0])}")
        elif pasted > 1:
            self.statusBar().showMessage(f"已粘贴 {pasted} 个副本")

    def _paste_single(self, source_path, target_dir):
        """复制单个文件/文件夹到目标目录，处理重名"""
        # 防止把文件夹复制进它自身或其子目录，否则 copytree 会无限递归嵌套直到路径超长
        if os.path.isdir(source_path):
            src_norm = os.path.normcase(os.path.normpath(os.path.abspath(source_path)))
            tgt_norm = os.path.normcase(os.path.normpath(os.path.abspath(target_dir)))
            if tgt_norm == src_norm or tgt_norm.startswith(src_norm + os.sep):
                raise Exception(f'不能将文件夹复制到其自身或子目录中：{os.path.basename(source_path)}')

        base_name = os.path.basename(source_path)
        dest = os.path.join(target_dir, base_name)

        # 处理重名：如果目标已存在，则生成副本名称
        if os.path.exists(dest):
            name, ext = os.path.splitext(base_name)
            # 检查原文件名是否已以 "_副本数字" 结尾
            match = re.search(r'_副本(\d+)$', name)
            if match:
                # 原文件已经是副本，从该数字继续递增
                base_name_without_copy = name[:match.start()]
                counter = int(match.group(1)) + 1
            else:
                # 原文件不是副本，从1开始
                base_name_without_copy = name
                counter = 1
            while os.path.exists(dest):
                new_name = f"{base_name_without_copy}_副本{counter}{ext}"
                dest = os.path.join(target_dir, new_name)
                counter += 1
                if counter > 100000:  # 安全上限,避免极端情况无限循环
                    raise Exception(f"无法生成唯一的目标文件名(已尝试 {counter} 次),请清理目标目录中的 '副本' 文件")

        if os.path.isdir(source_path):
            shutil.copytree(source_path, dest)
        else:
            shutil.copy2(source_path, dest)
    
    def save_file_version(self, file_path):
        """保存文件版本：生成 文件名_YYYYMMDD[后缀].ext 的副本。

        命名规则（与手动习惯一致）：
          当天第一个版本 → S1200-10_20260708.dsn
          当天第二个版本 → S1200-10_20260708a.dsn
          当天第三个版本 → S1200-10_20260708b.dsn
        """
        try:
            if not os.path.isfile(file_path):
                return
            dir_name = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            ext = os.path.splitext(file_path)[1]
            # 若文件名末尾已有 _YYYYMMDD 或 _YYYYMMDD[a-z] 后缀，剥离它避免重复
            date_match = re.search(r'_\d{8}[a-z]?$', base_name)
            if date_match:
                base_name = base_name[:date_match.start()]
            today = time.strftime('%Y%m%d')
            # 查找当天已有的版本
            existing_suffixes = []
            for f in os.listdir(dir_name):
                if f.startswith(base_name + '_' + today) and f.endswith(ext):
                    # 提取后缀部分：_YYYYMMDD 之后、.ext 之前
                    # 对于无扩展名文件(ext='')，直接使用日期后内容
                    # 提取后缀：_YYYYMMDD 之后的部分
                    # 注意：当 ext='' 时，f[:-0] 变成 f[:0]='' 导致无法提取后缀
                    # 所以这里显式处理 ext 为空的情况
                    suffix_start = len(base_name + '_' + today)
                    if ext:
                        middle = f[suffix_start:-len(ext)]
                    else:
                        middle = f[suffix_start:]
                    if middle == '' or (len(middle) == 1 and middle.isalpha()):
                        existing_suffixes.append(middle)
            # 确定下一个后缀：'' → a → b → c → ...
            all_suffixes = set(existing_suffixes)
            candidate = ''
            if candidate in all_suffixes:
                candidate = 'a'
                while candidate in all_suffixes:
                    candidate = chr(ord(candidate) + 1)
            next_suffix = candidate
            new_name = f'{base_name}_{today}{next_suffix}{ext}'
            new_path = os.path.join(dir_name, new_name)
            shutil.copy2(file_path, new_path)
            self.statusBar().showMessage(f'已保存版本: {new_name}')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'保存版本失败: {str(e)}')

    def rename_item(self, file_path):
        """重命名文件或文件夹"""
        try:
            old_name = os.path.basename(file_path)
            parent_dir = os.path.dirname(file_path)
            
            dialog = RenameDialog(file_path, self)
            if not dialog.exec_():
                return
            
            new_name = dialog.get_new_name()
            new_path = os.path.join(parent_dir, new_name)
            if os.path.exists(new_path):
                QMessageBox.warning(self, '警告', f'名称 "{new_name}" 已存在')
                return
            
            os.rename(file_path, new_path)
            self.statusBar().showMessage(f"已重命名: {old_name} -> {new_name}")

            settings_changed = False
            if file_path in self.comments:
                self.comments[new_path] = self.comments.pop(file_path)
                self.save_comments()
            if file_path in self.pinned_folders:
                self.pinned_folders = [new_path if path == file_path else path for path in self.pinned_folders]
                settings_changed = True
            if file_path in self.hidden_folders:
                self.hidden_folders = [new_path if path == file_path else path for path in self.hidden_folders]
                settings_changed = True
            if settings_changed:
                self.save_settings_to_file(self.settings, self.include_subfolders)

            # 如果重命名的是当前项目文件夹，更新current_folder
            if file_path == self.current_folder:
                self.current_folder = new_path
                self.file_model.setRootPath(new_path)
                self.file_tree.setRootIndex(self.file_model.index(new_path))
                self._update_folder_status_bar()
                # 面包屑焦点回到新的项目根
                self._breadcrumb_path = new_path
                self._rebuild_breadcrumb()
            # 同步更新「上次项目」记录，避免下次启动恢复指向旧名字
            if getattr(self, 'last_project_path', None) == file_path:
                self.last_project_path = new_path
        except Exception as e:
            QMessageBox.warning(self, "错误", f"重命名失败: {str(e)}")
    
    def add_to_zip(self, source_path):
        """将文件或文件夹添加到zip压缩包"""
        try:
            import zipfile

            parent_dir = os.path.dirname(source_path)
            base_name = os.path.basename(source_path)
            zip_path, zip_name = self._create_unique_zip_path(parent_dir, base_name)

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                self._write_path_to_zip(zf, source_path, parent_dir)

            self.statusBar().showMessage(f"已创建压缩包: {zip_name}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"创建压缩包失败: {str(e)}")

    def add_paths_to_zip(self, source_paths):
        """将多个文件或文件夹添加到同一个zip压缩包"""
        valid_paths = []
        seen = set()
        for source_path in source_paths:
            normalized_path = os.path.normpath(source_path)
            if os.path.exists(source_path) and normalized_path not in seen:
                seen.add(normalized_path)
                valid_paths.append(source_path)

        if not valid_paths:
            QMessageBox.warning(self, "警告", "没有可压缩的文件或文件夹")
            return

        try:
            import zipfile

            parent_dirs = {os.path.dirname(path) for path in valid_paths}
            if len(parent_dirs) == 1:
                target_dir = parent_dirs.pop()
                base_dir = target_dir
            else:
                common_path = os.path.commonpath(valid_paths)
                base_dir = common_path if os.path.isdir(common_path) else os.path.dirname(common_path)
                target_dir = base_dir

            zip_path, zip_name = self._create_unique_zip_path(target_dir, '选中文件')

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for source_path in valid_paths:
                    self._write_path_to_zip(zf, source_path, base_dir)

            self.statusBar().showMessage(f"已创建压缩包: {zip_name}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"创建压缩包失败: {str(e)}")

    @staticmethod
    def _resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

    def _find_7z_tool(self):
        """查找7-Zip路径"""
        import os
        import sys
        
        # 优先使用用户指定的路径
        configured_path = getattr(self, 'archive_tool_path', '')
        if (
            isinstance(configured_path, str)
            and os.path.basename(configured_path).casefold() == '7z.exe'
            and os.path.isfile(configured_path)
        ):
            return configured_path
        
        # 7z.exe 可能的位置
        sevenzip_paths = []
        
        # 程序所在目录（支持打包后的exe）
        if hasattr(sys, '_MEIPASS'):
            app_dir = os.path.dirname(os.path.abspath(sys.executable))
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))
        sevenzip_paths.append(os.path.join(app_dir, '7z.exe'))
        
        # 7-Zip安装目录
        sevenzip_paths.extend([
            r"C:\Program Files\7-Zip\7z.exe",
            r"C:\Program Files (x86)\7-Zip\7z.exe",
        ])
        
        for path in sevenzip_paths:
            if os.path.isfile(path):
                return path
        
        return None
    
    def _extract_with_7z(self, archive_path, extract_dir):
        """兼容接口：仅允许解压到空目录，且不覆盖任何已存在条目。"""
        sevenzip = self._find_7z_tool()
        if not sevenzip:
            raise ArchiveSafetyError('未找到7-Zip，请在设置中指定7z.exe路径或安装7-Zip')
        os.makedirs(extract_dir, exist_ok=True)
        if os.listdir(extract_dir):
            raise ArchiveSafetyError('安全解压要求目标临时目录为空')
        _extract_7z_to_stage(sevenzip, archive_path, extract_dir)
        return True
    
    def _list_archive_with_7z(self, archive_path):
        """使用 UTF-8 7-Zip 输出列出压缩包内容，供只读预览使用。"""
        if not getattr(self, 'enable_7zip', False):
            raise ArchiveSafetyError('7-Zip 压缩包预览尚未启用')
        sevenzip = self._find_7z_tool()
        if not sevenzip:
            raise ArchiveSafetyError('未找到7-Zip')
        entries = _inspect_7z_archive(sevenzip, archive_path)
        return [(item['name'], int(item.get('size') or 0), bool(item.get('is_dir'))) for item in entries]
    
    def smart_extract(self, archive_path):
        """后台执行事务式智能解压；既有目标自动改名，绝不覆盖。"""
        archive_path = os.path.abspath(archive_path)
        ext = os.path.splitext(archive_path)[1].lower()
        if ext not in ARCHIVE_EXTS or not os.path.isfile(archive_path):
            QMessageBox.warning(self, '错误', '请选择有效的 zip、rar 或 7z 压缩包')
            return
        sevenzip = None
        if ext in ('.rar', '.7z'):
            sevenzip = self._find_7z_tool()
            if not sevenzip:
                QMessageBox.warning(
                    self,
                    '错误',
                    '未找到7-Zip，请在设置中指定7z.exe路径或安装7-Zip',
                )
                return

        progress = QProgressDialog('正在准备安全解压...', '取消', 0, 0, self)
        progress.setWindowTitle('安全解压')
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)

        thread = ArchiveExtractThread(archive_path, sevenzip, self)
        result = {'state': None, 'value': None}
        self._extract_jobs[thread] = progress

        def mark_completed(destination):
            result['state'] = 'completed'
            result['value'] = destination

        def mark_failed(message):
            result['state'] = 'failed'
            result['value'] = message

        def mark_canceled():
            result['state'] = 'canceled'

        def cancel_extract():
            thread.cancel()
            self.statusBar().showMessage('正在取消解压并清理临时文件...')

        def finish_job():
            job_progress = self._extract_jobs.pop(thread, None)
            if job_progress is not None:
                job_progress.close()
                job_progress.deleteLater()
            state = result['state']
            if state == 'completed':
                destination = result['value']
                self.statusBar().showMessage(f'已安全解压到: {destination}')
            elif state == 'failed':
                QMessageBox.warning(self, '解压失败', str(result['value']))
            elif state == 'canceled':
                self.statusBar().showMessage('已取消解压，临时文件已清理')
            else:
                QMessageBox.warning(self, '解压失败', '解压线程异常结束')
            thread.deleteLater()
            if self._close_after_extract_cancel and not self._extract_jobs:
                self._close_after_extract_cancel = False
                QTimer.singleShot(0, self.close)

        thread.status_changed.connect(progress.setLabelText)
        thread.extract_completed.connect(mark_completed)
        thread.extract_failed.connect(mark_failed)
        thread.extract_canceled.connect(mark_canceled)
        thread.finished.connect(finish_job)
        progress.canceled.connect(cancel_extract)
        thread.start()
        progress.show()
    
    def _update_folder_status_bar(self):
        """更新状态栏显示当前项目文件夹信息"""
        if self.current_folder:
            folder_name = os.path.basename(self.current_folder)
            mb_match = self.folder_regex_mb.match(folder_name)
            db_match = self.folder_regex_db.match(folder_name)
            match = mb_match or db_match
            if match:
                number = match.group(1)
                self.statusBar().showMessage(f"当前文件夹：{number}")
        # 同步刷新文件数/大小统计
        self._refresh_folder_stats()

    def _refresh_folder_stats(self):
        """启动后台线程统计 current_folder 的文件数与总大小。token 机制丢弃过期结果。"""
        self._stats_token += 1
        token = self._stats_token
        # 中断并等待上一个线程结束，再启动新的，避免僵尸线程堆积
        old = getattr(self, '_stats_thread', None)
        if old is not None:
            try:
                old.stats_ready.disconnect(self._on_stats_ready)
            except (TypeError, RuntimeError):
                pass
            if old.isRunning():
                old.requestInterruption()
                old.quit()
                old.wait(2000)
            old.deleteLater()
            self._stats_thread = None
        root = getattr(self, 'current_folder', None)
        if not root or not os.path.isdir(root):
            self.folder_stats_label.setText('')
            return
        self.folder_stats_label.setText('统计中…')
        self._stats_thread = FolderStatsThread(root, token)
        self._stats_thread.stats_ready.connect(self._on_stats_ready)
        self._stats_thread.start()

    def _on_stats_ready(self, token, count, size, truncated):
        """统计线程回调：仅接受当前 token 的结果。"""
        if token != self._stats_token:
            return  # 过期，丢弃
        if truncated:
            self.folder_stats_label.setText(f'{count}+ 个文件 · ≥{self.format_file_size(size)}')
        else:
            self.folder_stats_label.setText(f'{count} 个文件 · {self.format_file_size(size)}')

    def archive_to_old_folder(self, file_paths):
        """将选中文件/文件夹移入各自目录下的 old 文件夹。
        支持跨目录多选，每个文件各自移入自己目录下的 old/。"""
        try:
            if not file_paths:
                return
            moved = 0
            for file_path in file_paths:
                if not os.path.exists(file_path):
                    continue
                parent_dir = os.path.dirname(file_path)
                target_dir = os.path.join(parent_dir, 'old')
                os.makedirs(target_dir, exist_ok=True)
                dest = os.path.join(target_dir, os.path.basename(file_path))
                # 重名时追加数字
                if os.path.exists(dest):
                    base, ext = os.path.splitext(os.path.basename(file_path))
                    i = 1
                    while os.path.exists(os.path.join(target_dir, f'{base}_{i}{ext}')):
                        i += 1
                    dest = os.path.join(target_dir, f'{base}_{i}{ext}')
                shutil.move(file_path, dest)
                moved += 1
            self.statusBar().showMessage(f'已归档 {moved} 个项目到 old/')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'归档失败: {str(e)}')

    def move_to_recycle(self, file_path):
        """移入回收站"""
        self._move_paths_to_recycle([file_path])
    
    def _open_with_shell(self, path):
        r"""用系统默认程序打开文件/文件夹,失败弹错。
        对 UNC 路径(//server/share 或 \\server\share),os.startfile 会报 WinError 2,
        改用 explorer.exe 直接打开。"""
        try:
            if path.startswith('\\\\') or path.startswith('//'):
                # UNC 路径:用 explorer 打开
                subprocess.Popen(['explorer', path])
            else:
                # 使用 ShellExecuteW 更可靠，能正确处理无关联程序的情况
                result = ctypes.windll.shell32.ShellExecuteW(None, 'open', path, None, None, 1)
                if result <= 32:
                    # ShellExecuteW 失败，回退到 os.startfile
                    os.startfile(path)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开：{os.path.basename(path)}\n{str(e)}")

    def _open_url(self, url):
        """用默认浏览器打开网址。"""
        try:
            os.startfile(url)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开网页：\n{url}\n\n{str(e)}")

    def _terminal_launch_candidates(self, folder_path):
        """生成安全的终端候选；路径只作为 cwd 或独立 argv，不拼入命令文本。"""
        system_root = os.environ.get('SystemRoot', r'C:\Windows')
        local_app_data = os.environ.get('LOCALAPPDATA', '')
        is_unc = _is_unc_path(folder_path)
        candidates = []

        if local_app_data:
            wt_path = os.path.join(local_app_data, 'Microsoft', 'WindowsApps', 'wt.exe')
            if os.path.exists(wt_path):
                # wt 的 -d 接受独立 argv；同时传 cwd，确保本地路径和支持 UNC 的版本都正确落点。
                candidates.append(('Windows 终端', wt_path, ['-d', folder_path], folder_path))

        powershell_path = os.path.join(
            system_root,
            'System32',
            'WindowsPowerShell',
            'v1.0',
            'powershell.exe',
        )
        if os.path.exists(powershell_path):
            if is_unc:
                # Win32 进程 cwd 不能可靠表示 UNC；用 EncodedCommand 安全设置 PowerShell provider 路径。
                ps_args = _powershell_unc_location_args(folder_path)
                ps_working_dir = os.path.expanduser('~')
            else:
                ps_args = ['-NoLogo', '-NoProfile', '-NoExit']
                ps_working_dir = folder_path
            candidates.append(('PowerShell', powershell_path, ps_args, ps_working_dir))

        # cmd.exe 原生不支持 UNC 当前目录；UNC 时宁可明确失败，也不打开到错误位置。
        if not is_unc:
            cmd_path = os.path.join(system_root, 'System32', 'cmd.exe')
            if os.path.exists(cmd_path):
                candidates.append(('命令提示符', cmd_path, ['/D'], folder_path))
        return candidates

    def _shell_execute_terminal(self, executable, args, working_directory):
        """通过 ShellExecuteW 启动终端，working_directory 是所选路径。"""
        from ctypes import wintypes

        params = subprocess.list2cmdline(args) if args else None
        shell_execute = ctypes.windll.shell32.ShellExecuteW
        shell_execute.argtypes = [
            wintypes.HWND,
            wintypes.LPCWSTR,
            wintypes.LPCWSTR,
            wintypes.LPCWSTR,
            wintypes.LPCWSTR,
            ctypes.c_int,
        ]
        shell_execute.restype = ctypes.c_ssize_t
        result = shell_execute(
            int(self.winId()),
            'open',
            executable,
            params,
            working_directory or None,
            1,
        )
        if result <= 32:
            raise OSError(f'ShellExecuteW 返回错误码 {result}')

    def open_folder_in_terminal(self, folder_path):
        """以普通权限在指定文件夹打开终端，并确保启动位置与所选路径一致。"""
        folder_path = os.path.abspath(os.path.normpath(folder_path))
        if not os.path.isdir(folder_path):
            QMessageBox.warning(self, '警告', f'文件夹不存在：{folder_path}')
            return False

        last_error = None
        candidates = self._terminal_launch_candidates(folder_path)
        for terminal_name, executable, args, working_directory in candidates:
            try:
                self._shell_execute_terminal(executable, args, working_directory)
                self.statusBar().showMessage(
                    f'已在{terminal_name}中打开所选路径: {folder_path}'
                )
                return True
            except Exception as e:
                last_error = e
        detail = str(last_error) if last_error else '未找到可用的终端程序'
        QMessageBox.warning(
            self,
            '错误',
            f'无法在终端中打开所选路径：{folder_path}\n{detail}',
        )
        return False

    def on_file_double_clicked(self, index):
        """双击文件树项：直接打开文件或展开目录；双击空白区域打开当前文件夹"""
        self._cancel_pending_preview()
        if not index.isValid():
            # 双击空白区域：打开当前文件夹
            if self.current_folder:
                self._open_with_shell(self.current_folder)
        else:
            # 双击有效项：直接打开文件或文件夹
            file_path = self.file_model.filePath(index)
            self._open_with_shell(file_path)
    def new_project(self):
        # 获取默认新建项目文件夹：优先使用用户上次选择的目录，再回退到第一条项目根目录
        if hasattr(self, 'default_new_project_folder') and self.default_new_project_folder and os.path.isdir(self.default_new_project_folder):
            default_folder = self.default_new_project_folder
        elif hasattr(self, 'settings') and self.settings is not None and self.settings:
            default_folder = self.settings[0][1]
        else:
            default_folder = os.path.expanduser("~")

        dialog = NewProjectDialog(self, default_folder=default_folder)
        if dialog.exec_():
            project_info = dialog.get_project_info()
            self.load_filtered_folders()
            self.locate_new_folder(project_info['full_path'])
    
    def locate_new_folder(self, folder_path):
        folder_name = os.path.basename(folder_path)
        mb_match = self.folder_regex_mb.match(folder_name)
        db_match = self.folder_regex_db.match(folder_name)
        if mb_match:
            table = self.motherboard_table
        elif db_match:
            table = self.daughterboard_table
        else:
            return
            for row in range(table.rowCount()):
                row_path = table.item(row, 0).data(Qt.UserRole)
                if row_path == folder_path:
                    table.selectRow(row)
                    table.scrollToItem(table.item(row, 0))
                    self.current_folder = folder_path
                    self.file_tree.setRootIndex(self.file_model.index(folder_path))
                    self.new_structure_btn.setEnabled(True)
                    break
    
    def new_folder_structure(self):
        if not self.current_folder:
            QMessageBox.warning(self, '警告', '请先选择一个项目文件夹')
            return
        folder_name = os.path.basename(self.current_folder)
        if not (self.folder_regex_mb.match(folder_name) or self.folder_regex_db.match(folder_name)):
            QMessageBox.warning(self, '警告', '当前选择的不是有效的项目文件夹')
            return
        try:
            dialog = NewStructureDialog(self.current_folder, self)
            if dialog.exec_():
                structure_info = dialog.get_structure_info()
                version = structure_info['version']
                all_folders = structure_info['selected_folders'] + structure_info['custom_folders']
                if not all_folders:
                    QMessageBox.warning(self, '警告', '请至少选择一个文件夹')
                    return
                version_folder = os.path.join(self.current_folder, f'V{version}')
                if not os.path.exists(version_folder):
                    os.makedirs(version_folder)
                created = []
                failed = []
                for folder in all_folders:
                    folder_path = os.path.join(version_folder, folder)
                    if os.path.exists(folder_path):
                        continue
                    try:
                        os.makedirs(folder_path)
                        created.append(folder)
                    except Exception as fe:
                        failed.append(f'{folder}（{fe}）')
                if failed:
                    msg = f'已创建 {len(created)} 个子文件夹，以下创建失败：\n' + '\n'.join('· ' + f for f in failed)
                    msg += '\n\n请检查目标是否为只读、网络盘是否断开或权限不足。'
                    QMessageBox.warning(self, '部分完成', msg)
                elif created:
                    QMessageBox.information(self, '成功', f'已创建版本文件夹 V{version} 和 {len(created)} 个子文件夹')
                else:
                    QMessageBox.information(self, '提示', f'版本文件夹 V{version} 已存在，所有子文件夹也已存在')
                self.file_tree.setRootIndex(self.file_model.index(self.current_folder))
        except Exception as e:
            QMessageBox.critical(self, '错误', f'创建子文件夹失败: {str(e)}')
    
    def _preview_text(self, file_path):
        # 大文件只读前 PREVIEW_TEXT_LIMIT 字节,避免一次性读全文导致界面卡顿
        PREVIEW_TEXT_LIMIT = 1 * 1024 * 1024  # 1 MB
        file_size = os.path.getsize(file_path)
        truncated = file_size > PREVIEW_TEXT_LIMIT

        # 尝试顺序：GBK(中文 Windows 最常见) → UTF-8 → 兜底 UTF-8 + replace
        content = None
        for encoding in ('gbk', 'utf-8'):
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read(PREVIEW_TEXT_LIMIT) if truncated else f.read()
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read(PREVIEW_TEXT_LIMIT) if truncated else f.read()

        if truncated:
            content += f'\n\n--- 文件过大,仅显示前 {self.format_file_size(PREVIEW_TEXT_LIMIT)} (共 {self.format_file_size(file_size)}) ---\n'
        self.preview_tab.setPlainText(content)

    def _preview_pdf(self, file_path):
        if not PdfReader:
            self.preview_tab.setPlainText('PDF文件预览功能需要安装PyPDF2库')
            return
        content = f'PDF文件: {os.path.basename(file_path)}\n\n'
        try:
            reader = PdfReader(file_path)
            content += f'页数: {len(reader.pages)}\n\n'
            for i, page in enumerate(reader.pages[:PREVIEW_PDF_PAGES]):
                content += f'第 {i+1} 页:\n{page.extract_text()}\n\n'
        except Exception as e:
            content += f'PDF读取错误: {str(e)}'
        self.preview_tab.setPlainText(content)

    def _get_preview_size(self):
        """获取预览区域可用尺寸"""
        container_size = self.preview_container.size()
        return container_size.width() - 20, container_size.height() - 20

    def _scale_image(self, image):
        """按预览区域缩放图片"""
        available_width, available_height = self._get_preview_size()
        return image.scaled(available_width, available_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)

    # 大图安全阈值：超过此尺寸先缩到该尺寸再加载，避免 OOM
    IMAGE_SAFE_MAX_DIM = 4096

    def _load_image_safe(self, file_path):
        """安全加载图片：超大图先缩到 IMAGE_SAFE_MAX_DIM 以内，避免 OOM。"""
        from PyQt5.QtGui import QImageReader
        reader = QImageReader(file_path)
        if not reader.canRead():
            return QImage(file_path)  # 回退到原逻辑
        size = reader.size()
        if size.width() > self.IMAGE_SAFE_MAX_DIM or size.height() > self.IMAGE_SAFE_MAX_DIM:
            # 等比缩到安全尺寸
            scaled = size.scaled(self.IMAGE_SAFE_MAX_DIM, self.IMAGE_SAFE_MAX_DIM, Qt.KeepAspectRatio)
            reader.setScaledSize(scaled)
            image = QImage()
            reader.read(image)
            return image
        return QImage(file_path)

    def _preview_image(self, file_path):
        try:
            self.preview_tab.hide()
            self.image_scroll_area.show()
            image = self._load_image_safe(file_path)
            if image.isNull():
                self.preview_tab.show()
                self.image_scroll_area.hide()
                self.preview_tab.setPlainText(f'无法加载图片: {os.path.basename(file_path)}')
                return
            scaled_image = self._scale_image(image)
            pixmap = QPixmap.fromImage(scaled_image)
            self.image_label.setPixmap(pixmap)
            self.image_label.setToolTip(f'点击查看大图\n图片: {os.path.basename(file_path)}\n原始尺寸: {image.width()}x{image.height()}')
            self.current_image_path = file_path
        except Exception as e:
            self.preview_tab.show()
            self.image_scroll_area.hide()
            self.preview_tab.setPlainText(f'图片预览错误: {str(e)}')


    def _preview_video(self, file_path):
        try:
            self.current_video_path = file_path
            frames = self.generate_video_thumbnails()
            if not frames:
                # fall back 到单帧
                single = self.generate_video_thumbnail(file_path)
                if single is not None:
                    frames = [QPixmap.fromImage(single)]
            if frames:
                self.preview_tab.hide()
                self.image_scroll_area.show()
                # 横向拼接所有帧为一张宽图
                total_w = sum(pm.width() for pm in frames) + 4 * (len(frames) - 1)
                max_h = max(pm.height() for pm in frames)
                composite = QPixmap(total_w, max_h)
                composite.fill(Qt.transparent)
                painter = QPainter(composite)
                x = 0
                for pm in frames:
                    y = (max_h - pm.height()) // 2
                    painter.drawPixmap(x, y, pm)
                    x += pm.width() + 4  # 4px 间距
                painter.end()
                self.image_label.setPixmap(composite)
                self.image_label.setToolTip(
                    f'视频: {os.path.basename(file_path)}\n左→右分别对应 {",".join(f"{int(p*100)}%" for p in VIDEO_PREVIEW_POSITIONS)} 位置\n点击查看大图')
                self.current_image_path = file_path
            else:
                self.preview_tab.show()
                self.image_scroll_area.hide()
                if HAS_OPENCV:
                    self.preview_tab.setPlainText(f'视频: {os.path.basename(file_path)}\n无法生成视频缩略图')
                else:
                    self.preview_tab.setPlainText(f'视频: {os.path.basename(file_path)}\n视频缩略图功能需要安装OpenCV (pip install opencv-python numpy pillow)')
        except Exception as e:
            self.preview_tab.show()
            self.image_scroll_area.hide()
            self.preview_tab.setPlainText(f'视频预览错误: {str(e)}')

    def _build_file_tree(self, files):
        """构建文件树（使用元组列表格式）"""
        tree = {}
        for filename, size, is_dir in files:
            path_parts = filename.replace('\\', '/').rstrip('/').split('/')
            current = tree
            for i, part in enumerate(path_parts):
                if not part:
                    continue
                if part not in current:
                    current[part] = {'_type': 'dir', '_children': {}, '_size': 0}
                if i == len(path_parts) - 1:
                    if is_dir:
                        current[part]['_type'] = 'dir'
                    else:
                        current[part] = {'_type': 'file', '_size': size, '_children': {}}
                else:
                    current = current[part]['_children']
        return tree

    def _print_tree(self, tree, prefix='', is_last=True):
        result = ''
        items = list(tree.items())
        for i, (name, node) in enumerate(items):
            is_last_item = i == len(items) - 1
            if node['_type'] == 'dir':
                result += f'{prefix}{"└── " if is_last_item else "├── "}{name}/\n'
                new_prefix = prefix + ("    " if is_last_item else "│   ")
                result += self._print_tree(node['_children'], new_prefix, is_last_item)
            else:
                result += f'{prefix}{"└── " if is_last_item else "├── "}{name} ({node["_size"] / 1024:.2f} KB)\n'
        return result

    def _preview_archive(self, file_path, ext):
        if ext in ('.rar', '.7z') and not getattr(self, 'enable_7zip', False):
            self._preview_7z_disabled_hint(file_path)
            return
        try:
            archive_info = f'压缩包: {os.path.basename(file_path)}\n\n'
            files_list = []
            total_size = 0
            if ext == '.zip':
                import zipfile
                with zipfile.ZipFile(file_path, 'r') as zf:
                    for item in zf.infolist():
                        filename = _decode_zip_name(item.filename)
                        files_list.append((filename, item.file_size, item.is_dir()))
                        total_size += item.file_size
            elif ext in ['.rar', '.7z']:
                items = self._list_archive_with_7z(file_path)
                for filename, size, is_dir in items:
                    files_list.append((filename, size, is_dir))
                    total_size += size
            
            file_tree = self._build_file_tree(files_list)
            archive_info += f'文件总数: {len(files_list)}\n'
            archive_info += f'总大小: {total_size / 1024:.2f} KB\n\n'
            archive_info += '文件树结构:\n'
            archive_info += '=' * 60 + '\n'
            archive_info += self._print_tree(file_tree)
            self.preview_tab.setPlainText(archive_info)
        except Exception as e:
            self.preview_tab.setPlainText(f'压缩包预览错误: {str(e)}')

    def _preview_excel(self, file_path, ext):
        if ext in ('.xlsx', '.xlsm'):
            is_macro = (ext == '.xlsm')
            if not load_workbook:
                self.preview_tab.setPlainText(
                    ('Excel宏文件预览功能需要安装openpyxl库' if is_macro else 'Excel文件预览功能需要安装openpyxl库'))
                return
            content = (f'Excel宏文件: {os.path.basename(file_path)}\n\n' if is_macro
                       else f'Excel文件: {os.path.basename(file_path)}\n\n')
            try:
                workbook = load_workbook(file_path, read_only=True, keep_vba=is_macro)
                for sheet_name in workbook.sheetnames:
                    content += f'工作表: {sheet_name}\n'
                    worksheet = workbook[sheet_name]
                    for row in worksheet.iter_rows(min_row=1, max_row=PREVIEW_EXCEL_MAX_ROWS, values_only=True):
                        content += '\t'.join(str(cell) if cell is not None else '' for cell in row) + '\n'
                    content += '\n'
            except Exception as e:
                content += (f'Excel宏文件读取错误: {str(e)}' if is_macro else f'Excel读取错误: {str(e)}')
            self.preview_tab.setPlainText(content)
        elif ext == '.xls':
            if not xlrd:
                self.preview_tab.setPlainText('Excel 97-2003文件预览功能需要安装xlrd库')
                return
            content = f'Excel文件(97-2003): {os.path.basename(file_path)}\n\n'
            try:
                workbook = xlrd.open_workbook(file_path)
                for sheet_idx in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_idx)
                    content += f'工作表: {sheet.name}\n'
                    max_rows = min(sheet.nrows, PREVIEW_EXCEL_MAX_ROWS)
                    for row_idx in range(max_rows):
                        row_data = [str(sheet.cell(row_idx, col_idx).value) for col_idx in range(sheet.ncols)]
                        content += '\t'.join(row_data) + '\n'
                    content += '\n'
            except Exception as e:
                content += f'Excel读取错误: {str(e)}'
            self.preview_tab.setPlainText(content)

    def _preview_word(self, file_path, ext):
        if ext == '.docx':
            if not Document:
                self.preview_tab.setPlainText('Word文件预览功能需要安装python-docx库')
                return
            content = f'Word文件: {os.path.basename(file_path)}\n\n'
            try:
                doc = Document(file_path)
                for para in doc.paragraphs[:PREVIEW_DOCX_PARAGRAPHS]:
                    content += para.text + '\n'
            except Exception as e:
                content += f'Word读取错误: {str(e)}'
            self.preview_tab.setPlainText(content)
        elif ext == '.doc':
            if not olefile:
                self.preview_tab.setPlainText('Word 97-2003文件预览功能需要安装olefile库')
                return
            content = f'Word文件(97-2003): {os.path.basename(file_path)}\n\n'
            ole = None
            try:
                ole = olefile.OleFileIO(file_path)
                if ole.exists('WordDocument'):
                    stream = ole.openstream('WordDocument')
                    data = stream.read()
                    text_parts = [chr(b) for b in data if 32 <= b < 127 or b in (10, 13, 9)]
                    text = ''.join(text_parts)
                    lines = [l.strip() for l in text.split('\n') if l.strip()]
                    content += '\n'.join(lines[:PREVIEW_DOC_LINES])
            except Exception as e:
                content += f'Word 97-2003读取错误: {str(e)}'
            finally:
                # 即使读取异常也要关闭句柄，否则文件被占用无法删除/重命名
                if ole is not None:
                    try:
                        ole.close()
                    except Exception:
                        pass
            self.preview_tab.setPlainText(content)

    def _preview_7z_disabled_hint(self, file_path):
        """在未授权 7-Zip 预览时显示说明，不探测工具也不读取压缩包。"""
        self.preview_tab.setPlainText(
            '7-Zip 压缩包预览未启用\n\n'
            f'文件: {os.path.basename(file_path)}\n\n'
            '程序尚未读取此 .rar / .7z 压缩包。\n'
            '如需预览，请打开“设置 → 7-Zip路径设置”，确认 7z.exe 来源可信后启用。\n'
            '建议使用最新版本的官方 7-Zip，并仅处理来源可信的压缩包。\n\n'
            '.zip 文件使用程序内置能力预览，不受此开关影响。'
        )

    def _preview_binary(self, file_path, ext):
        content = f'文件: {os.path.basename(file_path)}\n\n不支持的文件格式: {ext}\n\n文件前1000字节:\n\n'
        with open(file_path, 'rb') as f:
            content += f.read(1000).hex(' ')
        self.preview_tab.setPlainText(content)

    def _preview_category(self, ext):
        """把扩展名归类到一个预览类别 key，及其中文名。返回 (key, 中文名) 或 (None, None) 表示不受开关控制。"""
        if ext in TEXT_EXTS or ext in ['.bom', '.drc', '.rep', '.rpt']:
            return 'text', '文本'
        if ext == '.pdf':
            return 'pdf', 'PDF'
        if ext in IMAGE_EXTS:
            return 'image', '图片'
        if ext in VIDEO_EXTS:
            return 'video', '视频'
        if ext in ARCHIVE_EXTS:
            return 'archive', '压缩包'
        if ext in ['.xlsx', '.xlsm', '.xls']:
            return 'excel', '表格'
        if ext in ['.docx', '.doc']:
            return 'word', '文档'
        # 二进制兜底（仅读前1000字节，开销小）与加密类型（只显示提示）不参与开关
        return None, None

    def preview_file(self, file_path):
        """预览入口：若对应文件类型的预览被关闭，则只显示「显示预览」按钮，不立即读取文件。"""
        ext = os.path.splitext(file_path)[1].lower()
        key, type_name = self._preview_category(ext)
        if key and not getattr(self, f'preview_{key}_enabled', True):
            self._show_preview_button(file_path, type_name)
            return
        self._do_preview(file_path)

    def _show_preview_button(self, file_path, type_name):
        """隐藏预览区域，改为显示一个按钮，点击后才加载预览。"""
        self.preview_tab.hide()
        self.image_scroll_area.hide()
        self._manual_preview_path = file_path
        self.preview_button.setText(f'显示预览（{type_name}）：{os.path.basename(file_path)}')
        self.preview_button.show()

    def _on_preview_button_clicked(self):
        path = getattr(self, '_manual_preview_path', None)
        if not path or not os.path.exists(path):
            self._reset_preview()
            if path:
                self.preview_tab.setPlainText('文件已不存在，无法显示预览。')
            return
        self._manual_preview_path = None
        self.preview_button.hide()
        self._do_preview(path)

    def _reset_preview(self):
        """恢复预览区域到文本视图并清空，隐藏按钮与图片区域。"""
        self._cancel_pending_preview()
        self._manual_preview_path = None
        self.preview_button.hide()
        self.image_scroll_area.hide()
        self.preview_tab.show()
        self.preview_tab.clear()

    def _do_preview(self, file_path):
        try:
            self._manual_preview_path = None
            ext = os.path.splitext(file_path)[1].lower()
            text_exts = TEXT_EXTS
            image_exts = IMAGE_EXTS
            video_exts = VIDEO_EXTS
            archive_exts = ARCHIVE_EXTS

            self.preview_button.hide()
            self.preview_tab.show()
            self.image_scroll_area.hide()

            # 无扩展名的 dotfile 按文件名匹配
            base_name = os.path.basename(file_path)
            if ext == '' and base_name in TEXT_DOTFILES:
                self._preview_text(file_path)
            elif ext in text_exts:
                self._preview_text(file_path)
            elif ext == '.pdf':
                self._preview_pdf(file_path)
            elif ext in image_exts:
                self._preview_image(file_path)
            elif ext in video_exts:
                self._preview_video(file_path)
            elif ext in archive_exts:
                if ext in ('.rar', '.7z') and not getattr(self, 'enable_7zip', False):
                    self._preview_7z_disabled_hint(file_path)
                else:
                    self._preview_archive(file_path, ext)
            elif ext in ['.xlsx', '.xlsm', '.xls']:
                self._preview_excel(file_path, ext)
            elif ext in ['.docx', '.doc']:
                self._preview_word(file_path, ext)
            elif ext not in ['.opj', '.dsn', '.sch', '.brd', '.dbk', '.dsnlck']:
                self._preview_binary(file_path, ext)
            else:
                self.preview_tab.setPlainText(f'该文件类型（{ext}）涉及加密，无法预览')
        except Exception as e:
            self.preview_tab.setPlainText(f'预览文件时出错: {str(e)}')
    
    def extract_metadata(self, file_info):
        metadata = [
            f'文件名: {file_info.fileName()}',
            f'路径: {file_info.absoluteFilePath()}',
            f'大小: {self.format_file_size(file_info.size())}',
            f'创建时间: {file_info.created().toString()}',
            f'修改时间: {file_info.lastModified().toString()}',
            f'访问时间: {file_info.lastRead().toString()}',
            f'是否为文件: {file_info.isFile()}',
            f'是否为目录: {file_info.isDir()}',
            f'是否可读: {file_info.isReadable()}',
            f'是否可写: {file_info.isWritable()}',
        ]
        self.metadata_tab.setPlainText('\n'.join(metadata))
    
    def format_file_size(self, size):
        if size < 1024:
            return f'{size} B'
        elif size < 1024 * 1024:
            return f'{size / 1024:.2f} KB'
        elif size < 1024 * 1024 * 1024:
            return f'{size / (1024 * 1024):.2f} MB'
        else:
            return f'{size / (1024 * 1024 * 1024):.2f} GB'

    def _format_duration(self, seconds):
        if seconds is None or seconds < 0:
            return '--:--'
        seconds = int(seconds)
        minutes, sec = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        if hours:
            return f'{hours:02d}:{minutes:02d}:{sec:02d}'
        return f'{minutes:02d}:{sec:02d}'

    def _parse_version_tuple(self, version_text):
        """把 v0.2.3 / 0.2.3 解析为可比较的数字元组。"""
        cleaned = str(version_text or '').strip().lstrip('vV')
        parts = []
        for part in cleaned.split('.'):
            match = re.match(r'^(\d+)', part)
            parts.append(int(match.group(1)) if match else 0)
        while len(parts) < 3:
            parts.append(0)
        return tuple(parts)

    def _is_newer_version(self, latest_version, current_version):
        return self._parse_version_tuple(latest_version) > self._parse_version_tuple(current_version)

    def _versioned_exe_name(self, asset_name, tag_name):
        """生成带版本号的 exe 文件名，如 SeavoExplorer_V0p2p1.exe。"""
        base = os.path.splitext(os.path.basename(asset_name or 'SeavoExplorer.exe'))[0] or 'SeavoExplorer'
        version = str(tag_name or '').strip().lstrip('vV')
        version_part = 'V' + version.replace('.', 'p') if version else 'Vunknown'
        return f'{base}_{version_part}.exe'

    def _get_download_default_dir(self):
        downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
        if os.path.isdir(downloads):
            return downloads
        if getattr(self, 'app_dir', None) and os.path.isdir(self.app_dir):
            return self.app_dir
        return os.path.expanduser('~')

    def _fetch_latest_release(self):
        request = urllib.request.Request(
            GITHUB_LATEST_RELEASE_API,
            headers={
                'Accept': 'application/vnd.github+json',
                'User-Agent': f'SeavoExplorer/{APP_VERSION}',
            },
        )
        with urllib.request.urlopen(request, timeout=20) as response:
            return json.loads(response.read().decode('utf-8'))

    def _pick_release_exe_asset(self, release_data):
        assets = release_data.get('assets') or []
        exe_assets = [asset for asset in assets if str(asset.get('name', '')).lower().endswith('.exe')]
        if not exe_assets:
            return None
        for asset in exe_assets:
            if 'seavoexplorer' in str(asset.get('name', '')).lower():
                return asset
        return exe_assets[0]

    def _build_update_download_message(self, asset_name, downloaded, total, speed, eta, resumed, attempt):
        lines = [f'正在下载更新：{asset_name}']
        if total > 0:
            percent = downloaded * 100 / total if total else 0
            lines.append(f'{self.format_file_size(downloaded)} / {self.format_file_size(total)} ({percent:.1f}%)')
        else:
            lines.append(f'已下载 {self.format_file_size(downloaded)}')
        if speed > 0:
            lines.append(f'速度：{self.format_file_size(int(speed))}/s')
        if eta >= 0:
            lines.append(f'预计剩余：{self._format_duration(eta)}')
        if resumed:
            lines.append('状态：断点续传')
        if attempt > 1:
            lines.append(f'重试：第 {attempt} 次')
        return '\n'.join(lines)

    def _ask_open_release_page(self, title, message, release_url):
        box = QMessageBox(self)
        box.setWindowTitle(title)
        box.setIcon(QMessageBox.Question)
        box.setText(message)
        open_btn = box.addButton('打开发布页', QMessageBox.AcceptRole)
        box.addButton('取消', QMessageBox.RejectRole)
        box.setDefaultButton(open_btn)
        box.exec_()
        if box.clickedButton() == open_btn:
            self._open_url(release_url)
            return True
        return False

    def _ask_download_update(self, message):
        box = QMessageBox(self)
        box.setWindowTitle('发现新版本')
        box.setIcon(QMessageBox.Question)
        box.setText(message)
        download_btn = box.addButton('下载更新', QMessageBox.AcceptRole)
        browser_btn = box.addButton('浏览器打开', QMessageBox.ActionRole)
        box.addButton('取消', QMessageBox.RejectRole)
        box.setDefaultButton(download_btn)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == download_btn:
            return 'download'
        if clicked == browser_btn:
            return 'browser'
        return 'cancel'

    def _ask_open_download_folder(self, downloaded_path):
        box = QMessageBox(self)
        box.setWindowTitle('下载完成')
        box.setIcon(QMessageBox.Information)
        box.setText(f'更新文件已下载完成。\n\n保存位置：\n{downloaded_path}')
        open_btn = box.addButton('打开文件夹', QMessageBox.AcceptRole)
        box.addButton('关闭', QMessageBox.RejectRole)
        box.setDefaultButton(open_btn)
        box.exec_()
        return box.clickedButton() == open_btn

    def _download_update_asset(self, url, save_path, asset_name='', expected_size=0):
        progress = QProgressDialog('正在下载更新...', '取消', 0, 0, self)
        progress.setWindowTitle('下载更新')
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)

        result = {'path': None, 'error': None, 'canceled': False}
        # 清理之前的下载线程
        old_thread = getattr(self, 'update_download_thread', None)
        if old_thread is not None:
            try:
                old_thread.requestInterruption()
                old_thread.quit()
                old_thread.wait(2000)
            except Exception:
                pass
            try:
                old_thread.deleteLater()
            except Exception:
                pass
        thread = UpdateDownloadThread(url, save_path, expected_size, self)
        self.update_download_thread = thread

        def update_progress(downloaded, total, speed, eta, resumed, attempt):
            if total > 0:
                if progress.maximum() != total:
                    progress.setRange(0, total)
            else:
                if progress.maximum() != 0:
                    progress.setRange(0, 0)
            progress.setValue(downloaded if total > 0 else 0)
            progress.setLabelText(self._build_update_download_message(asset_name or os.path.basename(save_path), downloaded, total, speed, eta, resumed, attempt))

        def update_status(text):
            progress.setLabelText(text)

        def download_completed(path):
            # 校验文件大小是否匹配
            if expected_size > 0:
                actual_size = os.path.getsize(path)
                if actual_size != expected_size:
                    result['error'] = f'下载文件大小不匹配（期望 {expected_size}，实际 {actual_size}）'
                    progress.accept()
                    return
            result['path'] = path
            progress.accept()
        def download_failed(message):
            result['error'] = message
            progress.reject()

        def download_canceled(part_path, downloaded):
            result['canceled'] = True
            result['path'] = part_path
            progress.reject()

        def cancel_download():
            result['canceled'] = True
            thread.requestInterruption()

        thread.progress_changed.connect(update_progress)
        thread.status_changed.connect(update_status)
        thread.download_completed.connect(download_completed)
        thread.download_failed.connect(download_failed)
        thread.download_canceled.connect(download_canceled)
        progress.canceled.connect(cancel_download)

        try:
            thread.start()
            progress.exec_()
            if thread.isRunning():
                thread.requestInterruption()
                thread.wait(5000)
            if result['path'] and not result['canceled'] and not result['error']:
                return result['path']
            if result['canceled']:
                return None
            if result['error']:
                raise RuntimeError(result['error'])
            raise RuntimeError('更新下载失败')
        finally:
            # 先断开所有信号,避免 wait 超时后线程继续往已销毁对象发信号
            for sig_name in ('progress_changed', 'status_changed', 'download_completed', 'download_failed', 'download_canceled'):
                try:
                    getattr(thread, sig_name).disconnect()
                except (TypeError, RuntimeError):
                    pass
            if thread.isRunning():
                thread.requestInterruption()
                thread.quit()
                thread.wait(5000)
            thread.deleteLater()
            if getattr(self, 'update_download_thread', None) is thread:
                self.update_download_thread = None

    def check_for_updates(self):
        """检查 GitHub Releases 并可选下载最新 exe。"""
        self.statusBar().showMessage('正在检查更新...')
        try:
            release_data = self._fetch_latest_release()
            tag_name = release_data.get('tag_name') or release_data.get('name') or ''
            latest_version = str(tag_name).lstrip('vV')
            release_url = release_data.get('html_url') or GITHUB_RELEASES_URL
            asset = self._pick_release_exe_asset(release_data)
            if not tag_name:
                QMessageBox.warning(self, '检查更新', f'未能获取最新版本号。\n\n发布页：{release_url}')
                return
            if not self._is_newer_version(latest_version, APP_VERSION):
                QMessageBox.information(self, '检查更新', f'当前已是最新版本。\n\n当前版本：{APP_VERSION}\n发布页：{release_url}')
                return
            if not asset:
                self._ask_open_release_page(
                    '发现新版本',
                    f'发现新版本：{tag_name}\n当前版本：{APP_VERSION}\n\n该版本未提供 exe 安装包。\n发布页：{release_url}',
                    release_url,
                )
                return

            asset_name = asset.get('name') or 'SeavoExplorer.exe'
            asset_size = int(asset.get('size') or 0)
            size_mb = asset_size / 1024 / 1024
            asset_line = f'更新文件：{asset_name}'
            if size_mb > 0:
                asset_line += f'（{size_mb:.1f} MB）'
            msg = (
                f'发现新版本：{tag_name}\n'
                f'当前版本：{APP_VERSION}\n'
                f'{asset_line}\n'
                f'发布页：{release_url}\n\n'
                '是否下载更新文件？\n'
                '如果下载失败，可以改用浏览器打开发布页下载。'
            )
            action = self._ask_download_update(msg)
            if action == 'browser':
                self._open_url(release_url)
                return
            if action != 'download':
                return

            default_name = self._versioned_exe_name(asset_name, tag_name)
            default_path = os.path.join(self._get_download_default_dir(), default_name)
            save_path, _ = QFileDialog.getSaveFileName(self, '保存更新文件', default_path, '可执行文件 (*.exe)')
            if not save_path:
                return
            if not save_path.lower().endswith('.exe'):
                save_path += '.exe'

            downloaded_path = self._download_update_asset(asset.get('browser_download_url'), save_path, asset_name, asset_size)
            if not downloaded_path:
                return
            self.statusBar().showMessage(f'更新已下载: {downloaded_path}')
            if self._ask_open_download_folder(downloaded_path):
                self._open_with_shell(os.path.dirname(downloaded_path))
        except urllib.error.HTTPError as e:
            self._ask_open_release_page(
                '检查更新失败',
                f'GitHub 返回错误：HTTP {e.code}\n\n发布页：{GITHUB_RELEASES_URL}',
                GITHUB_RELEASES_URL,
            )
        except urllib.error.URLError as e:
            self._ask_open_release_page(
                '检查更新失败',
                f'无法连接 GitHub。\n原因：{e.reason}\n\n发布页：{GITHUB_RELEASES_URL}',
                GITHUB_RELEASES_URL,
            )
        except RuntimeError as e:
            message = str(e)
            if message:
                self._ask_open_release_page(
                    '下载失败',
                    f'{message}\n\n发布页：{GITHUB_RELEASES_URL}\n\n可改用浏览器下载。',
                    GITHUB_RELEASES_URL,
                )
        except Exception as e:
            QMessageBox.warning(self, '检查更新失败', f'检查更新时发生错误：\n{str(e)}')
        finally:
            self.statusBar().clearMessage()


    def show_about(self):
        about_text = (
            '<h3>SeavoExplorer - 主板项目文件浏览器</h3>'
            f'<p>版本 {APP_VERSION}</p>'
            '<p>本版本新增文件版本管理（保存版本）、归档到old文件夹；'
            '修复F5刷新文件树不更新、右键重命名失效等bug；'
            '视频预览支持 5 帧截图(10%/30%/50%/70%/90%)并可在查看器中逐帧切换；'
            '预览大图支持滚轮缩放、按钮缩放、鼠标拖拽平移；视频预览默认关闭(需手动开启)；'
            '修复文件搜索结果分组与双击定位、快捷访问双击打开资源管理器、窗口最大化记忆等。</p>'
            f'<p>GitHub：<a href="{GITHUB_REPO_URL}">{GITHUB_REPO_URL}</a></p>'
        )
        # 关于页 logo 优先用高清 PNG 源（清晰放大），回退到多尺寸 ico
        png_path = self._resource_path('favicon_src.png')
        ico_path = self._resource_path('favicon.ico')
        pixmap = None
        if os.path.exists(png_path):
            pixmap = QPixmap(png_path)
        elif os.path.exists(ico_path):
            pixmap = QIcon(ico_path).pixmap(128, 128)

        if pixmap is not None and not pixmap.isNull():
            scaled = pixmap.scaled(128, 128, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            msg = QMessageBox(self)
            msg.setWindowTitle('关于')
            msg.setTextFormat(Qt.RichText)
            msg.setTextInteractionFlags(Qt.TextBrowserInteraction)
            msg.setText(about_text)
            msg.setIconPixmap(scaled)
            msg.exec_()
        else:
            QMessageBox.about(self, '关于', about_text)
    
    def show_wizard(self):
        """打开新手向导，并标记为已显示。"""
        dialog = WizardDialog(self)
        dialog.exec_()
        if not getattr(self, 'wizard_shown', False):
            self.wizard_shown = True
            self.save_settings_to_file(self.settings, self.include_subfolders)

    def show_help(self):
        help_dialog = QDialog(self)
        help_dialog.setWindowTitle('使用帮助')
        help_dialog.setGeometry(200, 100, 700, 600)
        layout = QVBoxLayout(help_dialog)
        
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setHtml(f'''
<h2 style="color: #2c3e50;">SeavoExplorer 使用帮助</h2>
<p style="color: #7f8c8d;">主板/子卡项目文件浏览器 —— 快速定位项目、预览工程文档、整理版本目录。当前版本：<b>{APP_VERSION}</b>。
首次使用建议先看 <b>帮助 → 新手向导</b>，再按本页完成路径、预览和压缩相关设置。</p>

<h3 style="color: #2980b9;">一、首次使用流程</h3>
<ol>
<li>打开 <b>设置 → 项目文件夹设置</b>，添加一个或多个项目根目录。</li>
<li>按需要勾选 <b>包含子文件夹</b>、<b>按编号排序</b>，点击确定后程序会自动刷新项目列表。</li>
<li>在左侧搜索或选择项目，右侧文件树会显示该项目内容。</li>
<li>单击文件查看预览与元数据，双击文件则用系统默认程序打开。</li>
<li>如需预览/解压 <code>.rar</code>、<code>.7z</code>，请在 <b>设置 → 7-Zip路径设置</b> 中确认 7z.exe 可用。</li>
</ol>

<h3 style="color: #2980b9;">二、检查更新与下载更新</h3>
<ul>
<li>点击 <b>帮助 → 检查更新</b>，程序会读取 GitHub Releases 上的最新版本并与当前版本比较。</li>
<li>发现新版本时，弹窗会显示版本号、更新文件大小和发布页链接，可选择 <b>下载更新</b> 或 <b>浏览器打开</b>。</li>
<li>程序内下载会在后台进行，进度窗口显示下载量、速度和预计剩余时间；网络中断时会自动重试，已有临时文件时会尽量断点续传。</li>
<li>取消时会保留 <code>.part</code> 临时文件以便稍后续传；下载最终失败后会清理临时文件，避免长期占用磁盘。</li>
<li>如果 GitHub 连接较慢或下载失败，可使用弹窗中的发布页链接，在浏览器中手动下载最新 <code>SeavoExplorer.exe</code>。</li>
</ul>

<h3 style="color: #2980b9;">三、项目文件夹管理</h3>
<p><b>1. 配置项目根目录</b></p>
<p>点击菜单 <b>设置 → 项目文件夹设置</b>，添加包含项目文件夹的根目录。程序会扫描这些目录下符合命名规则的项目文件夹。</p>
<p>命名规则：以 <b>S</b>（主板）或 <b>M</b>（子卡）开头，后跟 <b>3~4 位数字</b>，可选 <code>_注释</code> 后缀。例如：<code>S001</code>、<code>M1234</code>、<code>S002_样机</code>、<code>M003_说明</code>。</p>
<ul>
<li><b>包含子文件夹</b>：递归扫描根目录下面的子目录，适合项目按客户/年份分层存放的场景</li>
<li><b>按编号排序</b>：忽略来源目录分组，所有项目统一按编号大小排序；不勾选时先按根目录添加顺序分组，组内再按编号排序</li>
<li><b>刷新</b>：按 <b>F5</b> 或重新打开设置后确认，可重新扫描项目与文件树</li>
</ul>

<p><b>2. 项目列表操作</b></p>
<ul>
<li><b>单击</b>项目行：在右侧文件树中显示该项目的文件</li>
<li><b>双击编号列</b>：在系统资源管理器中打开该项目文件夹</li>
<li><b>双击注释列</b>：编辑项目注释，注释会自动保存到 <code>seavo_comments.json</code></li>
<li><b>右键项目行</b>：置顶 / 取消置顶、在终端中打开、隐藏项目；置顶项会加粗并排在列表前部</li>
<li><b>隐藏项目</b>：不常用或已归档项目可隐藏；如需找回，请使用 <b>设置 → 恢复已隐藏项目</b></li>
<li><b>文件夹搜索框</b>：输入编号、注释或路径关键词实时过滤项目列表</li>
</ul>
<p style="color: #7f8c8d;">注释显示优先级：若 <code>seavo_comments.json</code> 中对该文件夹有注释则优先显示，否则使用文件夹名的 <code>_注释</code> 后缀。</p>

<h3 style="color: #2980b9;">四、文件浏览与操作</h3>
<p><b>1. 文件浏览</b></p>
<ul>
<li>单击文件：在下方 <b>文件预览</b> 区显示内容；<b>元数据</b> 标签页显示大小、创建/修改时间等详情</li>
<li>双击文件：用系统默认程序打开</li>
<li>双击文件夹：在资源管理器中打开</li>
<li>双击空白区域：在资源管理器中打开当前项目文件夹</li>
</ul>

<p><b>2. 多选操作</b></p>
<p>文件树支持多选：按住 <b>Ctrl</b> 点选多个项目，或按住 <b>Shift</b> 选择连续范围。多选后可批量复制、删除、压缩。</p>

<p><b>3. 右键菜单</b></p>
<p>选中<b>单个</b>文件/文件夹时：</p>
<ul>
<li><b>复制</b>：复制到剪贴板，既可在资源管理器中粘贴，也可用程序内"粘贴副本"</li>
<li><b>粘贴副本</b>：把剪贴板中的内容复制到当前位置，自动处理重名（追加 <code>_副本N</code>）</li>
<li><b>重命名</b>：重命名文件或文件夹（也可按 <b>F2</b>）</li>
<li><b>保存版本</b>：为文件生成日期版本副本（如 <code>S1200-10_20260708.dsn</code>），当天多次保存自动递增字母后缀（a/b/c...）</li>
<li><b>归档到old文件夹</b>：将文件移入同目录下的 <code>old/</code> 文件夹（自动创建），支持多选</li>
<li><b>显示隐藏文件</b>：菜单 <b>设置 → 显示隐藏文件</b>（可勾选开关），勾选后在文件树中显示以 <code>.</code> 开头的文件和系统隐藏属性的文件</li>
<li><b>添加到zip压缩包</b>：压缩为同名 <code>.zip</code> 文件</li>
<li><b>智能解压</b>：仅对 <code>.zip</code>、<code>.rar</code>、<code>.7z</code> 显示</li>
<li><b>移入回收站</b>：移入系统回收站，避免直接永久删除</li>
<li><b>在终端中打开</b>：仅文件夹显示；优先用 Windows Terminal 打开，失败后回退到 PowerShell / cmd，并始终定位到所选路径</li>
</ul>
<p>选中<b>多个</b>项目时，菜单仅保留可批量执行的项：<b>复制</b>、<b>添加到zip压缩包</b>、<b>归档到old文件夹</b>、<b>移入回收站</b>。</p>
<p>在<b>空白处</b>右键：仅显示<b>粘贴副本</b>，粘贴到当前项目文件夹。</p>

<p><b>4. 复制与粘贴的目标规则</b></p>
<ul>
<li>复制（单个或多个）后，“粘贴副本”会把<b>全部</b>已复制项粘到目标位置</li>
<li>粘贴时若<b>选中了一个文件夹</b>，粘贴到该文件夹内；若选中的是文件，则粘到其所在目录</li>
<li>若<b>选中了多个</b>项目，则粘贴到当前项目根文件夹</li>
<li>粘贴遇到重名文件或文件夹时，会自动追加副本序号，避免覆盖原文件</li>
</ul>

<h3 style="color: #2980b9;">四、文件预览</h3>
<p>单击文件树中的文件，下方预览区会自动显示内容：</p>
<table border="1" cellpadding="4" cellspacing="0" style="border-collapse: collapse;">
<tr style="background: #ecf0f1;"><th>文件类型</th><th>支持格式</th><th>说明</th></tr>
<tr><td>文本文件</td><td>.txt .csv .log .bom .drc .rep .rpt .md .json .xml .html .htm .ini .cfg</td><td>直接显示文本（UTF-8/GBK 自动识别）</td></tr>
<tr><td>PDF 文件</td><td>.pdf</td><td>多页预览，可翻页查看</td></tr>
<tr><td>Excel 文件</td><td>.xlsx .xlsm .xls</td><td>表格形式预览</td></tr>
<tr><td>Word 文件</td><td>.docx .doc</td><td>文档内容预览</td></tr>
<tr><td>图片文件</td><td>.jpg .jpeg .png .bmp .gif .tiff .tif .webp .svg</td><td>缩略图预览，点击查看大图</td></tr>
<tr><td>视频文件</td><td>.mp4 .avi .mov .mkv .flv .wmv .m4v .webm .mpg .mpeg .3gp</td><td>5 帧截图预览(10%/30%/50%/70%/90%)，点击放大后可逐帧切换(方向键/按钮)</td></tr>
<tr><td>压缩包</td><td>.zip .rar .7z</td><td>树状结构显示内容</td></tr>
</table>
<p style="color: #7f8c8d;">加密工程文件（.opj .dsn .sch .brd .dbk .dsnlck）无法预览，会显示提示信息而非二进制内容；请用对应 EDA 软件打开。</p>
<p><b>预览开关（降低卡顿）：</b>菜单 <b>设置 → 预览设置</b> 可按类型分别开关自动预览（文本/PDF/图片/视频/压缩包/表格/文档）。关闭某类后，点击该类文件不会立即读取，预览区改为显示一个 <b>显示预览</b> 按钮，需手动点击才加载——适合大文件或慢速磁盘。各类开关会被记住。</p>

<h3 style="color: #2980b9;">五、压缩包操作</h3>
<p><b>1. 智能解压</b></p>
<p>右键压缩包选择“智能解压”，程序自动判断：</p>
<ul>
<li>包内只有一个顶层项目 → 直接解压到当前目录</li>
<li>包内有多个顶层项目 → 创建与压缩包同名的文件夹，解压到其中，避免文件散落</li>
</ul>
<p><code>.zip</code> 使用内置解压；<code>.rar</code> / <code>.7z</code> 需要 7-Zip 支持。</p>

<p><b>2. 添加到 zip 压缩包</b></p>
<ul>
<li>选中单个项目：在同目录下生成同名 <code>.zip</code>，重名时自动追加序号</li>
<li>选中多个项目：一并打包为一个 <code>.zip</code></li>
<li>压缩文件会保留原有目录层级，便于发给他人后直接解压使用</li>
</ul>

<p><b>3. 7-Zip 路径设置</b></p>
<p>菜单 <b>设置 → 7-Zip路径设置</b> 可手动指定 7z.exe。程序按以下顺序自动查找：</p>
<ol>
<li>设置中手动指定的路径</li>
<li>程序所在目录下的 7z.exe</li>
<li>C:\\Program Files\\7-Zip\\7z.exe</li>
<li>C:\\Program Files (x86)\\7-Zip\\7z.exe</li>
</ol>

<h3 style="color: #2980b9;">六、快捷访问栏</h3>
<p>菜单栏下方的快捷访问栏提供常用文件夹的快速入口：</p>
<ul>
<li><b>普通按钮</b>（默认样式）：点击后在文件树中显示该文件夹内容</li>
<li><b>不显示预览按钮</b>（灰色斜体）：点击后直接在资源管理器中打开</li>
<li><b>+</b> 按钮：位于“快捷访问”标题右侧，点击可打开快捷访问设置</li>
<li><b>右键快捷访问按钮</b>：可在终端中打开该目录、删除该快捷访问（需二次确认）、或打开快捷访问设置</li>
</ul>
<p>菜单 <b>设置 → 快捷访问设置</b> 可添加、删除、排序快捷项，并为每项设置名称、路径和是否不显示预览。</p>
<p style="color: #7f8c8d;">提示：磁盘根目录、网络文件夹等大目录建议设为“不显示预览”，避免文件树加载缓慢。</p>

<h3 style="color: #2980b9;">七、界面与导航</h3>
<ul>
<li><b>面包屑路径栏</b>：文件树上方显示从项目根到当前点选项的路径（如 <code>S1234 › V01 › BOM</code>）。点击中间任意一段，可在文件树中快速选中并定位到该文件夹；路径过长时中间会自动省略。</li>
<li><b>状态栏文件统计</b>：选中项目后，状态栏右侧常驻显示该项目递归的<b>文件数与总大小</b>，在后台计算不卡界面；切换项目会自动更新。</li>
<li><b>记住窗口与项目</b>：退出时记住窗口大小/位置、左右分栏宽度、是否最大化，以及上次打开的项目；下次启动自动恢复。若上次项目已被删除/改名/隐藏，则安全跳过不报错。</li>
<li><b>全屏已禁用</b>：本程序不支持全屏模式（避免菜单与关闭按钮不可见），按 F11 等不会进入全屏。</li>
</ul>

<h3 style="color: #2980b9;">八、视频预览与大图查看</h3>
<p><b>1. 多帧视频预览</b></p>
<p>视频预览默认<b>关闭</b>，需在 <b>设置 → 预览设置</b> 中手动开启(视频类)。开启后点击视频文件，会显示 5 张截图,分别对应视频 10%、30%、50%、70%、90% 位置。</p>
<p><b>2. 视频帧查看器</b></p>
<p>点击视频预览图可打开帧查看器,左右键或 ← → ↑ ↓ 方向键切换帧,也可点击底部「◀ 上一张」「下一张 ▶」按钮。顶部标题栏显示当前帧位置(如 [2/5] 30%)。</p>
<p><b>3. 大图缩放与拖拽</b></p>
<p>图片/视频帧大图查看器支持:</p>
<ul>
<li><b>滚轮</b>：向上放大、向下缩小</li>
<li><b>按钮</b>：底部「−」「+」「1:1」分别控制缩小、放大、重置</li>
<li><b>拖拽</b>：左键按住拖动平移图片(图片大于视口时)</li>
</ul>

<h3 style="color: #2980b9;">九、新建项目与版本结构</h3>
<p><b>1. 新建项目文件夹</b></p>
<p>点击左侧 <b>新建项目文件夹</b>，选择类型（S/M）、输入编号和保存位置，程序自动创建符合命名规则的项目文件夹。可选注释会作为文件夹名后缀保存，便于后续识别。</p>
<p><b>2. 新建文件夹内部结构</b></p>
<p>选中一个项目后，点击 <b>新建文件夹内部结构</b>，可创建版本文件夹（如 <code>V01</code>）及标准子文件夹（BOM、SCH、物料、评审、信号测试），也可自定义子文件夹。上次选择的模板会被记住。</p>
<p style="color: #7f8c8d;">建议同一项目内按版本目录归档资料，例如 <code>V01</code>、<code>V02</code>，减少不同阶段文件混放。</p>

<h3 style="color: #2980b9;">十、快捷键</h3>
<table border="1" cellpadding="4" cellspacing="0" style="border-collapse: collapse;">
<tr style="background: #ecf0f1;"><th>快捷键</th><th>功能</th></tr>
<tr><td>F5</td><td>刷新项目列表和文件树</td></tr>
<tr><td>F2</td><td>重命名文件树中选中的单个文件/文件夹</td></tr>
<tr><td>Ctrl+C</td><td>复制选中的文件/文件夹（支持多选）</td></tr>
<tr><td>Ctrl+V</td><td>粘贴副本到选中文件夹或当前项目</td></tr>
<tr><td>Delete</td><td>将选中的文件/文件夹移入回收站（支持多选）</td></tr>
<tr><td>← ↑</td><td>视频帧查看器:切换到上一帧</td></tr>
<tr><td>→ ↓</td><td>视频帧查看器:切换到下一帧</td></tr>
</table>

<h3 style="color: #2980b9;">十一、配置文件与数据保存</h3>
<ul>
<li><b>应用设置</b>：项目路径、排序选项、快捷访问、7-Zip 路径、预览开关、窗口大小/位置/分栏、上次打开的项目等保存到 <code>seavoexplorer.json</code></li>
<li><b>项目注释</b>：手动编辑的注释保存到 <code>seavo_comments.json</code></li>
<li><b>保存位置</b>：开发运行时保存在脚本所在目录；打包为 exe 后保存在 exe 所在目录</li>
<li><b>隐藏属性</b>：在 Windows 下配置文件会尽量设置为隐藏，避免误删</li>
</ul>

<h3 style="color: #2980b9;">十二、常见问题</h3>
<ul>
<li><b>找不到项目</b>：检查根目录是否添加正确、项目文件夹是否符合 S/M + 3~4 位数字规则；若项目在更深层目录，请勾选“包含子文件夹”</li>
<li><b>列表顺序不符合预期</b>：检查是否启用了“按编号排序”；关闭后会按根目录添加顺序分组显示</li>
<li><b>.rar / .7z 无法预览或解压</b>：安装 7-Zip，或在 <b>设置 → 7-Zip路径设置</b> 中手动指定 7z.exe</li>
<li><b>大目录打开慢</b>：把该路径加入快捷访问并设置为“不显示预览”，或避免把磁盘根目录作为普通预览入口</li>
<li><b>不小心删除文件</b>：程序会移入系统回收站，可点击状态栏右侧回收站按钮打开并恢复</li>
<li><b>预览内容乱码</b>：文本预览会尝试 UTF-8 / GBK；若仍乱码，请用专业编辑器或对应软件打开原文件</li>
</ul>
<h3 style="color: #2980b9;">十三、关于更新</h3>
<p>如果你使用“帮助 → 检查更新”，程序会优先尝试在应用内下载更新；若网络较慢或下载失败，可直接打开 GitHub Releases 页面用浏览器下载。</p>
'''
        )

        layout.addWidget(help_text)
        
        close_btn = QPushButton('关闭')
        close_btn.clicked.connect(help_dialog.close)
        layout.addWidget(close_btn)
        
        help_dialog.exec_()
    
    def show_full_image(self, event):
        """点击图片或视频缩略图时显示大图。图片单张;视频多帧可上下帧切换。"""
        if not hasattr(self, 'current_image_path') or not self.current_image_path:
            return
        try:
            ext = os.path.splitext(self.current_image_path)[1].lower()
            if ext in IMAGE_EXTS:
                self._show_single_image(self.current_image_path)
            elif ext in VIDEO_EXTS:
                self._show_video_frames(self.current_image_path)
        except Exception as e:
            QMessageBox.warning(self, '警告', f'无法显示大图: {str(e)}')

    def _show_single_image(self, path):
        """图片大图预览(单张,支持缩放)。"""
        dialog = QDialog(self)
        dialog.setAttribute(Qt.WA_DeleteOnClose)
        dialog.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMaximizeButtonHint)
        dialog.setWindowTitle(f'图片查看 - {os.path.basename(path)}')
        image = self._load_image_safe(path)
        if image.isNull():
            return
        image_label = ZoomableImageLabel()
        image_label.setPixmap(QPixmap.fromImage(image))
        scroll_area = QScrollArea(dialog)
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(image_label)
        # 底部缩放控制栏
        zoom_bar = self._make_zoom_bar(image_label, dialog)
        layout = QVBoxLayout(dialog)
        layout.addWidget(scroll_area, 1)
        layout.addLayout(zoom_bar)
        layout.setContentsMargins(4, 4, 4, 4)
        screen = QApplication.desktop().screenGeometry()
        dialog.resize(int(screen.width() * 0.85), int(screen.height() * 0.85))
        dialog.exec_()

    def _show_video_frames(self, path):
        """视频多帧预览(上一张/下一张 + 缩放)。"""
        frames = self._capture_video_frames(path, target_height=480)
        if not frames:
            QMessageBox.information(self, '提示', '无法提取视频帧,请确认已安装 OpenCV (pip install opencv-python)')
            return

        dialog = QDialog(self)
        dialog.setAttribute(Qt.WA_DeleteOnClose)
        dialog.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMaximizeButtonHint)

        image_label = ZoomableImageLabel()
        scroll_area = QScrollArea(dialog)
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(image_label)

        # 底部控制栏:上一张/进度/下一张 + 缩放
        nav_layout = QHBoxLayout()
        prev_btn = QPushButton('◀ 上一张')
        next_btn = QPushButton('下一张 ▶')
        page_label = QLabel()
        page_label.setAlignment(Qt.AlignCenter)
        nav_layout.addWidget(prev_btn)
        nav_layout.addWidget(page_label, 1)
        nav_layout.addWidget(next_btn)

        zoom_bar = self._make_zoom_bar(image_label, dialog)

        layout = QVBoxLayout(dialog)
        layout.addWidget(scroll_area, 1)
        layout.addLayout(nav_layout)
        layout.addLayout(zoom_bar)
        layout.setContentsMargins(4, 4, 4, 4)

        state = {'idx': 0, 'frames': frames}

        def show_frame(idx):
            state['idx'] = idx
            image_label.setPixmap(frames[idx])  # 重置缩放并显示新帧
            pos_pct = int(VIDEO_PREVIEW_POSITIONS[idx] * 100)
            dialog.setWindowTitle(f'视频帧预览 - {os.path.basename(path)}  [{idx+1}/{len(frames)}]  {pos_pct}%')
            page_label.setText(f'{idx+1} / {len(frames)}  ({pos_pct}%)')
            prev_btn.setEnabled(idx > 0)
            next_btn.setEnabled(idx < len(frames) - 1)

        prev_btn.clicked.connect(lambda: show_frame(max(0, state['idx'] - 1)))
        next_btn.clicked.connect(lambda: show_frame(min(len(frames) - 1, state['idx'] + 1)))

        # 快捷键:左/上=上一张, 右/下=下一张(QShortcut 无需焦点)
        def go_prev():
            if state['idx'] > 0:
                show_frame(state['idx'] - 1)
        def go_next():
            if state['idx'] < len(frames) - 1:
                show_frame(state['idx'] + 1)
        QShortcut(Qt.Key_Left, dialog, go_prev, context=Qt.WindowShortcut)
        QShortcut(Qt.Key_Up, dialog, go_prev, context=Qt.WindowShortcut)
        QShortcut(Qt.Key_Right, dialog, go_next, context=Qt.WindowShortcut)
        QShortcut(Qt.Key_Down, dialog, go_next, context=Qt.WindowShortcut)

        show_frame(0)
        screen = QApplication.desktop().screenGeometry()
        dialog.resize(int(screen.width() * 0.85), int(screen.height() * 0.85))
        dialog.exec_()

    def _make_zoom_bar(self, image_label, dialog):
        """创建底部缩放控制栏:放大/缩小/重置 + 缩放比例显示。"""
        from PyQt5.QtWidgets import QHBoxLayout, QPushButton, QLabel
        layout = QHBoxLayout()
        zoom_out_btn = QPushButton('−')  # 缩小
        zoom_out_btn.setFixedSize(32, 24)
        zoom_out_btn.setToolTip('缩小 (或滚轮向下)')
        zoom_in_btn = QPushButton('+')  # 放大
        zoom_in_btn.setFixedSize(32, 24)
        zoom_in_btn.setToolTip('放大 (或滚轮向上)')
        zoom_reset_btn = QPushButton('1:1')
        zoom_reset_btn.setFixedSize(40, 24)
        zoom_reset_btn.setToolTip('重置为原始大小')
        zoom_label = QLabel('100%')
        zoom_label.setAlignment(Qt.AlignCenter)
        zoom_label.setFixedWidth(50)

        def refresh_label():
            zoom_label.setText(image_label.get_zoom_text())

        # 按钮触发时刷新
        zoom_out_btn.clicked.connect(lambda: (image_label.zoom_out(), refresh_label()))
        zoom_in_btn.clicked.connect(lambda: (image_label.zoom_in(), refresh_label()))
        zoom_reset_btn.clicked.connect(lambda: (image_label.zoom_reset(), refresh_label()))
        # 滚轮/任意缩放变化时也刷新(通过回调)
        image_label.set_zoom_callback(refresh_label)

        layout.addStretch()
        layout.addWidget(zoom_out_btn)
        layout.addWidget(zoom_label)
        layout.addWidget(zoom_in_btn)
        layout.addWidget(zoom_reset_btn)
        layout.addStretch()
        return layout

    def _capture_video_frames(self, path, target_height=480):
        """在 VIDEO_PREVIEW_POSITIONS 各时间点截图,返回等比缩放后的 QPixmap 列表。"""
        if not HAS_OPENCV:
            return None
        frames = []
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            cap.release()
            return None
        try:
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if total_frames <= 0:
                return None
            for pos in VIDEO_PREVIEW_POSITIONS:
                frame_no = max(0, min(int(total_frames * pos), total_frames - 1))
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_no)
                ret, frame = cap.read()
                if not ret:
                    continue
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = frame_rgb.shape
                # 等比缩放到 target_height
                scale = target_height / h
                new_w = max(1, int(round(w * scale)))
                resized = cv2.resize(frame_rgb, (new_w, target_height), interpolation=cv2.INTER_AREA)
                q_img = QImage(resized.data, resized.shape[1], resized.shape[0], resized.shape[1] * ch, QImage.Format_RGB888).copy()
                frames.append(QPixmap.fromImage(q_img))
        finally:
            try: cap.release()
            except Exception: pass
        return frames if frames else None
    
    def generate_video_thumbnails(self, video_path=None, thumb_h=96):
        """在 VIDEO_PREVIEW_POSITIONS 各时间点截图,返回等比缩放后的 QPixmap 列表。
        可传入 video_path,否则使用 self.current_video_path。"""
        if video_path is None:
            video_path = getattr(self, 'current_video_path', None)
        return self._capture_video_frames(video_path, target_height=thumb_h)

    def generate_video_thumbnail(self, video_path, size=(320, 240)):
        """保留:单个视频缩略图(中点帧),向后兼容。"""
        if not HAS_OPENCV:
            return None
        try:
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                return None
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            frame_to_capture = total_frames // 2
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_to_capture)
            ret, frame = cap.read()
            cap.release()
            if not ret:
                cap = cv2.VideoCapture(video_path)
                ret, frame = cap.read()
                cap.release()
                if not ret:
                    return None
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            height, width, channel = frame_rgb.shape
            q_image = QImage(frame_rgb.data, width, height, channel * width, QImage.Format_RGB888)
            return q_image.copy()
        except Exception:
            return None
    
    def keyPressEvent(self, event):
        # 输入控件（搜索框、编辑框）获得焦点时，不拦截 Ctrl+C/V/Delete 等编辑快捷键
        focus = self.focusWidget()
        if isinstance(focus, (QLineEdit, QTextEdit)):
            super().keyPressEvent(event)
            return
        handled = True
        if event.key() == Qt.Key_F11:
            # 双保险：拦掉 F11，配合 changeEvent 阻全屏
            event.accept()
        elif event.key() == Qt.Key_F5:
            self.refresh_all()
        elif event.key() == Qt.Key_F2:
            selected_paths = self._get_selected_file_paths()
            if len(selected_paths) == 1:
                self.rename_item(selected_paths[0])
            elif len(selected_paths) > 1:
                self.statusBar().showMessage("请选择单个文件或文件夹进行重命名")
            else:
                selected_path = self._get_primary_selected_file_path()
                if selected_path:
                    self.rename_item(selected_path)
        elif event.key() == Qt.Key_C and event.modifiers() & Qt.ControlModifier:
            self.copy_selected_items()
        elif event.key() == Qt.Key_V and event.modifiers() & Qt.ControlModifier:
            self.paste_to_selected_target()
        elif event.key() == Qt.Key_Delete:
            selected_paths = self._get_selected_file_paths()
            if len(selected_paths) > 1:
                self._move_paths_to_recycle(selected_paths)
            else:
                selected_path = self._get_primary_selected_file_path()
                if selected_path:
                    self.move_to_recycle(selected_path)
        else:
            handled = False
        # 统一调用 super() 确保事件正确传播(无论是否被本窗口处理)
        super().keyPressEvent(event)
        if not handled:
            event.ignore()

if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        app.setAttribute(Qt.AA_DisableWindowContextHelpButton)

        # 设置全局界面字体：优先使用系统中可用的清晰字体
        apply_app_font(app)

        # 创建启动画面（使用预渲染缓存）
        splash = QSplashScreen(get_splash_pixmap())
        splash.setStyleSheet("""
            QSplashScreen {
                border: 2px solid #E94A16;
                border-radius: 10px;
                background-color: black;
            }
        """)
        splash.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint)
        splash.showMessage("正在初始化程序...", Qt.AlignBottom | Qt.AlignCenter, Qt.white)
        splash.show()

        # 显示启动画面
        app.processEvents()
        
        # 初始化主窗口（在后台进行）
        window = MainWindow()

        # 显示主窗口
        window.show()

        # 确保窗口获得焦点并显示在最前面
        window.raise_()
        window.activateWindow()
        if window.windowState() & Qt.WindowMinimized:
            window.setWindowState(window.windowState() & ~Qt.WindowMinimized | Qt.WindowActive)

        # 关闭启动画面
        splash.finish(window)
        
        sys.exit(app.exec_())
    except Exception as e:
        error_msg = "程序异常退出: {}\n\n详细信息:\n{}".format(str(e), traceback.format_exc())
        print(error_msg)
        try:
            with open(os.path.join(_get_app_dir(), 'error_details.log'), 'w', encoding='utf-8') as f:
                f.write(error_msg)
        except Exception:
            pass
        sys.exit(1)
